# -*- coding: utf-8 -*-
"""批量导入、导入模板管理、业绩报表模板管理路由"""
import os
import re
import json
from datetime import datetime
from flask import Blueprint, request, redirect, url_for, flash, render_template, jsonify
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from models import db, User, Order, Category, ImportTemplate, PerformanceReportTemplate, _now_bj
from helpers import role_required, get_unread_count, notify_users

bp = Blueprint('import_routes', __name__)

# 引入模板管理中的共享变量
from routes_templates import UPLOAD_FOLDER, allowed_file


# ============ 列数据预处理函数 ============

def apply_column_preprocessing(value, preprocessing_config):
    """
    对列值应用预处理配置
    
    Args:
        value: 原始值
        preprocessing_config: 预处理配置，支持以下类型：
            - regex_replace: {"type":"regex_replace","pattern":"正则","replacement":"替换"}
            - prefix_remove: {"type":"prefix_remove","prefix":"要去掉的前缀"}
            - suffix_remove: {"type":"suffix_remove","suffix":"要去掉的后缀"}
            - trim: {"type":"trim"}（可省略其他参数）
            - strip_chars: {"type":"strip_chars","chars":"要去掉的字符集"}
            - replace: {"type":"replace","old":"旧字符串","new":"新字符串"}
            - upper: {"type":"upper"}
            - lower: {"type":"lower"}
    
    Returns:
        处理后的值
    """
    if not value or not preprocessing_config:
        return value
    
    value = str(value) if value is not None else ''
    ptype = preprocessing_config.get('type', '')
    
    try:
        if ptype == 'regex_replace':
            pattern = preprocessing_config.get('pattern', '')
            replacement = preprocessing_config.get('replacement', '')
            if pattern:
                value = re.sub(pattern, replacement, value)
        
        elif ptype == 'prefix_remove':
            prefix = preprocessing_config.get('prefix', '')
            if prefix and value.startswith(prefix):
                value = value[len(prefix):]
        
        elif ptype == 'suffix_remove':
            suffix = preprocessing_config.get('suffix', '')
            if suffix and value.endswith(suffix):
                value = value[:-len(suffix)]
        
        elif ptype == 'strip_chars':
            chars = preprocessing_config.get('chars', '')
            if chars:
                value = value.strip(chars)
        
        elif ptype == 'replace':
            old_str = preprocessing_config.get('old', '')
            new_str = preprocessing_config.get('new', '')
            if old_str:
                value = value.replace(old_str, new_str)
        
        elif ptype == 'upper':
            value = value.upper()
        
        elif ptype == 'lower':
            value = value.lower()
        
        elif ptype == 'trim':
            value = value.strip()
    
    except Exception as e:
        print(f"[预处理] 应用预处理失败: {e}")
    
    return value


# ============ 批量导入快递信息 ============
@bp.route('/batch_import')
@role_required('shipper', 'admin')
def batch_import_page():
    """批量导入页面"""
    categories = Category.query.filter_by(is_active=True).order_by(Category.sort_order.asc()).all()
    import_templates = ImportTemplate.query.all()
    template_map = {t.category_id: t for t in import_templates}
    return render_template('batch_import.html', categories=categories,
                           import_templates=import_templates, template_map=template_map,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/admin/import_template/upload', methods=['POST'])
@role_required('admin')
def upload_import_template():
    """上传导入模板（支持从现有模板复制）"""
    category_id = request.form.get('category_id', type=int)
    file = request.files.get('file')
    copy_from_template_id = request.form.get('copy_from_template_id', type=int)

    if not category_id:
        flash('请选择产品类别！', 'danger')
        return redirect(url_for('templates.admin_templates'))

    # 收集字段映射
    field_mapping = {}
    idx = 0
    while True:
        excel_col = request.form.get(f'excel_col_{idx}', '').strip()
        order_field = request.form.get(f'order_field_{idx}', '').strip()
        if not excel_col and not order_field:
            break
        if excel_col and order_field:
            field_mapping[excel_col] = order_field
        idx += 1

    # 获取默认快递种类
    default_express_type = request.form.get('default_express_type', '').strip()

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    filename = None
    filepath = None

    # 处理文件：优先使用复制的文件，否则上传新文件
    if copy_from_template_id:
        # 从现有模板复制
        source_template = ImportTemplate.query.get(copy_from_template_id)
        if source_template and os.path.exists(source_template.filepath):
            filename = source_template.filename
            # 创建新文件名（添加时间戳）
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            saved_name = f"import_{category_id}_{timestamp}_{filename}"
            filepath = os.path.join(UPLOAD_FOLDER, saved_name)
            # 复制文件
            import shutil
            shutil.copy2(source_template.filepath, filepath)
    elif file and file.filename:
        # 上传新文件
        if not allowed_file(file.filename):
            flash('只允许上传Excel文件（.xlsx, .xls）！', 'danger')
            return redirect(url_for('templates.admin_templates'))
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        saved_name = f"import_{category_id}_{timestamp}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, saved_name)
        file.save(filepath)
    else:
        flash('请选择文件或从现有模板复制！', 'danger')
        return redirect(url_for('templates.admin_templates'))

    existing = ImportTemplate.query.filter_by(category_id=category_id).first()
    if existing:
        if os.path.exists(existing.filepath):
            os.remove(existing.filepath)
        existing.filename = filename
        existing.filepath = filepath
        existing.field_mapping = json.dumps(field_mapping, ensure_ascii=False)
        existing.default_express_type = default_express_type
        existing.create_time = _now_bj()
        flash('导入模板已更新！', 'success')
    else:
        template = ImportTemplate(category_id=category_id, filename=filename, filepath=filepath,
                                  field_mapping=json.dumps(field_mapping, ensure_ascii=False),
                                  default_express_type=default_express_type)
        db.session.add(template)
        flash('导入模板上传成功！', 'success')

    db.session.commit()
    return redirect(url_for('templates.admin_templates'))


@bp.route('/admin/import_template/delete/<int:template_id>', methods=['POST'])
@role_required('admin')
def delete_import_template(template_id):
    """删除导入模板"""
    template = ImportTemplate.query.get_or_404(template_id)
    if os.path.exists(template.filepath):
        os.remove(template.filepath)
    db.session.delete(template)
    db.session.commit()
    flash('导入模板删除成功！', 'success')
    return redirect(url_for('import_routes.batch_import_page'))


@bp.route('/admin/import_template/edit', methods=['POST'])
@role_required('admin')
def edit_import_template():
    """编辑导入模板（只更新字段映射，可选重新上传文件）"""
    template_id = request.form.get('template_id', type=int)
    template = ImportTemplate.query.get_or_404(template_id)

    # 收集字段映射
    field_mapping = {}
    idx = 0
    while True:
        excel_col = request.form.get(f'excel_col_{idx}', '').strip()
        order_field = request.form.get(f'order_field_{idx}', '').strip()
        if not excel_col and not order_field:
            break
        if excel_col and order_field:
            field_mapping[excel_col] = order_field
        idx += 1

    if not field_mapping:
        flash('至少需要配置一个字段映射！', 'danger')
        return redirect(url_for('templates.admin_templates'))

    # 处理可选的文件重新上传
    file = request.files.get('file')
    if file and file.filename:
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        new_filename = f"import_{template.category_id}_{timestamp}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, new_filename)
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(filepath)
        # 删除旧文件
        if os.path.exists(template.filepath):
            os.remove(template.filepath)
        template.filepath = filepath
        template.filename = filename

    template.field_mapping = json.dumps(field_mapping, ensure_ascii=False)
    template.default_express_type = request.form.get('default_express_type', '').strip() or None
    
    # 收集列预处理配置
    column_preprocessing = {}
    idx = 0
    while True:
        pp_col = request.form.get(f'pp_col_{idx}', '').strip()
        pp_type = request.form.get(f'pp_type_{idx}', '').strip()
        if not pp_col:
            break
        if pp_type:
            config = {'type': pp_type}
            if pp_type == 'regex_replace':
                config['pattern'] = request.form.get(f'pp_pattern_{idx}', '')
                config['replacement'] = request.form.get(f'pp_replacement_{idx}', '')
            elif pp_type == 'prefix_remove':
                config['prefix'] = request.form.get(f'pp_prefix_{idx}', '')
            elif pp_type == 'suffix_remove':
                config['suffix'] = request.form.get(f'pp_suffix_{idx}', '')
            elif pp_type == 'strip_chars':
                config['chars'] = request.form.get(f'pp_chars_{idx}', '')
            elif pp_type == 'replace':
                config['old'] = request.form.get(f'pp_old_{idx}', '')
                config['new'] = request.form.get(f'pp_new_{idx}', '')
            column_preprocessing[pp_col] = config
        idx += 1
    template.column_preprocessing = json.dumps(column_preprocessing, ensure_ascii=False)
    
    db.session.commit()
    flash(f'导入模板 "{template.category.name}" 更新成功！', 'success')
    return redirect(url_for('templates.admin_templates'))


@bp.route('/api/import_template/mapping/<int:category_id>')
@login_required
def api_import_template_mapping(category_id):
    """获取某类别的导入模板映射配置"""
    template = ImportTemplate.query.filter_by(category_id=category_id).first()
    if template:
        return jsonify({
            'mapping': json.loads(template.field_mapping) if template.field_mapping else {},
            'default_express_type': template.default_express_type or ''
        })
    return jsonify({'mapping': {}, 'default_express_type': ''})


# ============ 业绩报表模板管理 ============
@bp.route('/admin/performance_templates')
@role_required('admin')
def admin_performance_templates():
    """业绩报表模板管理页面"""
    templates = PerformanceReportTemplate.query.filter_by(is_active=True).all()
    return render_template('admin_performance_templates.html',
                           templates=templates,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/admin/performance_template/save', methods=['POST'])
@role_required('admin')
def save_performance_template():
    """保存业绩报表模板配置"""
    template_id = request.form.get('template_id', type=int)
    name = request.form.get('name', '').strip()
    template_file = request.files.get('template_file')

    # 字段映射
    field_mapping = {}
    excel_fields = request.form.getlist('excel_field[]')
    order_fields = request.form.getlist('order_field[]')
    for ef, of in zip(excel_fields, order_fields):
        if ef and of:
            field_mapping[ef] = of

    # 处理模板文件上传
    filepath = None
    if template_file and template_file.filename:
        filename = secure_filename(template_file.filename)
        filepath = os.path.join('uploads', 'performance_templates', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        template_file.save(filepath)

    if template_id:
        template = PerformanceReportTemplate.query.get_or_404(template_id)
        template.name = name
        template.field_mapping = json.dumps(field_mapping, ensure_ascii=False)
        if filepath:
            template.filepath = filepath
    else:
        if not filepath:
            flash('请上传模板文件！', 'danger')
            return redirect(url_for('import_routes.admin_performance_templates'))
        template = PerformanceReportTemplate(
            name=name,
            field_mapping=json.dumps(field_mapping, ensure_ascii=False),
            filepath=filepath
        )
        db.session.add(template)

    db.session.commit()
    flash('业绩报表模板保存成功！', 'success')
    return redirect(url_for('import_routes.admin_performance_templates'))


@bp.route('/admin/performance_template/delete/<int:template_id>', methods=['POST'])
@role_required('admin')
def delete_performance_template(template_id):
    """删除业绩报表模板"""
    template = PerformanceReportTemplate.query.get_or_404(template_id)
    template.is_active = False
    db.session.commit()
    flash('业绩报表模板已删除！', 'success')
    return redirect(url_for('import_routes.admin_performance_templates'))


def extract_digits(s: str) -> str:
    """提取字符串中的所有数字，拼接成纯数字字符串"""
    if not s:
        return ""
    # 匹配所有数字字符并拼接

    ss = ''.join(re.findall(r'\d+', s))
    print(f"===========拼接后的字符串{ss}")
    return ss


def col_letter_to_index(col_str: str) -> int:
    """Excel列号字母转数字索引，如 A->1, B->2, Z->26, AA->27"""
    col_str = col_str.strip().upper()
    if not col_str or not col_str[0].isalpha():
        return 0
    # 限制长度防止数值过大
    if len(col_str) > 3:
        return 0
    result = 0
    for ch in col_str:
        result = result * 26 + (ord(ch) - ord('A') + 1)
        # 防止结果过大
        if result > 1000:
            return 0
    return result


def resolve_col_index(excel_col, header_map):
    """解析Excel列标识为列索引，支持列名或列号(A,B,C...)"""
    # 先尝试列名匹配（原始值）
    if excel_col in header_map:
        return header_map[excel_col]
    
    # 尝试标准化的列名匹配（去除空格等）
    excel_col_clean = ''.join(str(excel_col).strip().split())
    if excel_col_clean in header_map:
        return header_map[excel_col_clean]
    
    # 检查是否是有效的列字母格式（只包含A-Z，不超过3个字符）
    excel_col_upper = str(excel_col).strip().upper()
    if len(excel_col_upper) > 3:
        return None
    if not all(c.isalpha() for c in excel_col_upper):
        return None
    
    # 再尝试列号匹配（A->1, B->2...）
    col_idx = col_letter_to_index(excel_col_upper)
    # 检查是否在合理的列索引范围内（1-1000）
    if 0 < col_idx <= 1000:
        return col_idx
    return None


def is_col_letter_mapping(mapping):
    """判断映射是否全部使用列号（A,B,C...）"""
    for key in mapping.keys():
        key = key.strip().upper()
        if not key:
            return False
        
        # 严格检查：必须是纯英文字母（A-Z），长度1-3
        if len(key) > 3:
            return False
        
        # 检查每个字符是否都是A-Z
        for c in key:
            if not ('A' <= c <= 'Z'):
                return False
        
        # 确认是有效的列号（A-Z, AA-ZZ, AAA-ZZZ）
        continue
    return True


def _is_valid_tracking_number(s):
    """判断是否是有效的快递单号"""
    if not s:
        return False
    s = s.strip()
    # 顺丰：12位纯数字（或以SF开头）
    if re.match(r'^(SF\d{10,15}|\d{12,20})$', s):
        return True
    # 中通、圆通、韵达等：纯数字，通常10-20位
    if re.match(r'^\d{10,20}$', s):
        return True
    # 京东：以JD开头
    if re.match(r'^JD\d+', s):
        return True
    # 包含字母和数字混合（如YT、YZ开头等）
    if re.match(r'^[A-Za-z]{1,4}\d{10,20}$', s):
        return True
    return False


def _find_tracking_from_row(row_data):
    """从行数据所有字段中查找有效的快递单号"""
    for key, value in row_data.items():
        val = str(value).strip()
        if val and _is_valid_tracking_number(val):
            return val
    return ''


# ============ 自动检测表头行 (V2) ============

def _auto_detect_header_row_v2(ws, mapping, is_xls):
    """
    自动检测 Excel 中哪一行是真正的表头行（0-based 索引）

    策略：
      1. 列号映射模式：直接返回 0
      2. 列名映射模式：扫描前10行，找与 mapping 中列名匹配度最高的行
         - 优先排除只有少数非空单元格的行（如日期标题行）
         - 优先排除纯数字行
         - 优先选择与 mapping 列名精确匹配/部分匹配的行
    """
    if is_col_letter_mapping(mapping):
        return 0

    target_cols = set(mapping.keys())
    if not target_cols:
        return 0

    max_rows = ws.nrows if is_xls else ws.max_row
    total_cols = ws.ncols if is_xls else ws.max_column
    scan_limit = min(10, max_rows)

    best_idx = 1
    best_score = -1
    best_debug = {}

    for r in range(scan_limit):
        cells = []
        for c in range(total_cols):
            try:
                if is_xls:
                    v = ws.cell(r, c).value
                else:
                    v = ws.cell(row=r + 1, column=c + 1).value
                cells.append(str(v).strip() if v is not None else '')
            except Exception:
                cells.append('')

        non_empty = [c for c in cells if c]
        if not non_empty:
            continue

        score = 0
        matched = set()

        for val in non_empty:
            if val in target_cols:
                score += 100
                matched.add(val)
                continue
            simplified = ''.join(val.split())
            for t in target_cols:
                ts = ''.join(t.split())
                if simplified == ts:
                    score += 90
                    matched.add(t)
                    break
            stripped = re.sub(r'\d+$', '', simplified)
            if stripped in target_cols:
                score += 70
                matched.add(stripped)

        # 惩罚：只有第一列有值的行（如日期标题）
        first_col_only = (cells[0] and all(c == '' for c in cells[1:]))
        if first_col_only:
            score -= 200

        # 惩罚：全数字内容（非表头特征）
        has_digit = sum(1 for c in non_empty if re.match(r'^[\d\.\-]+$', c))
        if has_digit == len(non_empty) and len(non_empty) > 1:
            score -= 150

        # 惩罚：含 "日期/报表/时间/统计" 等标题关键词
        title_keywords = ('日期', '报表', '统计', '汇总', '时间', '序号', '总表', '部门')
        if len(non_empty) <= 2 and any(k in cells[0] for k in title_keywords if cells[0]):
            score -= 100

        # 没匹配到任何列的行
        if len(matched) <= 0:
            score -= 50

        # 轻偏好前几行
        if r <= 1:
            score += 5

        debug_line = f"行{r+1}: score={score}, 匹配={matched}, 内容={non_empty[:8]}"
        best_debug[r] = debug_line

        if score > best_score:
            best_score = score
            best_idx = r

    print(f"[DEBUG] === 表头自动检测结果 ===")
    for line in best_debug.values():
        print(f"[DEBUG]   {line}")
    print(f"[DEBUG]   最终选: 第 {best_idx + 1} 行 (index={best_idx}, score={best_score})")

    if best_score <= 0:
        best_idx = 1
        print(f"[DEBUG]   无高分匹配，退回第 2 行作为表头")

    return best_idx


def _auto_detect_header_row(ws, mapping, is_xls, no_header):
    """旧接口兼容"""
    return _auto_detect_header_row_v2(ws, mapping, is_xls)


@bp.route('/batch_import/execute', methods=['POST'])
@role_required('shipper', 'admin')
def execute_batch_import():
    """执行批量导入"""
    category_id = request.form.get('category_id', type=int)
    file = request.files.get('file')

    if not category_id:
        flash('请选择产品类别！', 'danger')
        return redirect(url_for('import_routes.batch_import_page'))
    if not file or file.filename == '':
        flash('请选择文件！', 'danger')
        return redirect(url_for('import_routes.batch_import_page'))

    template = ImportTemplate.query.filter_by(category_id=category_id).first()
    if not template:
        flash('该类别未配置导入模板！', 'danger')
        return redirect(url_for('import_routes.batch_import_page'))

    mapping = json.loads(template.field_mapping) if template.field_mapping else {}
    preprocessing = json.loads(template.column_preprocessing) if template.column_preprocessing else {}
    skip_rows = template.skip_rows or 0
    if not mapping:
        flash('导入模板未配置字段映射！', 'danger')
        return redirect(url_for('import_routes.batch_import_page'))

    # 读取上传的Excel
    is_xls = file.filename.lower().endswith('.xls')

    try:
        # 判断是否全部使用列号映射（无表头模式）
        no_header = is_col_letter_mapping(mapping)

        if is_xls:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file.read())
            ws = wb.sheet_by_index(0)

            print(f"[DEBUG] === Excel文件信息 (xls) ===")
            print(f"[DEBUG] 总行数: {ws.nrows}, 总列数: {ws.ncols}")
            
            # 1) 先确定表头行位置（从检测到的行建立 header_map）
            if skip_rows > 0:
                header_row_index = skip_rows
            elif no_header:
                header_row_index = 0
            else:
                header_row_index = _auto_detect_header_row_v2(ws, mapping, is_xls)
            print(f"[DEBUG] skip_rows={skip_rows}, no_header={no_header}, 表头行索引={header_row_index} (Excel行号={header_row_index+1})")

            # 2) 从检测到的表头行建立 header_map
            header_map = {}
            if not no_header:
                print(f"[DEBUG] 从第 {header_row_index + 1} 行读取列名:")
                for col in range(ws.ncols):
                    cell_value = ws.cell(header_row_index, col).value
                    col_letter = chr(ord('A') + col)
                    print(f"[DEBUG]   列 {col_letter}: {repr(cell_value)}")
                    if cell_value is not None:
                        cleaned_value = str(cell_value).strip()
                        cleaned_value = ''.join(cleaned_value.split())
                        if cleaned_value:
                            header_map[cleaned_value] = col
                            header_map[str(cell_value).strip()] = col
                print(f"[DEBUG] 建立的header_map: {header_map}")

            # 3) 从表头行的下一行开始读取数据
            start_row = header_row_index + 1
            print(f"[DEBUG] 数据从第 {start_row + 1} 行开始读取")

            rows_data = []
            for row in range(start_row, ws.nrows):
                row_data = {}
                for excel_col, field in mapping.items():
                    col_idx = resolve_col_index(excel_col, header_map)
                    if col_idx is not None:
                        raw_value = ws.cell(row, col_idx).value
                        if excel_col in preprocessing:
                            raw_value = apply_column_preprocessing(raw_value, preprocessing[excel_col])
                        row_data[field] = raw_value
                rows_data.append(row_data)
                if len(rows_data) <= 5:
                    print(f"[DEBUG]   读取第 {row+1} 行数据: {row_data}")

            print(f"[DEBUG] 从Excel共读取了 {len(rows_data)} 行原始数据")
        else:
            import openpyxl
            file.seek(0)
            wb = openpyxl.load_workbook(file, read_only=False, data_only=True)

            print(f"[DEBUG] === Excel文件信息 (xlsx) ===")
            print(f"[DEBUG] Workbook工作表: {wb.sheetnames}")

            ws = wb.active
            print(f"[DEBUG] 当前工作表: {ws.title}")
            print(f"[DEBUG] 工作表总行数: {ws.max_row}, 总列数: {ws.max_column}")

            # 调试：遍历前10行前15列
            print(f"[DEBUG] 查看前10行前15列的内容:")
            for row in range(1, min(11, ws.max_row + 1)):
                row_content = []
                for col in range(1, min(16, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    val = cell.value
                    col_letter = chr(ord('A') + col - 1)
                    row_content.append(f"{col_letter}{row}={repr(val)}")
                print(f"[DEBUG]   行 {row}: {' | '.join(row_content)}")

            # 1) 先确定表头行位置
            if skip_rows > 0:
                header_row_index = skip_rows
            elif no_header:
                header_row_index = 0
            else:
                header_row_index = _auto_detect_header_row_v2(ws, mapping, is_xls)
            print(f"[DEBUG] skip_rows={skip_rows}, no_header={no_header}, 表头行索引={header_row_index} (Excel行号={header_row_index+1})")

            # 2) 从检测到的表头行建立 header_map（xlsx 行号是 1-based）
            header_map = {}
            if not no_header:
                header_row_1based = header_row_index + 1
                print(f"[DEBUG] 从第 {header_row_1based} 行读取列名:")
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=header_row_1based, column=col).value
                    col_letter = chr(ord('A') + col - 1)
                    print(f"[DEBUG]   列 {col_letter}: {repr(cell_value)}")
                    if cell_value is not None:
                        cleaned_value = str(cell_value).strip()
                        cleaned_value = ''.join(cleaned_value.split())
                        if cleaned_value:
                            header_map[cleaned_value] = col
                            header_map[str(cell_value).strip()] = col
                print(f"[DEBUG] 建立的header_map: {header_map}")

            # 3) 从表头行的下一行开始读取数据
            start_row_1based = header_row_index + 2
            print(f"[DEBUG] 数据从第 {start_row_1based} 行开始读取")

            rows_data = []
            for row in range(start_row_1based, ws.max_row + 1):
                row_data = {}
                for excel_col, field in mapping.items():
                    col_idx = resolve_col_index(excel_col, header_map)
                    if col_idx:
                        raw_value = ws.cell(row=row, column=col_idx).value
                        if excel_col in preprocessing:
                            raw_value = apply_column_preprocessing(raw_value, preprocessing[excel_col])
                        row_data[field] = raw_value
                rows_data.append(row_data)
                if len(rows_data) <= 5:
                    print(f"[DEBUG]   读取第 {row} 行数据: {row_data}")

            print(f"[DEBUG] 从Excel共读取了 {len(rows_data)} 行原始数据")
            wb.close()

        # 更新订单
        updated_count = 0
        not_found_count = 0
        skipped_count = 0
        category = Category.query.get(category_id)
        default_express = template.default_express_type or ''

        print(f"[DEBUG] 导入模板映射: {mapping}")
        print(f"[DEBUG] Excel表头: {header_map}")
        print(f"[DEBUG] 默认快递种类: {default_express}")
        print(f"[DEBUG] 类别: {category.name}")
        print(f"[DEBUG] 共读取 {len(rows_data)} 行数据")

        # 字段预处理：大写字母、去特殊字符、只保留字母数字汉字
        def normalize_field(s):
            s = str(s).upper()
            s = re.sub(r'[^\w\u4e00-\u9fff]', '', s)
            return s

        # 预加载所有业务员并建立索引
        # 注意：mapping的值是字段名，不是键！
        mapping_values = list(mapping.values())
        has_group_in_mapping = 'group_name' in mapping_values
        has_salesman_in_mapping = 'salesman_name' in mapping_values
        print(f"[DEBUG] has_group_in_mapping={has_group_in_mapping}, has_salesman_in_mapping={has_salesman_in_mapping}")
        print(f"[DEBUG] mapping keys: {list(mapping.keys())}")
        print(f"[DEBUG] mapping values: {list(mapping.values())}")
        
        salesman_map = {}
        if has_salesman_in_mapping:
            all_salesmen = User.query.filter(User.roles.like('%salesman%')).all()
            for s in all_salesmen:
                salesman_map[normalize_field(s.name)] = s

        # 先查询所有订单，统计类别分布
        all_orders_all = Order.query.all()
        print(f"[DEBUG] 数据库中总共有 {len(all_orders_all)} 个订单")
        
        # 统计各个类别的订单数
        category_stats = {}
        for order in all_orders_all:
            cat = order.category
            if cat not in category_stats:
                category_stats[cat] = 0
            category_stats[cat] += 1
        print(f"[DEBUG] 各类别订单统计: {category_stats}")
        
        # 专门查询我们选择的类别的订单
        target_category_orders = Order.query.filter(Order.category == category.name).all()
        print(f"[DEBUG] 选择的类别'{category.name}'有 {len(target_category_orders)} 个订单")
        for i, order in enumerate(target_category_orders[:20]):
            print(f"[DEBUG]   目标订单{i+1}: id={order.id}, 类别={repr(order.category)}, 状态={repr(order.status)}, 组别={repr(order.group_name)}, 客户名={repr(order.customer_name)}, export_marked={getattr(order, 'export_marked', 'N/A')}")

        # 预加载所有待发货订单并建立索引
        # 构建查询条件：只查询待发货(status=submitted) + 已标记导出(export_marked=True) + 精确类别匹配
        orders_query = Order.query.filter(
            Order.category == category.name,
            Order.status == 'submitted',
            Order.export_marked == True
        )
        all_orders = orders_query.all()
        
        print(f"[DEBUG] 查询到 {len(all_orders)} 个待发货订单 (类别='{category.name}', status=submitted, export_marked=True)")
        print(f"[DEBUG] 提示：如果查询结果为0，请确认数据库里的订单类别名称是否与选择的类别完全一致！")
        
        # 输出前5个待发货订单信息
        print(f"[DEBUG] 前5个待发货订单:")
        for i, order in enumerate(all_orders[:5]):
            print(f"[DEBUG]   订单{i+1}: id={order.id}, 组别={repr(order.group_name)}, 业务员={repr(order.salesman.name if order.salesman else None)}, 客户名={repr(order.customer_name)}, 状态={order.status}")
        
        # 建立订单索引，严格按照模板配置的字段匹配
        order_index = {}
        
        for order in all_orders:
            # 获取业务员姓名（标准化）
            salesman_name_norm = ""
            if order.salesman:
                salesman_name_norm = normalize_field(order.salesman.name)
            
            group_name_norm = normalize_field(order.group_name)
            customer_name_norm = normalize_field(order.customer_name)
            
            # 根据模板中实际配置的字段构建索引key
            if has_group_in_mapping and has_salesman_in_mapping:
                # 模板配置了 组别+业务员+客户名 三个字段
                key = (group_name_norm, salesman_name_norm, customer_name_norm)
            elif has_group_in_mapping:
                # 模板配置了 组别+客户名 两个字段
                key = (group_name_norm, customer_name_norm)
            elif has_salesman_in_mapping:
                # 模板配置了 业务员+客户名 两个字段
                key = (salesman_name_norm, customer_name_norm)
            else:
                # 模板只配置了 客户名 一个字段
                key = customer_name_norm
            
            if key not in order_index:
                order_index[key] = []
            order_index[key].append(order)
        
        print(f"[DEBUG] 建立了 {len(order_index)} 个订单索引，匹配策略：{'组别+业务员+客户名' if has_group_in_mapping and has_salesman_in_mapping else '组别+客户名' if has_group_in_mapping else '业务员+客户名' if has_salesman_in_mapping else '仅客户名'}")

        # 按业务员收集发货的订单信息
        salesman_orders_map = {}

        print(f"[DEBUG] === 开始处理 {len(rows_data)} 行数据 ===")
        
        for idx, row_data in enumerate(rows_data):
            print(f"[DEBUG] 处理第 {idx+1} 行: {row_data}")
            group_name = str(row_data.get('group_name', '')).strip()
            salesman_name = str(row_data.get('salesman_name', '')).strip()
            customer_name = str(row_data.get('customer_name', '')).strip()
            tracking_number = str(row_data.get('tracking_number', '')).strip()
            express_type = str(row_data.get('express_type', '')).strip()
            new_product_info = str(row_data.get('product_info', '')).strip()  # 托寄物内容

            # 校验快递单号：如果不是有效单号，从该行所有数据中查找
            if tracking_number and not _is_valid_tracking_number(tracking_number):
                found_number = _find_tracking_from_row(row_data)
                if found_number:
                    tracking_number = found_number
                else:
                    tracking_number = ''  # 找不到有效单号则置空

            # 如果Excel中没有快递种类，使用默认值
            if not express_type and default_express:
                express_type = default_express

            group_name_norm = normalize_field(group_name)
            customer_name_norm = normalize_field(customer_name)

            # 必填字段校验（只有映射中有组别时才要求）
            if has_group_in_mapping and not group_name:
                print(f"[DEBUG]   第 {idx+1} 行跳过：缺少组别")
                skipped_count += 1
                continue

            if not tracking_number or tracking_number == '0':
                print(f"[DEBUG]   第 {idx+1} 行跳过：缺少快递单号 (tracking_number={repr(tracking_number)})")
                skipped_count += 1
                continue

            if not customer_name:
                print(f"[DEBUG]   第 {idx+1} 行跳过：缺少客户姓名")
                skipped_count += 1
                continue

            # 查找业务员（从预加载的索引中找）
            salesman = None
            if has_salesman_in_mapping and salesman_name:
                salesman_key = normalize_field(salesman_name)
                salesman = salesman_map.get(salesman_key)

            # 如果映射中有业务员字段但找不到，则跳过
            if has_salesman_in_mapping and not salesman:
                skipped_count += 1
                continue

            # 查找匹配的订单（严格按照模板配置的字段匹配）
            salesman_name_norm = normalize_field(salesman_name)
            
            # 根据模板中实际配置的字段构建查找key
            if has_group_in_mapping and has_salesman_in_mapping:
                # 模板配置了 组别+业务员+客户名 三个字段
                key = (group_name_norm, salesman_name_norm, customer_name_norm)
                match_desc = f"组别={repr(group_name)} (标准化={repr(group_name_norm)}), 业务员={repr(salesman_name)} (标准化={repr(salesman_name_norm)}), 客户名={repr(customer_name)} (标准化={repr(customer_name_norm)})"
            elif has_group_in_mapping:
                # 模板配置了 组别+客户名 两个字段
                key = (group_name_norm, customer_name_norm)
                match_desc = f"组别={repr(group_name)} (标准化={repr(group_name_norm)}), 客户名={repr(customer_name)} (标准化={repr(customer_name_norm)})"
            elif has_salesman_in_mapping:
                # 模板配置了 业务员+客户名 两个字段
                key = (salesman_name_norm, customer_name_norm)
                match_desc = f"业务员={repr(salesman_name)} (标准化={repr(salesman_name_norm)}), 客户名={repr(customer_name)} (标准化={repr(customer_name_norm)})"
            else:
                # 模板只配置了 客户名 一个字段
                key = customer_name_norm
                match_desc = f"客户名={repr(customer_name)} (标准化={repr(customer_name_norm)})"
            
            print(f"[DEBUG]   尝试匹配：{match_desc}")
            
            candidates = order_index.get(key, [])
            print(f"[DEBUG]   找到 {len(candidates)} 个候选订单")
            
            # 输出所有候选订单信息
            if candidates:
                print(f"[DEBUG]   候选订单列表:")
                for i, c in enumerate(candidates):
                    print(f"[DEBUG]     候选{i+1}: id={c.id}, 客户名={repr(c.customer_name)}, 状态={c.status}, 快递单号={repr(c.tracking_number)}")

            if not candidates:
                print(f"[DEBUG]   未找到匹配订单")
                not_found_count += 1
                continue

            # 如果有托寄物内容，用数字匹配筛选
            order = None
            if new_product_info:
                extract_digits_func = lambda s: ''.join(re.findall(r'\d+', str(s) or ''))
                target_digits = extract_digits_func(new_product_info)
                if target_digits:
                    # 尝试找数字匹配的订单（包含匹配：订单数字包含Excel数字）
                    for c in candidates:
                        order_digits = extract_digits_func(c.product_info)
                        # 检查订单数字是否包含Excel数字，或Excel数字是否包含订单数字
                        if target_digits in order_digits or order_digits in target_digits:
                            order = c
                            break

            # 如果没有找到数字匹配的，取第一个候选订单
            if not order and candidates:
                order = candidates[0]
            if order:
                print(f"[DEBUG]   找到订单: id={order.id}, status={order.status}")
                order.tracking_number = tracking_number
                order.express_type = express_type or order.express_type
                order.status = 'shipped'

                # 判断是否为主品，非主品发货即签收
                if category.is_main_product:
                    # 主品：正常发货流程
                    if express_type == '顺丰':
                        order.logistics_status = '已发货'
                else:
                    # 非主品：发货即签收
                    order.logistics_status = '已签收'
                    print(f"[DEBUG]   非主品类别，直接标记为已签收")

                updated_count += 1

                # 按业务员收集订单信息
                salesman_id = order.salesman_id
                if salesman_id not in salesman_orders_map:
                    salesman_orders_map[salesman_id] = []
                salesman_orders_map[salesman_id].append({
                    'group_name': order.group_name,
                    'express_type': express_type,
                    'tracking_number': tracking_number
                })
            else:
                print(f"[DEBUG]   未找到匹配订单")
                not_found_count += 1

        # 一次性提交所有数据库更新
        db.session.commit()

        # 异步按业务员汇总发送通知（不阻塞主线程）
        if salesman_orders_map:
            import threading
            def send_notifications_async():
                for salesman_id, orders in salesman_orders_map.items():
                    try:
                        # 构建汇总通知内容
                        order_lines = []
                        for idx, order_info in enumerate(orders, 1):
                            line = f"{idx}. {order_info['group_name']} - {order_info['express_type']}：{order_info['tracking_number']}"
                            order_lines.append(line)
                        
                        content = f"【批量发货通知】\n本次有 {len(orders)} 个订单已发货：\n" + "\n".join(order_lines) + "\n请注意查收！"
                        
                        notify_users([salesman_id], content)
                        print(f"[DEBUG] 已给业务员 {salesman_id} 发送批量通知")
                    except Exception as e:
                        print(f"[DEBUG] 发送通知失败: {e}")
            threading.Thread(target=send_notifications_async, daemon=True).start()

        msg = f'批量导入完成，共更新 {updated_count} 条订单！'
        if not_found_count > 0:
            msg += f'（{not_found_count} 条未匹配到订单）'
        if skipped_count > 0:
            msg += f'（{skipped_count} 条因数据不完整已跳过）'
        flash(msg, 'success')

    except Exception as e:
        import traceback
        print(f"[ERROR] 导入失败: {traceback.format_exc()}")
        flash(f'导入失败：{str(e)}', 'danger')

    return redirect(url_for('import_routes.batch_import_page'))


def merge_product_info(original_info, new_info):
    """
    合并托寄物内容：数量累加，产品名称不变
    原格式示例: "康欣胶囊*3+赠品*1" 或 "康欣胶囊3个"
    新格式示例: "康欣胶囊*2"
    结果: "康欣胶囊*5+赠品*1"
    """
    import re

    if not original_info:
        return new_info
    if not new_info:
        return original_info

    # 解析产品信息为字典 {产品名: 数量}
    def parse_product_info(info):
        products = {}
        # 支持多种格式: "产品*数量", "产品数量个", "产品数量盒" 等
        # 按 + 分割
        parts = re.split(r'[+、，,]', info)
        for part in parts:
            part = part.strip()
            if not part:
                continue

            # 尝试匹配 "产品名*数量" 或 "产品名数量单位"
            match = re.match(r'^(.+?)[*×xX](\d+)', part)
            if match:
                name = match.group(1).strip()
                qty = int(match.group(2))
                products[name] = products.get(name, 0) + qty
                continue

            # 尝试匹配 "产品名数量" (如 "康欣胶囊3")
            match = re.match(r'^(.+?)(\d+)(?:个|盒|瓶|袋|件)?$', part)
            if match:
                name = match.group(1).strip()
                qty = int(match.group(2))
                products[name] = products.get(name, 0) + qty
                continue

            # 没有数量的情况，默认为1
            products[part] = products.get(part, 0) + 1

        return products

    # 解析原有和新的产品信息
    original_products = parse_product_info(original_info)
    new_products = parse_product_info(new_info)

    # 合并数量（产品名称不变，数量累加）
    for name, qty in new_products.items():
        if name in original_products:
            original_products[name] += qty
        else:
            original_products[name] = qty

    # 重新组装为字符串
    result_parts = []
    for name, qty in original_products.items():
        result_parts.append(f"{name}*{qty}")

    return '+'.join(result_parts)
