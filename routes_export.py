# -*- coding: utf-8 -*-
"""导出相关路由"""
import os
import re
import json
import tempfile
import zipfile
from datetime import datetime
from collections import defaultdict
from flask import Blueprint, request, redirect, url_for, flash, render_template, jsonify, send_file
from flask_login import login_required, current_user

from models import db, Order, Category, ExcelTemplate
from helpers import role_required, get_unread_count
from routes_import import resolve_col_index

bp = Blueprint('export', __name__)


# ============ 模板映射API ============
@bp.route('/api/template/mapping/<int:category_id>')
@login_required
def api_template_mapping(category_id):
    """获取某类别的模板映射配置"""
    template = ExcelTemplate.query.filter_by(category_id=category_id).first()
    if template and template.field_mapping:
        return jsonify(json.loads(template.field_mapping))
    return jsonify({})


# ============ 导出待发货订单 ============
def extract_product_qty(product_info):
    """从产品信息中提取数量，如 '康欣胶囊x3' -> 3"""
    match = re.search(r'[xX×*]\s*(\d+)', product_info)
    if match:
        return int(match.group(1))
    match = re.search(r'(\d+)\s*[件盒瓶袋套个只箱]', product_info)
    if match:
        return int(match.group(1))
    return 1


def _get_order_field_value(order, order_field, default_zero=False):
    """根据字段名获取订单对应值

    支持多种格式：
    1. 简单字段名: customer_name, phone 等
    2. 正则表达式: /regex/ 或 /field_name/regex/
    3. 固定文本: {"type": "fixed", "value": "文本内容"}
    4. 条件匹配: {"type": "condition", "field": "字段名", "conditions": [{"match": "值1", "result": "结果1"}, ...], "default": "默认值"}

    参数:
        default_zero: 正则未匹配时是否返回'0'（导出Excel时使用）
    """
    import json as _json

    # 处理JSON对象格式（固定文本、条件匹配）
    if isinstance(order_field, dict) or (isinstance(order_field, str) and order_field.startswith('{')):
        try:
            config = _json.loads(order_field) if isinstance(order_field, str) else order_field
            config_type = config.get('type', '')

            # 固定文本
            if config_type == 'fixed':
                return config.get('value', '')

            # 条件匹配
            elif config_type == 'condition':
                source_field = config.get('field', 'product_info')
                source_value = getattr(order, source_field, '') or ''
                if not isinstance(source_value, str):
                    source_value = str(source_value)

                conditions = config.get('conditions', [])
                for cond in conditions:
                    match = cond.get('match', '')
                    result = cond.get('result', '')
                    # * 表示匹配所有
                    if match == '*' or match == '':
                        return result
                    # 支持包含匹配
                    if match in source_value:
                        return result

                return config.get('default', '')
        except _json.JSONDecodeError:
            pass

    # 正则表达式提取
    if isinstance(order_field, str) and order_field.startswith('/') and order_field.endswith('/'):
        import re as _re
        expr = order_field[1:-1]

        # 判断是否指定了源字段：/字段名/正则/组号
        # 已知的字段名列表
        known_fields = ['product_info', 'address', 'phone', 'customer_name', 'remark',
                       'group_name', 'gift_info', 'paid_amount', 'collect_amount']

        # 尝试解析：/field_name/regex/group
        source_field = 'product_info'  # 默认从产品信息提取
        parts = expr.split('/')

        if len(parts) >= 3 and parts[0] in known_fields:
            # 格式: /field_name/regex 或 /field_name/regex/group
            source_field = parts[0]
            regex_part = '/'.join(parts[1:])
        else:
            # 格式: /regex 或 /regex/group（默认从product_info提取）
            regex_part = expr

        # 解析组号
        rparts = regex_part.rsplit('/', 1)
        if len(rparts) == 2 and rparts[1].isdigit():
            pattern, group_idx = rparts[0], int(rparts[1])
        else:
            pattern, group_idx = regex_part, 1

        # 获取源字段值
        source_value = getattr(order, source_field, '') or ''
        if not isinstance(source_value, str):
            source_value = str(source_value)

        print(f"[DEBUG] 正则提取: pattern={pattern}, source_field={source_field}, source_value={source_value[:50]}...")

        try:
            m = _re.search(pattern, source_value)
            if m:
                result = m.group(group_idx) if group_idx <= len(m.groups()) else (m.group(0) if group_idx == 0 else '')
                print(f"[DEBUG] 正则匹配成功: groups={m.groups()}, result={result}")
                return result
            else:
                print(f"[DEBUG] 正则未匹配到")
                # 导出Excel时，正则未匹配且是数字提取模式，返回'0'
                if default_zero and (r'\d' in pattern or any(c.isdigit() for c in pattern)):
                    print(f"[DEBUG] 导出模式，数字提取未匹配返回0")
                    return '0'
        except _re.error as e:
            print(f"[WARN] 正则表达式错误: {order_field}, 错误: {e}")
        return ''

    # 处理字符串格式的字段名
    order_field_str = str(order_field) if order_field else ''

    if order_field_str == 'customer_name':
        return order.customer_name or ''
    elif order_field_str == 'phone':
        return order.phone or ''
    elif order_field_str == 'address':
        return order.address or ''
    elif order_field_str == 'product_info':
        return order.product_info or ''
    elif order_field_str == 'has_gift':
        return 1 if order.has_gift else 0
    elif order_field_str == 'gift_info':
        return order.gift_info or ''
    elif order_field_str == 'remark':
        return order.remark or ''
    elif order_field_str == 'paid_amount':
        return order.paid_amount or ''
    elif order_field_str == 'collect_amount':
        return order.collect_amount if order.collect_amount else 0
    elif order_field_str == '__extract_qty__':
        return extract_product_qty(order.product_info)
    elif order_field_str == 'group_name':
        return order.group_name or ''
    elif order_field_str == 'salesman_name':
        return order.salesman.name if order.salesman else ''
    elif order_field_str == '__date__':
        return datetime.now().strftime('%Y-%m-%d')
    elif order_field_str == '__seq__':
        return ''
    return ''


@bp.route('/export/preview')
@role_required('shipper', 'admin')
def export_preview():
    """导出预览页"""
    customer_keyword = request.args.get('customer_keyword', '').strip()
    tracking_keyword = request.args.get('tracking_keyword', '').strip()
    category_filter = request.args.get('category', '').strip()
    salesman_filter = request.args.get('salesman_id', '').strip()

    # 查询待发货订单（排除已标记导出的）
    query = Order.query.filter(Order.status == 'submitted', Order.export_marked == False)
    if customer_keyword:
        query = query.filter(Order.customer_name.contains(customer_keyword))
    if tracking_keyword:
        query = query.filter(Order.tracking_number.contains(tracking_keyword))
    if category_filter:
        query = query.filter(Order.category == category_filter)
    if salesman_filter:
        query = query.filter(Order.salesman_id == salesman_filter)
    orders = query.order_by(Order.category, Order.create_time.desc()).all()

    # 按类别分组
    orders_by_category = defaultdict(list)
    for order in orders:
        orders_by_category[order.category].append(order)

    return render_template('export_preview.html',
                          orders_by_category=orders_by_category,
                          total_count=len(orders),
                          customer_keyword=customer_keyword,
                          tracking_keyword=tracking_keyword,
                          category_filter=category_filter,
                          salesman_filter=salesman_filter,
                          categories=list(orders_by_category.keys()),
                          unread_count=get_unread_count(current_user.id))


@bp.route('/api/export/preview')
@role_required('shipper', 'admin')
def api_export_preview():
    """获取导出预览数据（按模板字段映射）"""
    print(f"[EXPORT_PREVIEW] 收到请求, user={current_user.name}, roles={current_user.get_roles()}")

    customer_keyword = request.args.get('customer_keyword', '').strip()
    tracking_keyword = request.args.get('tracking_keyword', '').strip()
    category_filter = request.args.get('category', '').strip()
    salesman_filter = request.args.get('salesman_id', '').strip()

    # 查询待发货订单（排除已标记导出的）
    query = Order.query.filter(Order.status == 'submitted', Order.export_marked == False)
    if customer_keyword:
        query = query.filter(Order.customer_name.contains(customer_keyword))
    if tracking_keyword:
        query = query.filter(Order.tracking_number.contains(tracking_keyword))
    if category_filter:
        query = query.filter(Order.category == category_filter)
    if salesman_filter:
        query = query.filter(Order.salesman_id == salesman_filter)
    orders = query.order_by(Order.category, Order.create_time.desc()).all()

    if not orders:
        return jsonify({'error': '没有待发货的订单可导出'})

    # 按类别分组
    orders_by_category = defaultdict(list)
    for order in orders:
        orders_by_category[order.category].append(order)

    # 获取模板映射
    categories = Category.query.filter(Category.name.in_(orders_by_category.keys())).all()
    category_id_map = {c.name: c.id for c in categories}
    templates = ExcelTemplate.query.filter(
        ExcelTemplate.category_id.in_(category_id_map.values())
    ).all()
    template_map = {t.category_id: t for t in templates}

    # 构建预览数据
    result = []
    for cat_name, cat_orders in orders_by_category.items():
        cat_id = category_id_map.get(cat_name)
        template = template_map.get(cat_id)

        if not template or not template.field_mapping:
            continue

        mapping = json.loads(template.field_mapping)

        # 表头
        headers = list(mapping.keys())

        # 数据行
        rows = []
        seq = 1
        for order in cat_orders:
            row = []
            for excel_col, order_field in mapping.items():
                if order_field == '__seq__':
                    row.append(seq)
                else:
                    row.append(_get_order_field_value(order, order_field, default_zero=True) or '')
            rows.append(row)
            seq += 1

        result.append({
            'category': cat_name,
            'count': len(cat_orders),
            'headers': headers,
            'rows': rows
        })

    print(f"[EXPORT_PREVIEW] 返回结果: total={len(orders)}, categories={len(result)}")

    return jsonify({
        'total': len(orders),
        'categories': result
    })


@bp.route('/export/orders', methods=['POST'])
@role_required('shipper', 'admin')
def export_orders():
    """导出当前筛选条件下的待发货订单，按类别匹配模板"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    # 获取筛选参数
    customer_keyword = request.form.get('customer_keyword', '').strip()
    tracking_keyword = request.form.get('tracking_keyword', '').strip()
    category_filter = request.form.get('category', '').strip()
    salesman_filter = request.form.get('salesman_id', '').strip()

    # 查询待发货订单（排除已标记导出的）
    query = Order.query.filter(Order.status == 'submitted', Order.export_marked == False)
    if customer_keyword:
        query = query.filter(Order.customer_name.contains(customer_keyword))
    if tracking_keyword:
        query = query.filter(Order.tracking_number.contains(tracking_keyword))
    if category_filter:
        query = query.filter(Order.category == category_filter)
    if salesman_filter:
        query = query.filter(Order.salesman_id == salesman_filter)
    orders = query.order_by(Order.create_time.desc()).all()

    if not orders:
        flash('没有待发货的订单可导出！', 'warning')
        return redirect(request.referrer or url_for('orders.dashboard'))

    # 按类别分组
    orders_by_category = defaultdict(list)
    for order in orders:
        orders_by_category[order.category].append(order)

    # 获取所有相关模板
    categories = Category.query.filter(Category.name.in_(orders_by_category.keys())).all()
    category_id_map = {c.name: c.id for c in categories}
    templates = ExcelTemplate.query.filter(
        ExcelTemplate.category_id.in_(category_id_map.values())
    ).all()
    template_map = {t.category_id: t for t in templates}

    # 检查是否所有类别都有模板
    missing = []
    for cat_name in orders_by_category.keys():
        cat_id = category_id_map.get(cat_name)
        if not cat_id or cat_id not in template_map:
            missing.append(cat_name)
    if missing:
        flash(f'以下类别未配置Excel模板：{", ".join(missing)}，请先在系统配置中上传模板！', 'danger')
        return redirect(request.referrer or url_for('orders.dashboard'))

    # 为每个类别生成Excel
    output_files = []
    for cat_name, cat_orders in orders_by_category.items():
        cat_id = category_id_map[cat_name]
        template = template_map[cat_id]
        mapping = json.loads(template.field_mapping) if template.field_mapping else {}

        if not mapping:
            flash(f'类别 "{cat_name}" 的模板未配置字段映射，请先配置！', 'danger')
            return redirect(request.referrer or url_for('orders.dashboard'))

        # 检查模板文件是否存在
        if not os.path.exists(template.filepath):
            flash(f'类别 "{cat_name}" 的模板文件不存在，请重新上传模板！', 'danger')
            return redirect(request.referrer or url_for('orders.dashboard'))

        # 根据文件格式选择处理方式
        is_xls = template.filepath.lower().endswith('.xls')

        if is_xls:
            import xlrd
            import xlwt
            import xlutils.copy as xlcopy

            # 读取.xls模板
            try:
                rb = xlrd.open_workbook(template.filepath, formatting_info=True)
            except Exception as e:
                flash(f'类别 "{cat_name}" 的模板文件损坏或格式不支持，请检查模板文件！错误：{str(e)}', 'danger')
                return redirect(request.referrer or url_for('orders.dashboard'))
            wb = xlcopy.copy(rb)
            ws = wb.get_sheet(0)

            # 创建居中样式
            center_style = xlwt.XFStyle()
            center_alignment = xlwt.Alignment()
            center_alignment.horz = xlwt.Alignment.HORZ_CENTER
            center_alignment.vert = xlwt.Alignment.VERT_CENTER
            center_style.alignment = center_alignment

            # 找到表头行（第一行），建立列名->列号的映射
            header_map = {}
            sheet0 = rb.sheet_by_index(0)
            for col in range(sheet0.ncols):
                cell_value = sheet0.cell(0, col).value
                if cell_value:
                    header_map[str(cell_value).strip()] = col

            print(f"[DEBUG] 类别 {cat_name} 表头映射: {header_map}")
            print(f"[DEBUG] 配置映射: {mapping}")

            # 从第二行开始填充数据
            data_row = 1
            seq = 1
            for order in cat_orders:
                print(f"[DEBUG] 处理订单: {order.group_name}, 客户: {order.customer_name}")
                for excel_col_name, order_field in mapping.items():
                    col_idx = resolve_col_index(excel_col_name, header_map)
                    print(f"[DEBUG]   列 '{excel_col_name}' -> 索引 {col_idx}, 字段 {order_field}")
                    if col_idx is None or col_idx < 0 or col_idx > 1000:
                        print(f"[WARN]   无效的列索引 {col_idx}，跳过此列")
                        continue
                    # 序号自动递增
                    if order_field == '__seq__':
                        value = seq
                    else:
                        value = _get_order_field_value(order, order_field, default_zero=True)
                    print(f"[DEBUG]   值: {value}")
                    ws.write(data_row, col_idx, value, center_style)
                data_row += 1
                seq += 1

            # 保存
            tmp = tempfile.NamedTemporaryFile(suffix='.xls', delete=False)
            wb.save(tmp.name)
            tmp.close()
            output_files.append((cat_name, tmp.name))
        else:
            import openpyxl
            from openpyxl.styles import Alignment, Font, Border, Side
            from openpyxl import Workbook

            # 读取.xlsx模板
            wb = None
            try:
                wb = openpyxl.load_workbook(template.filepath, data_only=False, keep_links=False)
            except Exception as e:
                print(f"[WARN] 模板加载失败({e})，根据字段映射直接生成Excel")

            if wb is None:
                # WPS兼容模式：XML级别填充，保留原模板格式和表头
                try:
                    from services import fill_wps_template

                    # 构建数据行
                    data_rows = []
                    seq = 1
                    for order in cat_orders:
                        row_data = {}
                        for excel_col_name, order_field in mapping.items():
                            if order_field == '__seq__':
                                row_data[order_field] = seq
                            else:
                                row_data[order_field] = _get_order_field_value(order, order_field, default_zero=True)
                        data_rows.append(row_data)
                        seq += 1

                    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                    tmp.close()
                    result = fill_wps_template(template.filepath, mapping, data_rows, tmp.name)
                    print(f"[DEBUG] fill_wps_template 返回结果: {result}")
                    if not result.get('success'):
                        error_msg = result.get('error', '未知错误')
                        print(f"[ERROR] 模板填充失败: {error_msg}")
                        flash(f'类别 "{cat_name}" 的模板填充失败：{error_msg}', 'danger')
                        return redirect(request.referrer or url_for('orders.dashboard'))
                    # 验证文件大小
                    file_size = os.path.getsize(tmp.name)
                    print(f"[DEBUG] 生成的文件大小: {file_size} bytes")
                    if file_size == 0:
                        flash(f'类别 "{cat_name}" 的导出文件为空', 'danger')
                        return redirect(request.referrer or url_for('orders.dashboard'))
                    output_files.append((cat_name, tmp.name))
                    continue  # 跳过下面的模板填充逻辑
                except Exception as e:
                    flash(f'类别 "{cat_name}" 的模板处理失败：{str(e)}', 'danger')
                    return redirect(request.referrer or url_for('orders.dashboard'))

            ws = wb.active

            # 居中对齐样式
            center_alignment = Alignment(horizontal='center', vertical='center')

            # 找到表头行（第一行），建立列名->列号的映射
            header_map = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    header_map[str(cell_value).strip()] = col

            print(f"[DEBUG] 类别 {cat_name} 表头映射: {header_map}")
            print(f"[DEBUG] 配置映射: {mapping}")
            print(f"[DEBUG] 配置映射的键: {list(mapping.keys())}")

            # 从第二行开始填充数据
            data_row = 2
            seq = 1
            for order in cat_orders:
                print(f"[DEBUG] 处理订单: {order.group_name}, 客户: {order.customer_name}")
                for excel_col_name, order_field in mapping.items():
                    col_idx = resolve_col_index(excel_col_name, header_map)
                    print(f"[DEBUG]   列 '{excel_col_name}' -> 索引 {col_idx}, 字段 {order_field}")
                    if col_idx is None or col_idx <= 0 or col_idx > 1000:
                        print(f"[WARN]   无效的列索引 {col_idx}，跳过此列")
                        continue
                    if order_field == '__seq__':
                        value = seq
                    else:
                        value = _get_order_field_value(order, order_field, default_zero=True)
                    print(f"[DEBUG]   值: {value}")
                    cell = ws.cell(row=data_row, column=col_idx, value=value)
                    cell.alignment = center_alignment
                data_row += 1
                seq += 1

            # 保存
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            wb.save(tmp.name)
            tmp.close()
            output_files.append((cat_name, tmp.name))

    # 先标记订单为已导出（不管是否发送钉钉都标记）
    print(f"[DEBUG] 标记 {len(orders)} 个订单为已导出")
    for order in orders:
        order.export_marked = True
        order.export_mark_time = datetime.now()
    db.session.commit()
    print("[DEBUG] 订单标记完成")

    # 根据当前用户权限决定是否发送钉钉
    can_dingtalk = getattr(current_user, 'can_dingtalk_export', False)
    print(f"[DEBUG] can_dingtalk_export = {can_dingtalk}")

    dingtalk_success = False
    dingtalk_error = None

    if can_dingtalk:
        try:
            from services import send_excel_as_online_sheet

            print(f"[DEBUG] 开始发送钉钉通知，共 {len(output_files)} 个文件...")
            all_success = True
            error_messages = []

            if len(output_files) == 1:
                cat_name, filepath = output_files[0]
                file_display = f'{cat_name}-{datetime.now().strftime("%Y%m%d")}.xlsx'
                print(f"[DEBUG] 发送单个文件: {file_display}")
                result = send_excel_as_online_sheet(filepath, file_display, len(orders))
                print(f"[DEBUG] 发送结果: {result}")

                if result.get('success'):
                    dingtalk_success = True
                else:
                    error_messages.append(result.get('error', '未知错误'))
            else:
                for cat_name, filepath in output_files:
                    file_display = f'{cat_name}-{datetime.now().strftime("%Y%m%d")}.xlsx'
                    print(f"[DEBUG] 发送文件: {file_display}")
                    result = send_excel_as_online_sheet(filepath, file_display, len(orders))
                    print(f"[DEBUG] 发送结果: {result}")

                    if not result.get('success'):
                        all_success = False
                        error_messages.append(f"{cat_name}: {result.get('error', '未知错误')}")

                if all_success:
                    dingtalk_success = True

            if not dingtalk_success:
                dingtalk_error = '; '.join(error_messages)
                print(f"[ERROR] 钉钉发送失败: {dingtalk_error}")

        except Exception as e:
            import traceback
            dingtalk_error = f"{str(e)}\n{traceback.format_exc()}"
            print(f"[ERROR] 钉钉发送异常: {dingtalk_error}")

    # 根据钉钉发送结果设置提示信息
    if can_dingtalk:
        if dingtalk_success:
            flash('导出成功！钉钉消息已发送。', 'success')
        else:
            flash(f'导出成功，但钉钉消息发送失败：{dingtalk_error}', 'warning')
    else:
        flash('导出成功！（您未开启钉钉导出权限，如需发送钉钉请联系管理员）', 'success')

    # 如果只有一个类别，直接下载
    if len(output_files) == 1:
        cat_name, filepath = output_files[0]
        return send_file(filepath, as_attachment=True,
                         download_name=f'{cat_name}-{datetime.now().strftime("%Y%m%d")}.xlsx')

    # 多个类别，打包为zip
    zip_tmp = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
    with zipfile.ZipFile(zip_tmp.name, 'w', zipfile.ZIP_DEFLATED) as zf:
        for cat_name, filepath in output_files:
            zf.write(filepath, f'{cat_name}-{datetime.now().strftime("%Y%m%d")}.xlsx')
            os.remove(filepath)
    zip_tmp.close()
    return send_file(zip_tmp.name, as_attachment=True,
                     download_name=f'待发货订单-{datetime.now().strftime("%Y%m%d")}.zip',
                     mimetype='application/zip')


@bp.route('/order/<int:order_id>/toggle_export_mark', methods=['POST'])
@role_required('shipper', 'admin')
def toggle_export_mark(order_id):
    """切换订单导出标记"""
    order = Order.query.get_or_404(order_id)

    # 权限检查
    if current_user.username != 'admin' and current_user.group_id:
        managed_group_ids = current_user.get_managed_group_ids()
        if order.group_id not in managed_group_ids:
            flash('只能操作本级及下级组的订单！', 'danger')
            return redirect(request.referrer or url_for('orders.dashboard'))

    order.export_marked = not order.export_marked
    order.export_mark_time = datetime.now() if order.export_marked else None
    db.session.commit()

    action = '标记为已导出' if order.export_marked else '取消导出标记'
    flash(f'订单 {order.customer_name} 已{action}！', 'success')
    return redirect(request.referrer or url_for('orders.dashboard'))


@bp.route('/download/<token>')
def download_file(token):
    """临时文件下载接口"""
    from services import get_download_info
    info = get_download_info(token)
    if not info:
        return '下载链接已过期或无效', 404
    return send_file(info['file_path'], as_attachment=True, download_name=info['file_name'])
