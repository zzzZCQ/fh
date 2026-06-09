# -*- coding: utf-8 -*-
"""行为轨迹处理路由"""
import os
import re
from datetime import datetime, timedelta
from flask import Blueprint, request, redirect, url_for, flash, render_template, send_file
from flask_login import current_user
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import func

from helpers import role_required, get_unread_count
from models import db, BehaviorTrackingRecord

bp = Blueprint('behavior', __name__)

# 行为轨迹文件存储目录（每个用户独立）
BEHAVIOR_DATA_DIR = os.path.join(os.path.dirname(__file__), 'behavior_tracking_data')


def get_user_behavior_dir(user_id):
    """获取用户的行为轨迹文件目录"""
    user_dir = os.path.join(BEHAVIOR_DATA_DIR, f'user_{user_id}')
    os.makedirs(user_dir, exist_ok=True)
    return user_dir


def get_processed_dates_file(user_id):
    """获取用户的已处理日期文件路径"""
    user_dir = get_user_behavior_dir(user_id)
    return os.path.join(user_dir, 'processed_dates.txt')


@bp.route('/behavior_tracking')
@role_required('salesman')
def behavior_tracking_page():
    """行为轨迹处理页面"""
    user_dir = get_user_behavior_dir(current_user.id)

    # 获取已处理的日期
    processed_dates = []
    processed_file = get_processed_dates_file(current_user.id)
    if os.path.exists(processed_file):
        with open(processed_file, 'r', encoding='utf-8') as f:
            processed_dates = [line.strip() for line in f if line.strip()]

    # 扫描已上传的文件，带详细信息
    files_info = []
    has_base_file = False
    if os.path.exists(user_dir):
        for f in os.listdir(user_dir):
            if f.endswith('.xlsx') or f.endswith('.xls'):
                if f == '行为轨迹表.xlsx':
                    files_info.append({
                        'name': f,
                        'type': 'base',
                        'date_str': '',
                        'is_processed': False
                    })
                    has_base_file = True
                else:
                    # 提取日期并标准化格式
                    date_tuple = extract_date_from_filename(f)
                    date_str = ''
                    is_processed = False
                    if date_tuple:
                        month, day = date_tuple
                        date_str = f"{month:02d}{day:02d}"
                        is_processed = date_str in processed_dates
                    # 确定文件类型
                    file_type = 'data'
                    if '未完播' in f:
                        file_type = 'incomplete'
                    elif '完播' in f:
                        file_type = 'complete'
                    files_info.append({
                        'name': f,
                        'type': file_type,
                        'date_str': date_str,
                        'is_processed': is_processed
                    })

    return render_template('behavior_tracking.html',
                           files_info=files_info,
                           has_base_file=has_base_file,
                           processed_dates=processed_dates,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/behavior_tracking/upload', methods=['POST'])
@role_required('salesman')
def upload_behavior_files():
    """上传行为轨迹文件（基础表已存在时可跳过）"""
    user_dir = get_user_behavior_dir(current_user.id)

    base_file = request.files.get('base_file')
    data_files = request.files.getlist('data_files')
    folder_files = request.files.getlist('folder_files')
    replace_base = request.form.get('replace_base') == 'on'  # 是否重新上传基础表

    upload_count = 0
    base_uploaded = False

    # 上传基础表（只有勾选重新上传或基础表不存在时才处理）
    has_existing_base = os.path.exists(os.path.join(user_dir, '行为轨迹表.xlsx'))
    
    if base_file and base_file.filename:
        if replace_base or not has_existing_base:
            save_name = '行为轨迹表.xlsx'
            base_file.save(os.path.join(user_dir, save_name))
            upload_count += 1
            base_uploaded = True

    # 上传手动选择的数据文件
    for f in data_files:
        if f and f.filename:
            save_name = f.filename
            f.save(os.path.join(user_dir, save_name))
            upload_count += 1

    # 上传文件夹中选择的数据文件（只保留Excel）
    for f in folder_files:
        if f and f.filename:
            # webkitdirectory会上传所有文件，可能包含路径，只取文件名部分
            fname = os.path.basename(f.filename)
            # 只保留Excel文件
            if fname.lower().endswith('.xlsx') or fname.lower().endswith('.xls'):
                f.save(os.path.join(user_dir, fname))
                upload_count += 1

    if upload_count > 0:
        message = f'上传成功！共 {upload_count} 个文件'
        if base_uploaded:
            message += '（含基础表）'
        flash(message, 'success')
    else:
        # 如果没有上传任何文件，但基础表已存在且勾选了重新上传但没选文件
        if replace_base and not base_file:
            flash('请选择要上传的基础表文件', 'warning')
        else:
            flash('没有选择文件', 'warning')

    return redirect(url_for('behavior.behavior_tracking_page'))


@bp.route('/behavior_tracking/download')
@role_required('salesman')
def download_behavior_result():
    """下载处理后的行为轨迹表"""
    user_dir = get_user_behavior_dir(current_user.id)
    result_file = os.path.join(user_dir, '行为轨迹表.xlsx')

    if not os.path.exists(result_file):
        flash('没有可下载的结果文件', 'warning')
        return redirect(url_for('behavior.behavior_tracking_page'))

    return send_file(result_file, as_attachment=True,
                     download_name=f'行为轨迹表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')


@bp.route('/behavior_tracking/clear_processed')
@role_required('salesman')
def clear_behavior_processed():
    """清空已处理记录"""
    processed_file = get_processed_dates_file(current_user.id)

    if os.path.exists(processed_file):
        os.remove(processed_file)
        flash('已清空处理记录，可以重新处理所有文件', 'success')
    else:
        flash('没有处理记录需要清空', 'info')

    return redirect(url_for('behavior.behavior_tracking_page'))


@bp.route('/behavior_tracking/clear_all')
@role_required('salesman')
def clear_behavior_all():
    """清空所有文件（新一期开始时用）"""
    user_dir = get_user_behavior_dir(current_user.id)
    
    # 清空整个文件夹内所有文件
    import shutil
    if os.path.exists(user_dir):
        shutil.rmtree(user_dir)
        os.makedirs(user_dir, exist_ok=True)
    
    # 同时清空数据库中的行为轨迹记录
    try:
        deleted_count = BehaviorTrackingRecord.query.filter_by(
            user_id=current_user.id
        ).delete(synchronize_session=False)
        db.session.commit()
        flash(f'已清空所有文件和数据库记录（{deleted_count}条），可以开始新一期了', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'清空文件成功，但数据库清空失败：{str(e)}', 'warning')

    return redirect(url_for('behavior.behavior_tracking_page'))


@bp.route('/behavior_tracking/process', methods=['POST'])
@role_required('salesman')
def process_behavior_tracking():
    """处理行为轨迹数据"""
    try:
        result = process_behavior_tracking_data(current_user.id)
        flash(result, 'success')
    except Exception as e:
        flash(f'处理失败：{str(e)}', 'danger')

    return redirect(url_for('behavior.behavior_tracking_page'))


def process_behavior_tracking_data(user_id):
    """核心处理逻辑 - 优化版"""
    user_dir = get_user_behavior_dir(user_id)

    DATA_START_ROW = 5
    TITLE_ROW = 4
    ORIGINAL_FILE = os.path.join(user_dir, "行为轨迹表.xlsx")
    PROCESSED_FILE = get_processed_dates_file(user_id)

    if not os.path.exists(ORIGINAL_FILE):
        raise Exception("找不到 行为轨迹表.xlsx，请先上传基础表")

    # 加载原文件（使用read_only模式提升性能）
    wb = load_workbook(ORIGINAL_FILE, read_only=False)
    sheet = wb.active

    # 识别列结构
    nick_col = None
    col_map = {}
    for col_idx, cell_value in enumerate(sheet[TITLE_ROW], start=1):
        cell_str = str(cell_value.value or "").strip()
        if '昵称' in cell_str:
            nick_col = col_idx
        if cell_str.isdigit():
            col_map[int(cell_str)] = col_idx

    if nick_col is None:
        raise Exception(f"第{TITLE_ROW}行未找到'昵称'列")

    # 优化：批量读取用户列表
    nick_row_map = {}
    empty_rows = []
    max_row = sheet.max_row

    # 使用列表推导式批量处理
    for row_idx in range(DATA_START_ROW, max_row + 1):
        nick_value = str(sheet.cell(row=row_idx, column=nick_col).value or "").strip()
        
        if nick_value == '':
            empty_rows.append(row_idx)
            continue
        
        clean_nick = clean_nickname(nick_value)
        
        if clean_nick in nick_row_map:
            # 标记重复行，后续统一处理
            empty_rows.append(row_idx)
        else:
            nick_row_map[clean_nick] = row_idx

    # 加载已处理日期
    processed_dates = set()
    if os.path.exists(PROCESSED_FILE):
        with open(PROCESSED_FILE, 'r', encoding='utf-8') as f:
            processed_dates = set(line.strip() for line in f if line.strip())

    # 扫描完播文件，先收集所有符合条件的文件信息
    all_candidate_files = []
    
    # 优化：提前编译条件
    processed_dates_set = processed_dates
    for f in os.listdir(user_dir):
        # 支持.xlsx和.xls格式，排除基础表
        if not (f.endswith('.xlsx') or f.endswith('.xls')) or f == '行为轨迹表.xlsx':
            continue

        date_tuple = extract_date_from_filename(f)
        if not date_tuple:
            continue

        month, day = date_tuple
        date_str = f"{month:02d}{day:02d}"
        
        # 检查是否已处理过
        if date_str in processed_dates_set:
            continue
        
        all_candidate_files.append({
            'filename': f,
            'month': month,
            'day': day,
            'date_str': date_str,
            'date_obj': parse_month_day(month, day)
        })

    if not all_candidate_files:
        return "未找到新的有效数据文件"
    
    # 第一阶段：先找到所有可用的列映射中最小的标题数字（1,2,3...）
    sorted_available_titles = sorted(col_map.keys())
    if not sorted_available_titles:
        return "行为轨迹表中没有找到有效的日期列"
    
    min_title_num = sorted_available_titles[0]
    ref_date = sorted(all_candidate_files, key=lambda x: x['date_obj'])[0]['date_obj']
    ref_title_num = min_title_num
    
    # 第四阶段：按日期分组处理所有文件
    date_groups = {}
    for file_info in all_candidate_files:
        date_groups.setdefault(file_info['date_str'], []).append(file_info)
    
    all_user_data = {}
    new_processed_dates = []
    target_titles = []
    
    # 优化：预定义颜色对象，避免重复创建
    fill_green = PatternFill(start_color='FF32CD32', end_color='FF32CD32', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    fill_red = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    color_map = {1: fill_green, 2: fill_yellow, 3: fill_red}
    
    for date_str in sorted(date_groups.keys()):
        date_files = date_groups[date_str]
        # 排序：完播在前，未完播在后
        date_files.sort(key=lambda x: 0 if '完播' in x['filename'] and '未' not in x['filename'] else 1)
        target_date = date_files[0]['date_obj']
        
        delta_days = (target_date - ref_date).days
        target_title_num = ref_title_num + delta_days
        
        if target_title_num not in col_map:
            continue
        
        # 读取完播和未完播用户名单
        complete_users = set()
        incomplete_users = set()
        for file_info in date_files:
            f = file_info['filename']
            data_file_path = os.path.join(user_dir, f)
            users = []
            try:
                users = read_user_list_from_file(data_file_path)
            except Exception as e:
                try:
                    if data_file_path.endswith('.xlsx'):
                        import xlrd
                        wb_temp = xlrd.open_workbook(data_file_path)
                        sheet_temp = wb_temp.sheet_by_index(0)
                        users = [clean_nickname(sheet_temp.cell(row_idx, 0).value) 
                                for row_idx in range(sheet_temp.nrows) 
                                if sheet_temp.cell(row_idx, 0).value is not None 
                                and str(sheet_temp.cell(row_idx, 0).value).strip() != '']
                    else:
                        wb_temp = load_workbook(data_file_path, read_only=True, data_only=True)
                        sheet_temp = wb_temp.active
                        users = [clean_nickname(row[0]) 
                                for row in sheet_temp.iter_rows(min_row=1, max_col=1, values_only=True) 
                                if row[0] is not None and str(row[0]).strip() != '']
                        wb_temp.close()
                except Exception:
                    continue
            
            if '未完播' in f:
                incomplete_users.update(users)
            else:
                complete_users.update(users)
        
        # 填充数据（完播优先）
        user_data_for_date = all_user_data
        target_num = target_title_num
        
        # 优化：使用集合操作代替逐个判断
        for user in complete_users:
            user_data_for_date.setdefault(user, {})[target_num] = 1
        
        # 只处理不在完播名单中的未完播用户
        for user in incomplete_users - complete_users:
            user_data_for_date.setdefault(user, {}).setdefault(target_num, 2)
        
        new_processed_dates.append(date_str)
        target_titles.append(target_title_num)

    if not target_titles:
        return "未找到新的有效数据文件"

    # 处理新增用户
    existing_users = set(nick_row_map.keys())
    new_users = all_user_data.keys() - existing_users
    
    for new_user in new_users:
        if empty_rows:
            new_row = empty_rows.pop(0)
        else:
            max_row += 1
            new_row = max_row
        sheet.cell(row=new_row, column=nick_col, value=new_user)
        nick_row_map[new_user] = new_row
    
    # 优化：批量预加载所有目标日期的通话记录
    all_call_records = {}
    for title_num in sorted(target_titles):
        target_date = ref_date + timedelta(days=title_num - ref_title_num)
        
        date_records = {}
        try:
            from wework_partition import get_partition_model
            PartitionModel = get_partition_model(target_date)
            records = PartitionModel.query.filter_by(uploader_id=user_id).all()
            
            # 优化：使用字典推导式
            for rec in records:
                clean_name = clean_nickname(rec.user_name)
                if clean_name:
                    date_records[clean_name] = date_records.get(clean_name, 0) + (rec.call_duration_seconds or 0)
        
        except Exception:
            pass
        
        all_call_records[title_num] = date_records
    
    # 第一步：排序用户（优化：向量化计算）
    user_rows = []
    for nick, row_idx in nick_row_map.items():
        play_count = 0
        for col_num in col_map.keys():
            cell_value = sheet.cell(row=row_idx, column=col_map[col_num]).value
            fill_num = extract_fill_value(cell_value)
            if fill_num == 1:
                play_count += 1
        user_rows.append((-play_count, nick, play_count))  # 负号用于升序排序
    
    # 优化：使用key排序
    user_rows.sort()
    user_rows_sorted = [{'play_count': p, 'nick': n} for _, n, p in user_rows]
    
    # 更新 nick_row_map，按排序后的顺序排列
    nick_order_map = {ui['nick']: DATA_START_ROW + idx for idx, ui in enumerate(user_rows_sorted)}
    
    # 计算需要清空的最大行号
    max_data_row = DATA_START_ROW + len(user_rows_sorted) - 1
    actual_max_row = max(max_row, max_data_row)
    
    # 优化：批量清空列
    sorted_target_titles = sorted(target_titles)
    for title_num in sorted_target_titles:
        col_idx = col_map[title_num]
        for row_idx in range(DATA_START_ROW, actual_max_row + 1):
            sheet.cell(row=row_idx, column=col_idx, value=None)
    
    # 先填充昵称列
    for user_info in user_rows_sorted:
        sheet.cell(row=nick_order_map[user_info['nick']], column=nick_col, value=user_info['nick'])
    
    # 优化：预计算日期到通话时长的映射
    title_to_date_call = {t: all_call_records.get(t, {}) for t in sorted_target_titles}
    
    # 填充数据（优化：减少嵌套循环中的查找次数）
    for title_num in sorted_target_titles:
        col_idx = col_map[title_num]
        date_call_data = title_to_date_call[title_num]
        
        for user_info in user_rows_sorted:
            nick = user_info['nick']
            row_idx = nick_order_map[nick]
            old_cell = sheet.cell(row=row_idx, column=col_idx)
            
            fill_value = all_user_data.get(nick, {}).get(title_num, 3)
            call_duration = date_call_data.get(nick, 0)
            
            if call_duration > 0:
                final_value = f"{fill_value}（{format_call_duration(call_duration)}）"
            else:
                final_value = str(fill_value)
            
            old_cell.value = final_value
            old_cell.fill = color_map[fill_value]

    # 保存已处理日期
    processed_dates.update(new_processed_dates)
    with open(PROCESSED_FILE, 'w', encoding='utf-8') as f:
        for d in sorted(processed_dates):
            f.write(f"{d}\n")

    # 优化：数据库批量操作
    try:
        # 先删除该用户目标日期的旧数据（批量删除）
        delete_dates = []
        for title_num in sorted_target_titles:
            target_date = ref_date + timedelta(days=title_num - ref_title_num)
            delete_dates.append((target_date.month, target_date.day))
        
        if delete_dates:
            from sqlalchemy import or_
            conditions = []
            for month, day in delete_dates:
                conditions.append(
                    (BehaviorTrackingRecord.month == month) & 
                    (BehaviorTrackingRecord.day == day)
                )
            BehaviorTrackingRecord.query.filter(
                BehaviorTrackingRecord.user_id == user_id,
                or_(*conditions)
            ).delete(synchronize_session=False)
        
        # 批量插入新数据
        records_to_insert = []
        for user_info in user_rows_sorted:
            nick = user_info['nick']
            play_count = user_info['play_count']
            
            for title_num in sorted_target_titles:
                target_date = ref_date + timedelta(days=title_num - ref_title_num)
                fill_value = all_user_data.get(nick, {}).get(title_num, 3)
                call_duration = all_call_records.get(title_num, {}).get(nick, 0)
                
                records_to_insert.append({
                    'user_id': user_id,
                    'nickname': nick,
                    'month': target_date.month,
                    'day': target_date.day,
                    'play_status': fill_value,
                    'call_duration_seconds': call_duration,
                    'play_order': play_count
                })
        
        if records_to_insert:
            db.session.bulk_insert_mappings(BehaviorTrackingRecord, records_to_insert)
            db.session.commit()
    
    except Exception as db_err:
        db.session.rollback()

    # 保存文件
    wb.save(ORIGINAL_FILE)
    wb.close()

    return f"处理完成！填充了 {len(target_titles)} 个日期列，共 {len(user_rows_sorted)} 个用户，已按完播次数降序排序"


def clean_nickname(name):
    """昵称清洗 - 与企业微信通话记录保持一致的规则（优化@微信等处理）"""
    if not name:
        return ''
    if not isinstance(name, str):
        name = str(name)
    
    text = name.strip()
    
    # 第一步：处理换行符，只保留第一行有效内容
    # 例如："秋天\n@微信" -> "秋天"
    if '\n' in text:
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if line and '@' not in line and '微信' not in line:
                text = line
                break
    
    # 第二步：处理@微信相关后缀
    # 例如："秋天@微信" -> "秋天", "秋天@" -> "秋天"
    if '@' in text:
        text = text.split('@')[0].strip()
    
    # 第三步：移除排除词
    exclude_words = ['语音通话', '视频通话', '正在呼叫', '企业微信', '微信', '通话', '接通', '等待', '结束', '取消', '的', '和', '与', '正在']
    for word in exclude_words:
        text = text.replace(word, '')
    
    # 第四步：处理"完"前缀
    # 例如：0522完清酒 -> 清酒, SB0211完苗先生 -> 苗先生
    prefix_pattern = r'(?:[A-Za-z]*\d+[A-Za-z]*完)?(.*)'
    match = re.search(prefix_pattern, text)
    if match:
        result_name = match.group(1).strip()
        # 清理可能残留的"完"字
        result_name = result_name.replace('完', '')
        if result_name and len(result_name) >= 2:
            text = result_name
    
    # 第五步：清理特殊字符，只保留中文/数字/字母
    text = re.sub(r'[^\w\u4e00-\u9fa5]', '', text)
    
    # 第六步：移除日期格式
    text = re.sub(r'[0-9]{1,2}[./-月]?[0-9]{1,2}', '', text)
    
    # 第七步：移除前缀
    prefixes = ['新-', '白-', 'SW-', 'AA-', '不要-', '白嫖-', 'SW', 'AA', '不要', '白嫖', '新', '白']
    for p in prefixes:
        if text.startswith(p):
            text = text[len(p):]
    
    # 第八步：再次移除完字前缀（以防万一）
    if '完' in text:
        text = text[text.index('完')+1:]
    
    # 第九步：移除括号内容
    text = re.sub(r'\(.*?\)', '', text)
    text = re.sub(r'（.*?）', '', text)
    
    return text.strip()


def extract_date_from_filename(filename):
    """从文件名提取日期"""
    try:
        # 先去掉文件扩展名
        name_without_ext = os.path.splitext(filename)[0]
        
        if '未完播' in name_without_ext:
            prefix = name_without_ext.split('未完播')[0].strip()
        elif '完播' in name_without_ext:
            prefix = name_without_ext.split('完播')[0].strip()
        else:
            return None
        
        # 从prefix中提取数字部分（处理前缀中可能含有的非数字字符）
        import re
        digits_match = re.search(r'(\d+)', prefix)
        if not digits_match:
            return None
        prefix = digits_match.group(1)
        
        if prefix.isdigit():
            if len(prefix) == 3:
                return (int(prefix[0]), int(prefix[1:]))
            elif len(prefix) == 4:
                return (int(prefix[:2]), int(prefix[2:]))
            elif len(prefix) == 2:
                # 也支持只有2位数字的情况（比如24表示5月24日？这里我们按默认月份处理）
                # 不过先保持严格一点，只处理3位或4位
                pass
        return None
    except Exception as e:
        print(f"[日期提取] 失败: {filename}, {e}")
        return None


@bp.route('/behavior_tracking/result')
@role_required('salesman')
def behavior_tracking_result():
    """行为轨迹处理结果展示页面（从数据库读取）"""
    user_id = current_user.id
    
    # 获取所有已处理的日期
    processed_dates = BehaviorTrackingRecord.query.filter_by(user_id=user_id) \
        .with_entities(BehaviorTrackingRecord.month, BehaviorTrackingRecord.day) \
        .distinct().all()
    
    if not processed_dates:
        flash('请先上传并处理行为轨迹表！', 'warning')
        return redirect(url_for('behavior.behavior_tracking_page'))
    
    # 整理日期列表并排序
    date_list = sorted(set((d.month, d.day) for d in processed_dates))
    
    # 获取所有昵称及其完播次数（用于排序）
    nickname_counts = db.session.query(
        BehaviorTrackingRecord.nickname,
        db.func.count(db.case((BehaviorTrackingRecord.play_status == 1, 1))).label('play_count')
    ).filter(
        BehaviorTrackingRecord.user_id == user_id
    ).group_by(BehaviorTrackingRecord.nickname).all()
    
    # 按完播次数降序排序
    sorted_nicknames = sorted(nickname_counts, key=lambda x: (-x[1], x[0]))
    
    # 构建表头
    headers = ['序号', '昵称'] + [f"{m:02d}{d:02d}" for m, d in date_list]
    
    # 构建数据
    rows_data = []
    for idx, (nickname, _) in enumerate(sorted_nicknames, 1):
        row = [{'value': idx}, {'value': nickname}]
        
        for month, day in date_list:
            record = BehaviorTrackingRecord.query.filter_by(
                user_id=user_id,
                nickname=nickname,
                month=month,
                day=day
            ).first()
            
            if record:
                # 通话时长
                call_str = f"（{format_call_duration(record.call_duration_seconds)}）" if record.call_duration_seconds > 0 else ""
                value = f"{record.play_status}{call_str}"
                
                # 颜色：1=绿色(#32cd32), 2=黄色, 3=红色
                color_map = {1: 'FF32CD32', 2: 'FFFFFF00', 3: 'FFFF0000'}
                fill_color = color_map.get(record.play_status, 'FFFFFFFF')
                
                row.append({
                    'value': value,
                    'fill_color': fill_color
                })
            else:
                row.append({'value': '', 'fill_color': None})
        
        rows_data.append(row)
    
    return render_template('behavior_tracking_result.html',
                           headers=headers,
                           rows_data=rows_data,
                           date_cols=[],
                           unread_count=get_unread_count(current_user.id))


def parse_month_day(month, day):
    """解析月日为日期对象（当前年份）"""
    from datetime import datetime
    current_year = datetime.now().year
    return datetime(current_year, month, day).date()





def read_user_list_from_file(file_path):
    """从文件读取用户列表（支持.xlsx和.xls）"""
    users = []

    if file_path.endswith('.xls') and not file_path.endswith('.xlsx'):
        # 使用xlrd读取.xls文件
        import xlrd
        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)
        for row_idx in range(sheet.nrows):
            cell_value = sheet.cell(row_idx, 0).value
            if cell_value is not None and str(cell_value).strip() != '':
                users.append(clean_nickname(cell_value))
    else:
        # 使用openpyxl读取.xlsx文件
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0] is not None and str(row[0]).strip() != '':
                users.append(clean_nickname(row[0]))
        wb.close()

    return users


def extract_bracket_content(text):
    """提取括号内容"""
    if not text:
        return ""
    text_str = str(text).strip()
    cn_bracket = re.search(r'（.*?）', text_str)
    if cn_bracket:
        return cn_bracket.group(0)
    en_bracket = re.search(r'\(.*?\)', text_str)
    if en_bracket:
        return en_bracket.group(0)
    return ""


def extract_fill_value(cell_value):
    """提取填充值"""
    if not cell_value:
        return None
    cell_str = str(cell_value).strip()
    num_match = re.match(r'^(\d+)', cell_str)
    if num_match:
        return int(num_match.group(1))
    return None


def format_call_duration(seconds):
    """格式化通话时长显示"""
    if not seconds:
        return ''
    seconds = int(seconds)
    if seconds < 60:
        return f"{seconds}秒"
    minutes, secs = divmod(seconds, 60)
    if minutes < 60:
        return f"{minutes}分{secs}秒"
    hours, mins = divmod(minutes, 60)
    return f"{hours}时{mins}分{secs}秒"


@bp.route('/admin/customer_tracking/save_stats', methods=['POST'])
def save_salesman_stats():
    """保存业务员统计数据（管理员可编辑本级及以下的所有人数据，业务员只能编辑自己的数据）"""
    from models import SalesmanDailyStats, User
    from datetime import datetime
    
    # 权限检查：必须是业务员或管理员
    if not current_user.has_role('admin') and not current_user.has_role('salesman'):
        return {'success': False, 'message': '没有权限'}
    
    data = request.get_json()
    user_id = data.get('user_id')
    date_str = data.get('date')
    total_incoming = data.get('total_incoming', 0)
    touched_count = data.get('touched_count', 0)
    activated_count = data.get('activated_count', 0)
    completed_count = data.get('completed_count', 0)
    note = data.get('note', '')
    
    # 获取目标用户
    target_user = User.query.get(user_id)
    if not target_user:
        return {'success': False, 'message': '目标用户不存在'}
    
    # 权限检查
    if current_user.has_role('admin'):
        # 管理员：检查目标用户是否在自己的管理范围内（本级及以下组别）
        managed_group_ids = current_user.get_managed_group_ids()
        if target_user.group_id not in managed_group_ids:
            return {'success': False, 'message': '只能编辑本级及以下组别的数据'}
    else:
        # 业务员：只能编辑自己的数据
        if str(user_id) != str(current_user.id):
            return {'success': False, 'message': '只能编辑自己的数据'}
    
    try:
        stat_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return {'success': False, 'message': '日期格式错误'}
    
    # 查找或创建记录
    stat = SalesmanDailyStats.query.filter_by(user_id=user_id, date=stat_date).first()
    if not stat:
        stat = SalesmanDailyStats(user_id=user_id, date=stat_date)
    
    stat.total_incoming = int(total_incoming)
    stat.touched_count = int(touched_count)
    stat.activated_count = int(activated_count)
    stat.completed_count = int(completed_count)
    stat.note = note
    
    db.session.add(stat)
    db.session.commit()
    
    # 计算百分比
    total = int(total_incoming) if total_incoming else 0
    touched_rate = round((int(touched_count) / total) * 100, 1) if total > 0 else 0
    activated_rate = round((int(activated_count) / total) * 100, 1) if total > 0 else 0
    completed_rate = round((int(completed_count) / total) * 100, 1) if total > 0 else 0
    
    return {
        'success': True,
        'touched_rate': touched_rate,
        'activated_rate': activated_rate,
        'completed_rate': completed_rate
    }


@bp.route('/admin/customer_tracking/batch_save_stats', methods=['POST'])
def batch_save_stats():
    """批量保存业务员统计数据（触达数和完播数从客户明细数据汇总，不允许手动修改）"""
    from models import User, SalesmanDailyStats, BehaviorTrackingRecord
    from datetime import datetime
    data = request.get_json()
    stats_list = data.get('stats', [])

    updated_stats = []

    # 管理员预获取管理的组别ID
    managed_group_ids = []
    if current_user.has_role('admin'):
        managed_group_ids = current_user.get_managed_group_ids()

    for stat_data in stats_list:
        user_id = stat_data.get('user_id')
        date_str = stat_data.get('date')
        total_incoming = stat_data.get('total_incoming', 0)
        activated_count = stat_data.get('activated_count', 0)
        # 获取前端传来的touched_count和completed_count（如果有的话）
        touched_count_input = stat_data.get('touched_count')
        completed_count_input = stat_data.get('completed_count')
        note = stat_data.get('note', '')

        # 权限检查
        can_edit = False
        if current_user.has_role('admin'):
            # 管理员：检查是否在管理组别范围内
            target_user = User.query.get(user_id)
            if target_user and target_user.group_id in managed_group_ids:
                can_edit = True
        elif current_user.has_role('salesman') and int(user_id) == current_user.id:
            # 业务员：只能编辑自己
            can_edit = True

        if not can_edit:
            continue

        try:
            stat_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            continue

        # 查找现有记录
        stat = SalesmanDailyStats.query.filter_by(user_id=user_id, date=stat_date).first()
        has_existing_data = False
        if stat and (stat.total_incoming > 0 or stat.activated_count > 0):
            has_existing_data = True

        # 从客户明细数据汇总触达数和完播数（作为默认值）
        records = BehaviorTrackingRecord.query.filter_by(
            user_id=user_id,
            month=stat_date.month,
            day=stat_date.day
        ).all()

        touched_count = 0
        completed_count = 0
        customer_records = {}
        for record in records:
            key = record.nickname
            if key not in customer_records:
                customer_records[key] = {'has_touched': False, 'has_completed': False}
            if record.call_duration_seconds > 0:
                customer_records[key]['has_touched'] = True
            if record.play_status == 1:
                customer_records[key]['has_completed'] = True

        # 统计总数
        touched_count = sum(1 for r in customer_records.values() if r['has_touched'])
        completed_count = sum(1 for r in customer_records.values() if r['has_completed'])

        # 确定最终使用的touched_count和completed_count
        if has_existing_data:
            # 已有数据：优先使用原数据，如果前端有传则用前端的
            if touched_count_input is not None:
                touched_count = int(touched_count_input)
            elif stat:
                touched_count = stat.touched_count
            if completed_count_input is not None:
                completed_count = int(completed_count_input)
            elif stat:
                completed_count = stat.completed_count
        else:
            # 没有数据：使用客户明细汇总，但如果前端有传也可以覆盖
            if touched_count_input is not None:
                touched_count = int(touched_count_input)
            if completed_count_input is not None:
                completed_count = int(completed_count_input)

        # 查找或创建记录
        if not stat:
            stat = SalesmanDailyStats(user_id=user_id, date=stat_date)

        stat.total_incoming = int(total_incoming)
        stat.touched_count = touched_count
        stat.activated_count = int(activated_count)
        stat.completed_count = completed_count
        stat.note = note
        
        db.session.add(stat)
        
        # 计算百分比
        total = int(total_incoming) if total_incoming else 0
        touched_rate = round((touched_count / total) * 100, 1) if total > 0 else 0
        activated_rate = round((int(activated_count) / total) * 100, 1) if total > 0 else 0
        completed_rate = round((completed_count / total) * 100, 1) if total > 0 else 0
        
        updated_stats.append({
            'user_id': user_id,
            'touched_rate': touched_rate,
            'activated_rate': activated_rate,
            'completed_rate': completed_rate,
            'touched_count': touched_count,
            'completed_count': completed_count
        })
    
    db.session.commit()
    
    return {
        'success': True,
        'updated_stats': updated_stats
    }


@bp.route('/admin/customer_tracking')
def customer_tracking():
    """客户跟踪页面 - 查看业务员的行为轨迹数据"""
    from models import User, Group, SalesmanDailyStats, CustomerInfo
    from datetime import date
    
    # 获取筛选参数
    is_salesman_only = current_user.has_role('salesman') and not current_user.has_role('admin')
    
    # 业务员默认选择自己的组别和自己
    if is_salesman_only:
        group_id = request.args.get('group_id', current_user.group_id, type=int)
        salesman_id = request.args.get('salesman_id', current_user.id, type=int)
    else:
        group_id = request.args.get('group_id', type=int)
        salesman_id = request.args.get('salesman_id', type=int)
    
    nickname = request.args.get('nickname', '', type=str)
    is_followed_filter = request.args.get('is_followed', '', type=str)
    
    # 获取当前用户的组别及其子组（用于数据权限控制：本级及以下）
    def get_user_group_hierarchy(user):
        if user.group:
            return user.group.get_all_children_ids() + [user.group_id]
        return []
    
    user_groups = get_user_group_hierarchy(current_user)
    
    # 获取组别列表（数据权限：本级及以下）
    if user_groups:
        groups = Group.query.filter(Group.id.in_(user_groups)).all()
    else:
        groups = []
    
    # 获取业务员列表（数据权限控制）
    if is_salesman_only:
        # 业务员只能看到自己
        salesmen = [current_user]
    elif group_id:
        group = Group.query.get(group_id)
        if group:
            child_groups = group.get_all_children_ids() + [group_id]
            salesmen = User.query.filter(User.group_id.in_(child_groups), User.roles.like('%salesman%')).all()
        else:
            salesmen = []
    else:
        # 默认：只能看到本级及以下的业务员
        if user_groups:
            salesmen = User.query.filter(User.group_id.in_(user_groups), User.roles.like('%salesman%')).all()
        else:
            salesmen = []
    
    # 获取可选的日期（根据数据权限过滤）
    dates_query = db.session.query(
        BehaviorTrackingRecord.month,
        BehaviorTrackingRecord.day
    ).distinct()
    
    # 如果是业务员，只能看到自己的数据
    if is_salesman_only:
        dates_query = dates_query.filter(BehaviorTrackingRecord.user_id == current_user.id)
    elif salesman_id:
        dates_query = dates_query.filter(BehaviorTrackingRecord.user_id == salesman_id)
    elif group_id:
        group = Group.query.get(group_id)
        if group:
            child_groups = group.get_all_children_ids() + [group_id]
            dates_query = dates_query.join(User).filter(User.group_id.in_(child_groups))
    elif user_groups:
        dates_query = dates_query.join(User).filter(User.group_id.in_(user_groups))
    
    dates = dates_query.order_by(BehaviorTrackingRecord.month, BehaviorTrackingRecord.day).all()
    
    # 根据筛选条件查询数据
    query = BehaviorTrackingRecord.query
    
    # 数据权限：本级及以下
    if user_groups:
        query = query.join(User).filter(User.group_id.in_(user_groups))
    
    # 如果是业务员，只能看到自己的数据
    if is_salesman_only:
        query = query.filter(BehaviorTrackingRecord.user_id == current_user.id)
    else:
        # 管理员必须选择业务员或组别才能显示数据
        if salesman_id:
            query = query.filter(BehaviorTrackingRecord.user_id == salesman_id)
        elif group_id:
            group = Group.query.get(group_id)
            if group:
                child_groups = group.get_all_children_ids() + [group_id]
                query = query.filter(User.group_id.in_(child_groups))
    
    # 判断是否应该显示数据
    show_data = False
    if is_salesman_only:
        # 业务员始终显示数据
        show_data = True
    else:
        # 管理员必须选择业务员或组别
        if salesman_id or group_id:
            show_data = True
    
    records = []
    if show_data:
        records = query.all()
    
    # 构建数据结构（按客户维度）并计算完播天数
    # 先收集所有user_id
    user_ids = list(set(record.user_id for record in records))
    # 批量查询客户信息 - 使用清洗后的昵称作为匹配键
    customer_infos = {}
    for c in CustomerInfo.query.filter(CustomerInfo.user_id.in_(user_ids)).all():
        # 存储时同时保存原始键和清洗后的键
        customer_infos[(c.user_id, c.nickname)] = c
        # 额外添加一个使用清洗后昵称的键
        clean_c_nick = clean_nickname(c.nickname)
        customer_infos[(c.user_id, f'__clean__{clean_c_nick}')] = c
    
    data = {}
    for record in records:
        key = (record.user_id, record.nickname)
        customer_info = customer_infos.get(key)
        
        # 如果原始键没找到，尝试用清洗后的昵称匹配
        if not customer_info:
            clean_record_nick = clean_nickname(record.nickname)
            customer_info = customer_infos.get((record.user_id, f'__clean__{clean_record_nick}'))
        
        if key not in data:
            user = User.query.get(record.user_id)
            data[key] = {
                'user_id': record.user_id,
                'salesman_name': user.name if user else '',
                'group_name': user.group.name if user and user.group else '',
                'nickname': record.nickname,
                'dates': {},
                'completed_days': 0,
                'customer_name': customer_info.customer_name if customer_info else '',
                'gender': customer_info.gender if customer_info else '',
                'age': customer_info.age if customer_info else 0,
                'phone': customer_info.phone if customer_info else '',
                'address': customer_info.address if customer_info else '',
                'health_condition': customer_info.health_condition if customer_info else '',
                'medication_status': customer_info.medication_status if customer_info else '',
                'is_followed': customer_info is not None and (
                    (customer_info.customer_name or customer_info.gender or 
                     customer_info.age > 0 or customer_info.phone or 
                     customer_info.address or customer_info.health_condition or 
                     customer_info.medication_status)
                )
            }
        date_key = f"{record.month:02d}{record.day:02d}"
        # 处理旧数据迁移：play_status=4 表示拒接，需要转换
        play_status = record.play_status
        is_rejected = record.is_rejected if hasattr(record, 'is_rejected') else False
        if play_status == 4:
            play_status = 0
            is_rejected = True
        
        data[key]['dates'][date_key] = {
            'play_status': play_status,
            'is_rejected': is_rejected,
            'call_duration': record.call_duration_seconds
        }
        if play_status == 1:
            data[key]['completed_days'] += 1
    
    # 按完播天数降序排序
    rows_data = sorted(data.values(), key=lambda x: -x['completed_days'])
    
    # 按客户昵称筛选
    if nickname:
        rows_data = [row for row in rows_data if nickname in row['nickname']]
    
    # 按是否已跟进筛选（只要填过任何字段都算已跟进）
    if is_followed_filter == '1':
        rows_data = [row for row in rows_data if row['is_followed']]
    elif is_followed_filter == '0':
        rows_data = [row for row in rows_data if not row['is_followed']]
    
    date_list = [f"{m:02d}{d:02d}" for m, d in dates]
    
    # 自动往后扩展一天的日期
    if dates:
        # 获取最后一个日期
        last_month, last_day = dates[-1]
        
        # 计算下一个日期
        from datetime import datetime, timedelta
        try:
            # 假设是当前年份
            current_year = datetime.now().year
            last_date = datetime(current_year, last_month, last_day)
            next_date = last_date + timedelta(days=1)
            
            # 添加上下一个日期
            dates.append((next_date.month, next_date.day))
            date_list.append(f"{next_date.month:02d}{next_date.day:02d}")
        except:
            # 如果计算失败，简单地假设 +1 天
            if last_day < 28:
                dates.append((last_month, last_day + 1))
                date_list.append(f"{last_month:02d}{last_day + 1:02d}")
    
    # 分页处理
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    total_records = len(rows_data)
    
    # 验证page_size
    valid_page_sizes = [20, 50, 100, 500]
    if page_size not in valid_page_sizes:
        page_size = 20
    
    # 计算分页
    start_index = (page - 1) * page_size
    end_index = start_index + page_size
    paged_rows_data = rows_data[start_index:end_index]
    total_pages = (total_records + page_size - 1) // page_size
    
    # 获取统计日期
    stats_date = request.args.get('stats_date')
    if stats_date:
        try:
            stats_date = date.fromisoformat(stats_date)
        except ValueError:
            stats_date = date.today()
    else:
        stats_date = date.today()
    
    # 将日期转换为MMDD格式用于匹配客户跟踪明细数据
    stats_date_str = f"{stats_date.month:02d}{stats_date.day:02d}"
    
    # 从客户明细数据中汇总统计信息
    # 统计每个业务员的：
    # - 触达数：当天有通话记录的客户数
    # - 完播数：当天完播的客户数
    salesmen_summary = {}
    for row in rows_data:
        user_id = row['user_id']
        if user_id not in salesmen_summary:
            salesmen_summary[user_id] = {
                'touched_count': 0,
                'completed_count': 0
            }
        # 检查统计日期是否有记录
        if stats_date_str in row['dates']:
            record = row['dates'][stats_date_str]
            # 有通话记录算触达
            if record.get('call_duration', 0) > 0:
                salesmen_summary[user_id]['touched_count'] += 1
            # 完播算完播数
            if record.get('play_status') == 1:
                salesmen_summary[user_id]['completed_count'] += 1
    
    # 获取业务员统计数据（从手动填写的表中读取，并用汇总数据覆盖）
    salesman_stats_list = []
    
    # 预获取管理员的管理组别ID
    managed_group_ids = []
    if current_user.has_role('admin'):
        managed_group_ids = current_user.get_managed_group_ids()
    
    # 获取当前筛选范围内的业务员（数据权限：本级及以下）
    if salesman_id:
        target_salesmen = [User.query.get(salesman_id)]
    elif group_id:
        group = Group.query.get(group_id)
        if group:
            child_groups = group.get_all_children_ids() + [group_id]
            target_salesmen = User.query.filter(User.group_id.in_(child_groups), User.roles.like('%salesman%')).all()
        else:
            target_salesmen = []
    else:
        # 默认：只能看到本级及以下的业务员
        if user_groups:
            target_salesmen = User.query.filter(User.group_id.in_(user_groups), User.roles.like('%salesman%')).all()
        else:
            target_salesmen = []
    
    # 如果是业务员，只能看到自己的数据
    if is_salesman_only:
        target_salesmen = [current_user]
    
    for salesman in target_salesmen:
        # 获取指定日期的统计记录，没有则创建默认记录
        stat = SalesmanDailyStats.query.filter_by(user_id=salesman.id, date=stats_date).first()
        if not stat:
            # 新记录：从客户明细数据汇总初始化
            summary = salesmen_summary.get(salesman.id, {'touched_count': 0, 'completed_count': 0})
            stat = SalesmanDailyStats(
                user_id=salesman.id,
                date=stats_date,
                total_incoming=0,
                touched_count=summary['touched_count'],
                activated_count=0,
                completed_count=summary['completed_count']
            )
        else:
            # 已存在记录：检查是否填写过数据（有total_incoming就算填写过）
            # 如果填写过数据，保留原有的touched_count和completed_count
            # 如果没有填写过数据，用客户明细汇总数据填充
            if stat.total_incoming == 0 and stat.activated_count == 0:
                # 完全没有填写过任何数据，用客户明细汇总
                summary = salesmen_summary.get(salesman.id, {'touched_count': 0, 'completed_count': 0})
                stat.touched_count = summary['touched_count']
                stat.completed_count = summary['completed_count']
        
        # 计算百分比
        total = stat.total_incoming or 0
        touched_rate = round((stat.touched_count / total) * 100, 1) if total > 0 else 0
        activated_rate = round((stat.activated_count / total) * 100, 1) if total > 0 else 0
        completed_rate = round((stat.completed_count / total) * 100, 1) if total > 0 else 0
        
        # 判断是否可以编辑：管理员可以编辑本级及以下的所有人，业务员只能编辑自己
        can_edit = False
        if current_user.has_role('admin'):
            # 管理员：检查是否在管理组别范围内
            if salesman.group_id in managed_group_ids:
                can_edit = True
        elif current_user.has_role('salesman') and salesman.id == current_user.id:
            # 业务员：只能编辑自己
            can_edit = True
        
        salesman_stats_list.append({
            'user_id': salesman.id,
            'salesman_name': salesman.name,
            'group_name': salesman.group.name if salesman.group else '',
            'date_str': stats_date.strftime('%Y-%m-%d'),
            'total_incoming': stat.total_incoming,
            'touched_count': stat.touched_count,
            'touched_rate': touched_rate,
            'activated_count': stat.activated_count,
            'activated_rate': activated_rate,
            'completed_count': stat.completed_count,
            'completed_rate': completed_rate,
            'note': stat.note or '',
            'can_edit': can_edit
        })
    
    salesman_stats_list.sort(key=lambda x: x['salesman_name'])
    
    # 计算汇总数据
    total_incoming_sum = sum(s['total_incoming'] for s in salesman_stats_list)
    total_touched_sum = sum(s['touched_count'] for s in salesman_stats_list)
    total_activated_sum = sum(s['activated_count'] for s in salesman_stats_list)
    total_completed_sum = sum(s['completed_count'] for s in salesman_stats_list)
    
    # 计算平均比率
    avg_touched_rate = round((total_touched_sum / total_incoming_sum) * 100, 1) if total_incoming_sum > 0 else 0
    avg_activated_rate = round((total_activated_sum / total_incoming_sum) * 100, 1) if total_incoming_sum > 0 else 0
    avg_completed_rate = round((total_completed_sum / total_incoming_sum) * 100, 1) if total_incoming_sum > 0 else 0
    
    salesman_stats_summary = {
        'total_count': len(salesman_stats_list),
        'total_incoming': total_incoming_sum,
        'total_touched': total_touched_sum,
        'total_activated': total_activated_sum,
        'total_completed': total_completed_sum,
        'avg_touched_rate': avg_touched_rate,
        'avg_activated_rate': avg_activated_rate,
        'avg_completed_rate': avg_completed_rate
    }
    
    return render_template('customer_tracking.html',
                           groups=groups,
                           salesmen=salesmen,
                           dates=date_list,
                           rows_data=paged_rows_data,
                           salesman_stats=salesman_stats_list,
                           salesman_stats_summary=salesman_stats_summary,
                           selected_group=group_id,
                           selected_salesman=salesman_id,
                           selected_nickname=nickname,
                           selected_followed=is_followed_filter,
                           show_data=show_data,
                           page=page,
                           page_size=page_size,
                           total_pages=total_pages,
                           total_records=total_records,
                           today_str=stats_date.strftime('%Y-%m-%d'),
                           unread_count=get_unread_count(current_user.id))


@bp.route('/admin/customer_detail')
def customer_detail():
    """客户详情页"""
    from models import User, CustomerInfo
    user_id = request.args.get('user_id', type=int)
    nickname = request.args.get('nickname', type=str)
    
    # 获取该用户的所有行为轨迹记录
    records = BehaviorTrackingRecord.query.filter_by(
        user_id=user_id,
        nickname=nickname
    ).order_by(BehaviorTrackingRecord.month, BehaviorTrackingRecord.day).all()
    
    user = User.query.get(user_id) if user_id else None
    
    # 获取客户详细信息
    customer_info = CustomerInfo.query.filter_by(user_id=user_id, nickname=nickname).first()
    
    # 计算统计信息
    total_days = len(records)
    completed_count = sum(1 for r in records if r.play_status == 1)
    touched_count = sum(1 for r in records if r.call_duration_seconds > 0)
    total_call_duration = sum(r.call_duration_seconds for r in records)
    
    # 构建日期数据
    date_list = [f"{r.month:02d}{r.day:02d}" for r in records]
    record_dict = {f"{r.month:02d}{r.day:02d}": r for r in records}
    
    return render_template('customer_detail.html',
                           user=user,
                           nickname=nickname,
                           customer_info=customer_info,
                           records=records,
                           date_list=date_list,
                           record_dict=record_dict,
                           total_days=total_days,
                           completed_count=completed_count,
                           touched_count=touched_count,
                           total_call_duration=total_call_duration,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/admin/customer_detail/save_info', methods=['POST'])
def save_customer_info():
    """保存客户详细信息"""
    from models import CustomerInfo
    import json
    data = request.get_json()
    
    print('[DEBUG] 接收到的数据:', json.dumps(data, ensure_ascii=False))
    
    user_id = data.get('user_id')
    nickname = data.get('nickname')
    clean_nick = clean_nickname(nickname)
    
    # 先尝试用原始昵称查找
    customer_info = CustomerInfo.query.filter_by(user_id=user_id, nickname=nickname).first()
    
    # 如果没找到，查找同一用户下清洗后昵称匹配的记录
    if not customer_info:
        all_customer_infos = CustomerInfo.query.filter_by(user_id=user_id).all()
        for c_info in all_customer_infos:
            if clean_nickname(c_info.nickname) == clean_nick:
                customer_info = c_info
                print(f'[DEBUG] 通过清洗后昵称匹配到记录: {c_info.nickname}')
                break
    
    print('[DEBUG] 查找结果:', '找到记录' if customer_info else '创建新记录')
    
    if not customer_info:
        customer_info = CustomerInfo(user_id=user_id, nickname=nickname)
    else:
        # 如果找到记录，更新昵称为最新版本
        if customer_info.nickname != nickname:
            print(f'[DEBUG] 更新昵称: {customer_info.nickname} -> {nickname}')
            customer_info.nickname = nickname
    
    # 更新字段 - 直接设置，允许空字符串覆盖原值
    customer_info.customer_name = data.get('customer_name') if data.get('customer_name') is not None else ''
    customer_info.gender = data.get('gender') if data.get('gender') is not None else ''
    customer_info.age = data.get('age', 0)
    customer_info.address = data.get('address') if data.get('address') is not None else ''
    customer_info.phone = data.get('phone') if data.get('phone') is not None else ''
    customer_info.health_condition = data.get('health_condition') if data.get('health_condition') is not None else ''
    customer_info.medication_status = data.get('medication_status') if data.get('medication_status') is not None else ''
    
    print('[DEBUG] 更新后的数据:')
    print(f'  customer_name: {repr(customer_info.customer_name)}')
    print(f'  phone: {repr(customer_info.phone)}')
    print(f'  address: {repr(customer_info.address)}')
    
    # 保存观看状态和通话时长修改
    record_changes = data.get('record_changes', [])
    if record_changes:
        print(f'[DEBUG] 记录修改: {len(record_changes)} 条')
        for change in record_changes:
            date_str = change.get('date')
            
            if date_str:
                try:
                    month = int(date_str[:2])
                    day = int(date_str[2:])
                    
                    # 查找或创建行为轨迹记录
                    record = BehaviorTrackingRecord.query.filter_by(
                        user_id=user_id,
                        nickname=nickname,
                        month=month,
                        day=day
                    ).first()
                    
                    if not record:
                        # 如果没有记录，创建一条新记录
                        record = BehaviorTrackingRecord(
                            user_id=user_id,
                            nickname=nickname,
                            month=month,
                            day=day,
                            play_status=0,
                            is_rejected=False,
                            call_duration_seconds=0,
                            play_order=0
                        )
                        db.session.add(record)
                        print(f'[DEBUG] 创建新记录: {date_str}')
                    
                    # 更新观看状态
                    if 'play_status' in change:
                        record.play_status = change['play_status']
                        print(f'[DEBUG] 更新观看状态: {date_str} -> {change["play_status"]}')
                    
                    # 更新拒接状态
                    if 'is_rejected' in change:
                        record.is_rejected = change['is_rejected']
                        print(f'[DEBUG] 更新拒接状态: {date_str} -> {change["is_rejected"]}')
                    
                    # 更新通话时长
                    if 'minutes' in change:
                        record.call_duration_seconds = change['minutes'] * 60
                        print(f'[DEBUG] 更新通话时长: {date_str} -> {change["minutes"]}分钟')
                    
                except Exception as e:
                    print(f'[DEBUG] 处理记录修改失败: {e}')
    
    try:
        db.session.add(customer_info)
        db.session.commit()
        print('[DEBUG] 数据库提交成功')
        return {'success': True}
    except Exception as e:
        db.session.rollback()
        print('[DEBUG] 数据库错误:', str(e))
        return {'success': False, 'message': str(e)}


@bp.route('/admin/customer_tracking/save_records', methods=['POST'])
@role_required('admin', 'salesman')
def save_tracking_records():
    """保存客户跟踪记录（可编辑观看状态）"""
    data = request.get_json()
    records = data.get('records', [])
    
    updated_count = 0
    
    for record in records:
        user_id = record.get('user_id')
        nickname = record.get('nickname')
        date_str = record.get('date')
        play_status = record.get('play_status')
        
        if not user_id or not nickname or not date_str or play_status is None:
            continue
        
        # 解析日期
        try:
            month = int(date_str[:2])
            day = int(date_str[2:])
        except:
            continue
        
        # 查找或创建记录
        existing_record = BehaviorTrackingRecord.query.filter_by(
            user_id=user_id,
            nickname=nickname,
            month=month,
            day=day
        ).first()
        
        if existing_record:
            existing_record.play_status = play_status
        else:
            new_record = BehaviorTrackingRecord(
                user_id=user_id,
                nickname=nickname,
                month=month,
                day=day,
                play_status=play_status,
                call_duration_seconds=0,
                play_order=0
            )
            db.session.add(new_record)
        
        updated_count += 1
    
    try:
        db.session.commit()
        return {'success': True, 'updated_count': updated_count}
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'message': str(e)}
