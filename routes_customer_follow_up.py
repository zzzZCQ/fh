"""客户对接管理"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from models import db, CustomerFollowUp, Order, Category, User, Group
from helpers import get_unread_count
import re
import os
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

bp = Blueprint('customer_follow_up', __name__)


def _extract_product_display(order):
    """从产品信息中提取产品+数量，如：康欣胶囊 x5
    
    逻辑：如果产品信息中包含类别名（或类别名核心部分），则提取数量；否则忽略（返回空）
    """
    info = order.product_info or ''
    category = order.category or ''
    
    if not info:
        return category
    
    # 检查产品信息中是否包含类别名
    if category and category in info:
        # 提取数量
        qty = ''
        qty_match = re.search(r'[xX×*]\s*(\d+)', info)
        if not qty_match:
            qty_match = re.search(r'[数量qty]*[：:]*\s*(\d+)', info)
        if qty_match:
            qty = qty_match.group(1)
        
        if qty:
            return f'{category} x{qty}'
        return category
    
    # 如果类别名包含括号（如"固本回元口服液（0526）"），尝试用核心名称匹配
    # 去掉括号及其内容
    category_core = re.sub(r'[（(].*[）)]', '', category).strip()
    
    if category_core and category_core != category and category_core in info:
        # 使用核心名称匹配
        qty = ''
        qty_match = re.search(r'[xX×*]\s*(\d+)', info)
        if not qty_match:
            qty_match = re.search(r'[数量qty]*[：:]*\s*(\d+)', info)
        if qty_match:
            qty = qty_match.group(1)
        
        if qty:
            return f'{category_core} x{qty}'
        return category_core
    
    # 产品信息中不包含类别名，忽略（可能是赠品）
    return ''


def _extract_amount(order):
    """提取金额：已付定金（只取数字部分）+ 代收金额，计算总金额"""
    total = 0
    # 已付定金：如"100企微0515"只取100
    if order.paid_amount:
        num_match = re.match(r'^(\d+(?:\.\d+)?)', str(order.paid_amount))
        if num_match:
            total += float(num_match.group(1))
    # 代收金额
    if order.collect_amount:
        total += float(order.collect_amount)
    return str(int(total)) if total == int(total) else str(total)


def _merge_products(existing, new_product):
    """合并产品，智能累计数量：康欣胶囊 x30 + 康欣胶囊 x1 = 康欣胶囊 x31"""
    if not existing:
        return new_product
    if not new_product:
        return existing
    
    # 解析现有产品
    products = {}
    for item in existing.split(','):
        item = item.strip()
        if not item:
            continue
        # 匹配 "产品名 x数量" 或 "产品名"
        match = re.match(r'^(.+?)\s*x(\d+)$', item)
        if match:
            name = match.group(1).strip()
            qty = int(match.group(2))
            products[name] = products.get(name, 0) + qty
        else:
            products[item] = products.get(item, 0) + 1
    
    # 解析新产品
    match = re.match(r'^(.+?)\s*x(\d+)$', new_product.strip())
    if match:
        name = match.group(1).strip()
        qty = int(match.group(2))
        products[name] = products.get(name, 0) + qty
    else:
        products[new_product.strip()] = products.get(new_product.strip(), 0) + 1
    
    # 生成结果
    result = []
    for name, qty in products.items():
        result.append(f'{name} x{qty}')
    return ', '.join(result)


def _get_filter_params():
    """获取筛选参数，用于重定向时保留"""
    return {
        'keyword': request.args.get('keyword', ''),
        'is_signed': request.args.get('is_signed', ''),
        'is_followed': request.args.get('is_followed', ''),
        'salesman_id': request.args.get('salesman_id', ''),
        'category': request.args.get('category', ''),
        'group_name': request.args.get('group_name', ''),
        'per_page': request.args.get('per_page', ''),
        'page': request.args.get('page', '')
    }

def _redirect_with_filters():
    """重定向并保留筛选参数"""
    return redirect(url_for('customer_follow_up.follow_up_list', **_get_filter_params()))


@bp.route('/customer_follow_up')
@login_required
def follow_up_list():
    """客户对接列表"""
    # 获取当前用户可管理的组ID列表
    managed_group_ids = current_user.get_managed_group_ids()
    
    # 获取可管理的业务员ID列表（本级及以下组的业务员和管理员）
    managed_salesman_ids = []
    if managed_group_ids:
        managed_salesmen = User.query.filter(
            User.group_id.in_(managed_group_ids),
            (User.roles.like('%salesman%') | User.roles.like('%admin%'))
        ).all()
        managed_salesman_ids = [u.id for u in managed_salesmen]
    
    # 对接员：只看分配给自己的
    if current_user.has_role('follow_up') and not current_user.has_role('admin'):
        query = CustomerFollowUp.query.filter_by(follow_up_person=current_user.name)
    # 管理员：看本级及以下，可按业务员筛选
    elif current_user.has_role('admin'):
        salesman_id = request.args.get('salesman_id', type=int)
        query = CustomerFollowUp.query
        if managed_salesman_ids:
            query = query.filter(CustomerFollowUp.salesman_id.in_(managed_salesman_ids))
        if salesman_id:
            query = query.filter_by(salesman_id=salesman_id)
    # 业务员：只看自己的
    else:
        query = CustomerFollowUp.query.filter_by(salesman_id=current_user.id)

    # 筛选
    keyword = request.args.get('keyword', '').strip()
    if keyword:
        query = query.filter(
            db.or_(
                CustomerFollowUp.customer_name.contains(keyword),
                CustomerFollowUp.phone.contains(keyword)
            )
        )

    is_signed = request.args.get('is_signed', '')
    if is_signed == '1':
        query = query.filter_by(is_main_signed=True)
    elif is_signed == '0':
        query = query.filter_by(is_main_signed=False)

    is_followed = request.args.get('is_followed', '')
    if is_followed == '1':
        query = query.filter_by(is_followed_up=True)
    elif is_followed == '0':
        query = query.filter_by(is_followed_up=False)

    category = request.args.get('category', '')
    if category:
        query = query.filter_by(category=category)
        
    group_name = request.args.get('group_name', '')
    if group_name:
        query = query.filter_by(group_name=group_name)

    # 分页
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in [20, 50, 100]:
        per_page = 20
        
    pagination = query.order_by(
        CustomerFollowUp.group_name,
        CustomerFollowUp.category,
        CustomerFollowUp.salesman_name
    ).paginate(page=page, per_page=per_page, error_out=False)
    
    records = pagination.items

    # 获取业务员列表（管理员用，只显示可管理的）
    salesmen = []
    if current_user.has_role('admin'):
        salesmen = User.query.filter(
            User.group_id.in_(managed_group_ids) if managed_group_ids else True,
            (User.roles.like('%salesman%') | User.roles.like('%admin%'))
        ).all()

    # 获取对接员列表（同步时选择用，只显示可管理的）
    follow_up_users = []
    if managed_group_ids:
        follow_up_users = User.query.filter(
            User.group_id.in_(managed_group_ids),
            User.roles.like('%follow_up%')
        ).all()
    else:
        follow_up_users = User.query.filter(User.roles.like('%follow_up%')).all()

    # 获取主品类别列表
    categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    
    # 获取组别列表
    groups = []
    if managed_group_ids:
        groups = Group.query.filter(Group.id.in_(managed_group_ids)).order_by(Group.name).all()
    elif current_user.group:
        groups = [current_user.group]
    
    # 构建筛选参数，用于分页链接
    filter_params = {}
    if keyword:
        filter_params['keyword'] = keyword
    if is_signed:
        filter_params['is_signed'] = is_signed
    if is_followed:
        filter_params['is_followed'] = is_followed
    if category:
        filter_params['category'] = category
    if group_name:
        filter_params['group_name'] = group_name
    if request.args.get('salesman_id'):
        filter_params['salesman_id'] = request.args.get('salesman_id')
    filter_params['per_page'] = per_page

    return render_template('customer_follow_up.html',
                           records=records,
                           salesmen=salesmen,
                           follow_up_users=follow_up_users,
                           categories=categories,
                           groups=groups,
                           pagination=pagination,
                           filter_params=filter_params,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/customer_follow_up/sync', methods=['POST'])
@login_required
def sync_follow_up():
    """从发货单同步主品数据到客户对接表"""
    # 获取选择的对接员
    follow_up_person = ''
    follow_up_user_id = request.form.get('follow_up_user_id', type=int)
    if follow_up_user_id:
        fu_user = db.session.get(User, follow_up_user_id)
        if fu_user:
            follow_up_person = fu_user.name

    # 获取选择的产品类别
    selected_category = request.form.get('category', '').strip()

    # 获取主品类别
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = [c.name for c in main_categories]

    if not main_cat_names:
        flash('没有配置主品类别，无法同步', 'warning')
        return _redirect_with_filters()

    # 确定要同步的类别
    sync_categories = [selected_category] if selected_category else main_cat_names

    # 根据角色查询发货单
    if current_user.has_role('admin'):
        # 管理员：获取本级及以下所有用户的发货单
        managed_group_ids = current_user.get_managed_group_ids()
        managed_salesman_ids = []
        if managed_group_ids:
            managed_salesmen = User.query.filter(
                User.group_id.in_(managed_group_ids),
                (User.roles.like('%salesman%') | User.roles.like('%admin%'))
            ).all()
            managed_salesman_ids = [u.id for u in managed_salesmen]
        
        if managed_salesman_ids:
            orders = Order.query.filter(
                Order.salesman_id.in_(managed_salesman_ids),
                Order.category.in_(sync_categories),
                Order.status.in_(['submitted', 'shipped'])
            ).all()
        else:
            orders = []
    else:
        # 普通用户：只同步自己的发货单
        orders = Order.query.filter(
            Order.salesman_id == current_user.id,
            Order.category.in_(sync_categories),
            Order.status.in_(['submitted', 'shipped'])
        ).all()

    if not orders:
        flash('没有找到主品发货单数据', 'warning')
        return _redirect_with_filters()

    sync_count = 0
    skip_count = 0

    # 按客户姓名+电话分组
    customer_orders = {}
    for order in orders:
        customer_name = order.customer_name or ''
        phone = order.phone or ''
        key = f'{customer_name}|{phone}'
        if key not in customer_orders:
            customer_orders[key] = []
        customer_orders[key].append(order)

    for key, order_list in customer_orders.items():
        customer_name, phone = key.split('|')
        first_order = order_list[0]
        salesman_name = first_order.salesman.name if first_order.salesman else ''

        # 规范化客户姓名：去除首尾空格
        customer_name_normalized = customer_name.strip()

        # 检查是否已存在（使用规范化后的姓名和电话查询）
        existing = CustomerFollowUp.query.filter(
            CustomerFollowUp.customer_name == customer_name_normalized,
            CustomerFollowUp.phone == phone
        ).first()

        # 重新汇总所有发货单的产品和金额
        total_amount = 0
        all_products = ''
        is_signed = False
        customer_status_list = []  # 收集所有备注
        best_customer_name = ''  # 记录最完整的客户姓名
        best_customer_wechat = ''  # 记录最完整的微信名
        best_gender = ''  # 记录性别
        best_address = ''  # 记录最完整的地址

        for o in order_list:
            # 汇总金额
            amount = _extract_amount(o)
            if amount:
                total_amount += float(amount)
            # 汇总产品
            product_display = _extract_product_display(o)
            if product_display:
                all_products = _merge_products(all_products, product_display)
            # 签收状态：任意一条签收即为签收
            if o.logistics_status == '已签收':
                is_signed = True
            # 收集备注
            if o.remark:
                customer_status_list.append(o.remark)
            # 收集最完整的客户信息
            if o.customer_name and len(o.customer_name.strip()) > len(best_customer_name.strip()):
                best_customer_name = o.customer_name.strip()
            if o.customer_wechat and len(o.customer_wechat) > len(best_customer_wechat):
                best_customer_wechat = o.customer_wechat
            if o.gender:
                best_gender = o.gender
            if o.address and len(o.address) > len(best_address):
                best_address = o.address

        # 使用最完整的客户姓名
        final_customer_name = best_customer_name if best_customer_name else customer_name_normalized

        # 合并备注
        customer_status = '\n'.join(customer_status_list) if customer_status_list else ''

        amount_str = str(int(total_amount)) if total_amount == int(total_amount) else str(total_amount)

        if existing:
            # 更新已存在的记录
            existing.is_main_signed = is_signed
            existing.purchased_products = all_products
            existing.amount = amount_str
            existing.category = first_order.category or ''
            if salesman_name and existing.salesman_name != salesman_name:
                existing.salesman_name = salesman_name
            if follow_up_person and not existing.follow_up_person:
                existing.follow_up_person = follow_up_person
            if best_customer_wechat:
                existing.customer_wechat = best_customer_wechat
            if best_gender:
                existing.gender = best_gender
            # 如果没有客户情况，从备注中提取（避免覆盖已有的内容）
            if not existing.customer_status and customer_status:
                existing.customer_status = customer_status
            skip_count += 1
        else:
            # 新增记录
            record = CustomerFollowUp(
                order_id=first_order.id,
                group_name=first_order.group_name or '',
                salesman_name=salesman_name,
                salesman_id=first_order.salesman_id,
                follow_up_person=follow_up_person,
                customer_name=final_customer_name,
                customer_wechat=best_customer_wechat,
                gender=best_gender or '',
                phone=phone,
                address=best_address or first_order.address or '',
                category=first_order.category or '',
                purchased_products=all_products,
                amount=amount_str,
                customer_status=customer_status,
                is_main_signed=is_signed,
                is_followed_up=False
            )
            db.session.add(record)
            sync_count += 1

    db.session.commit()
    flash(f'同步完成：新增 {sync_count} 条，更新 {skip_count} 条', 'success')
    return _redirect_with_filters()


@bp.route('/customer_follow_up/delete/<int:record_id>', methods=['POST'])
@login_required
def delete_follow_up(record_id):
    """删除客户对接记录"""
    record = CustomerFollowUp.query.get_or_404(record_id)
    # 业务员只能删除自己的，管理员可删所有
    if not current_user.has_role('admin') and record.salesman_id != current_user.id:
        flash('您没有权限删除此记录', 'danger')
        return _redirect_with_filters()
    db.session.delete(record)
    db.session.commit()
    flash('客户对接记录已删除', 'success')
    return _redirect_with_filters()


@bp.route('/customer_follow_up/edit/<int:record_id>', methods=['POST'])
@login_required
def edit_follow_up(record_id):
    """编辑客户对接信息"""
    record = CustomerFollowUp.query.get_or_404(record_id)

    # 权限检查：管理员、对接员（分配给自己的）、业务员（自己的）可编辑
    can_edit = False
    if current_user.has_role('admin'):
        can_edit = True
    elif current_user.has_role('salesman') and record.salesman_id == current_user.id:
        can_edit = True
    elif current_user.has_role('follow_up') and record.follow_up_person == current_user.name:
        can_edit = True

    if not can_edit:
        flash('您没有权限编辑此记录', 'danger')
        return _redirect_with_filters()

    record.customer_name = request.form.get('customer_name', '').strip()
    record.customer_wechat = request.form.get('customer_wechat', '').strip()
    record.gender = request.form.get('gender', '').strip()
    record.phone = request.form.get('phone', '').strip()
    record.address = request.form.get('address', '').strip()
    record.purchased_products = request.form.get('purchased_products', '').strip()
    record.amount = request.form.get('amount', '').strip()
    record.follow_up_person = request.form.get('follow_up_person', '').strip()
    record.customer_status = request.form.get('customer_status', '').strip()
    record.category = request.form.get('category', '').strip()

    db.session.commit()
    flash('客户对接信息已更新', 'success')
    return _redirect_with_filters()


@bp.route('/customer_follow_up/toggle_followed_up/<int:record_id>', methods=['POST'])
@login_required
def toggle_followed_up(record_id):
    """切换对接状态"""
    record = CustomerFollowUp.query.get_or_404(record_id)

    # 权限检查：只有对接员可以变更对接状态
    can_edit = False
    if current_user.has_role('follow_up') and record.follow_up_person == current_user.name:
        can_edit = True

    if not can_edit:
        flash('您没有权限操作此记录', 'danger')
        return _redirect_with_filters()

    # 切换状态
    record.is_followed_up = not record.is_followed_up
    db.session.commit()

    status_text = '已对接' if record.is_followed_up else '未对接'
    flash(f'对接状态已变更为：{status_text}', 'success')
    return _redirect_with_filters()


@bp.route('/api/customer_follow_up/<int:record_id>')
@login_required
def api_follow_up_detail(record_id):
    """获取对接记录详情"""
    record = CustomerFollowUp.query.get_or_404(record_id)
    return jsonify({
        'id': record.id,
        'order_id': record.order_id,
        'group_name': record.group_name,
        'salesman_name': record.salesman_name,
        'follow_up_person': record.follow_up_person,
        'customer_name': record.customer_name,
        'customer_wechat': record.customer_wechat,
        'gender': record.gender,
        'phone': record.phone,
        'address': record.address,
        'category': record.category,
        'purchased_products': record.purchased_products,
        'amount': record.amount,
        'customer_status': record.customer_status,
        'is_main_signed': record.is_main_signed,
        'is_followed_up': record.is_followed_up,
    })


@bp.route('/customer_follow_up/add_customer', methods=['POST'])
@login_required
def add_customer():
    """手动新增客户对接记录"""
    # 只有管理员和业务员可以新增
    if not (current_user.has_role('admin') or current_user.has_role('salesman')):
        flash('您没有权限执行此操作', 'danger')
        return _redirect_with_filters()

    # 获取表单数据
    group_name = request.form.get('group_name', '').strip()
    salesman_id = request.form.get('salesman_id', type=int)
    follow_up_person = request.form.get('follow_up_person', '').strip()
    customer_name = request.form.get('customer_name', '').strip()
    customer_wechat = request.form.get('customer_wechat', '').strip()
    gender = request.form.get('gender', '').strip()
    phone = request.form.get('phone', '').strip()
    address = request.form.get('address', '').strip()
    category = request.form.get('category', '').strip()
    purchased_products = request.form.get('purchased_products', '').strip()
    amount = request.form.get('amount', '').strip()
    customer_status = request.form.get('customer_status', '').strip()
    is_main_signed = request.form.get('is_main_signed') == 'on'
    is_followed_up = request.form.get('is_followed_up') == 'on'

    # 验证必填项
    if not customer_name or not phone:
        flash('客户姓名和电话不能为空！', 'danger')
        return _redirect_with_filters()

    # 获取业务员信息
    salesman_name = current_user.name
    if salesman_id:
        salesman = db.session.get(User, salesman_id)
        if salesman:
            salesman_name = salesman.name

    # 检查是否已存在（按客户姓名+电话去重）
    existing = CustomerFollowUp.query.filter_by(
        customer_name=customer_name,
        phone=phone
    ).first()

    if existing:
        flash('该客户已存在（相同客户姓名+电话）！', 'warning')
        return _redirect_with_filters()

    # 创建记录
    record = CustomerFollowUp(
        order_id=None,
        group_name=group_name,
        salesman_name=salesman_name,
        salesman_id=salesman_id if salesman_id else current_user.id,
        follow_up_person=follow_up_person,
        customer_name=customer_name,
        customer_wechat=customer_wechat,
        gender=gender,
        phone=phone,
        address=address,
        category=category,
        purchased_products=purchased_products,
        amount=amount,
        customer_status=customer_status,
        is_main_signed=is_main_signed,
        is_followed_up=is_followed_up
    )
    db.session.add(record)
    db.session.commit()

    flash(f'客户 {customer_name} 添加成功！', 'success')
    return _redirect_with_filters()


@bp.route('/customer_follow_up/export')
@login_required
def export_follow_up():
    """导出客户对接数据为Excel"""
    # 获取筛选条件
    keyword = request.args.get('keyword', '').strip()
    is_signed = request.args.get('is_signed', '')
    is_followed = request.args.get('is_followed', '')
    salesman_id = request.args.get('salesman_id', type=int)
    category = request.args.get('category', '')
    group_name = request.args.get('group_name', '')

    # 获取当前用户可管理的组ID列表
    managed_group_ids = current_user.get_managed_group_ids()
    
    # 获取可管理的业务员ID列表（本级及以下组的业务员和管理员）
    managed_salesman_ids = []
    if managed_group_ids:
        managed_salesmen = User.query.filter(
            User.group_id.in_(managed_group_ids),
            (User.roles.like('%salesman%') | User.roles.like('%admin%'))
        ).all()
        managed_salesman_ids = [u.id for u in managed_salesmen]

    # 构建查询
    if current_user.has_role('follow_up') and not current_user.has_role('admin'):
        query = CustomerFollowUp.query.filter_by(follow_up_person=current_user.name)
    elif current_user.has_role('admin'):
        query = CustomerFollowUp.query
        if managed_salesman_ids:
            query = query.filter(CustomerFollowUp.salesman_id.in_(managed_salesman_ids))
        if salesman_id:
            query = query.filter_by(salesman_id=salesman_id)
    else:
        query = CustomerFollowUp.query.filter_by(salesman_id=current_user.id)

    if keyword:
        query = query.filter(
            db.or_(
                CustomerFollowUp.customer_name.contains(keyword),
                CustomerFollowUp.phone.contains(keyword)
            )
        )
    if is_signed == '1':
        query = query.filter_by(is_main_signed=True)
    elif is_signed == '0':
        query = query.filter_by(is_main_signed=False)
    if is_followed == '1':
        query = query.filter_by(is_followed_up=True)
    elif is_followed == '0':
        query = query.filter_by(is_followed_up=False)
    if category:
        query = query.filter_by(category=category)
    if group_name:
        query = query.filter_by(group_name=group_name)

    records = query.order_by(
        CustomerFollowUp.group_name,
        CustomerFollowUp.category,
        CustomerFollowUp.salesman_name
    ).all()

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '客户档案交接登记表'

    # 设置列宽
    col_widths = [15, 12, 6, 15, 40, 20, 12, 12, 30, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 第1行：空行
    # 第2行：表头
    headers = ['微信名', '客户姓名', '性别', '电话号码', '系统下单详细地址', 
               '购买产品+数量', '购买金额', '一线客服', '客户身体情况、其他客情、特殊情况', '主品是否签收']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # 第3行开始：数据
    row_num = 3
    for r in records:
        ws.cell(row=row_num, column=1, value=r.customer_wechat or '')
        ws.cell(row=row_num, column=2, value=r.customer_name or '')
        ws.cell(row=row_num, column=3, value=r.gender or '')
        ws.cell(row=row_num, column=4, value=r.phone or '')
        ws.cell(row=row_num, column=5, value=r.address or '')
        ws.cell(row=row_num, column=6, value=r.purchased_products or '')
        ws.cell(row=row_num, column=7, value=r.amount or '')
        # 一线客服取业务员姓名
        ws.cell(row=row_num, column=8, value=r.salesman.name if r.salesman else r.salesman_name or '')
        ws.cell(row=row_num, column=9, value=r.customer_status or '')
        ws.cell(row=row_num, column=10, value='是' if r.is_main_signed else '否')
        row_num += 1

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'客户档案交接登记表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
