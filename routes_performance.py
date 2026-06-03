# -*- coding: utf-8 -*-
"""业绩报表路由"""
import os
import re
import json
from datetime import datetime, timedelta
from calendar import monthrange
from flask import Blueprint, request, redirect, url_for, flash, render_template, send_file, jsonify
from flask_login import current_user

from models import db, Order, PerformanceReportTemplate, User, Group, Category, _now_bj
from helpers import role_required, get_unread_count

bp = Blueprint('performance', __name__)


# ============ 业绩管理页面 ============
@bp.route('/salesman/performance_dashboard')
@role_required('salesman')
def performance_dashboard():
    """业绩管理页面"""
    now = _now_bj()
    return render_template('performance_dashboard.html',
                           now_year=now.year,
                           now_month=now.month,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/salesman/performance/api/overview')
@role_required('salesman')
def performance_overview():
    """业绩概览API - 本月、上月、同比、环比"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    # 本月
    # 总业绩（所有已发货）
    current_total_amount = calculate_total_performance(current_user.id, year, month)
    current_total_count = count_total_orders(current_user.id, year, month)
    # 已签收业绩
    current_signed_amount = calculate_performance(current_user.id, year, month)
    current_signed_count = count_orders(current_user.id, year, month)
    
    # 上月
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    prev_total_amount = calculate_total_performance(current_user.id, prev_year, prev_month)
    prev_signed_amount = calculate_performance(current_user.id, prev_year, prev_month)
    
    # 同比（去年同期）
    yoy_year = year - 1
    yoy_total_amount = calculate_total_performance(current_user.id, yoy_year, month)
    yoy_signed_amount = calculate_performance(current_user.id, yoy_year, month)
    
    # 计算环比和同比变化率（基于已签收业绩）
    mom_rate = ((current_signed_amount - prev_signed_amount) / prev_signed_amount * 100) if prev_signed_amount > 0 else 0
    yoy_rate = ((current_signed_amount - yoy_signed_amount) / yoy_signed_amount * 100) if yoy_signed_amount > 0 else 0
    
    return jsonify({
        'current_month': {
            'total_amount': current_total_amount,
            'total_count': current_total_count,
            'signed_amount': current_signed_amount,
            'signed_count': current_signed_count
        },
        'prev_month': {'total_amount': prev_total_amount, 'signed_amount': prev_signed_amount},
        'yoy': {'total_amount': yoy_total_amount, 'signed_amount': yoy_signed_amount},
        'mom_rate': round(mom_rate, 1),
        'yoy_rate': round(yoy_rate, 1)
    })


@bp.route('/salesman/performance/api/monthly_trend')
@role_required('salesman')
def monthly_trend():
    """月度业绩趋势API - 最近12个月"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    # 最近12个月
    months = []
    amounts = []
    counts = []
    
    for i in range(11, -1, -1):
        # 计算月份
        m = month - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1
        
        amount = calculate_performance(current_user.id, y, m)
        count = count_orders(current_user.id, y, m)
        
        months.append(f'{y}-{m:02d}')
        amounts.append(amount)
        counts.append(count)
    
    return jsonify({'months': months, 'amounts': amounts, 'counts': counts})


@bp.route('/salesman/performance/api/category_distribution')
@role_required('salesman')
def category_distribution():
    """产品类别分布API"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    # 获取主产品类别
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}

    # 不按时间筛选，后面再判断归属月份
    # 按类别统计（统计submitted和shipped状态的订单，排除退回已签收）
    orders = Order.query.filter(
        Order.salesman_id == current_user.id,
        Order.status.in_(['submitted', 'shipped']),
        Order.logistics_status != '退回已签收'
    ).all()

    category_data = {}
    for order in orders:
        cat = order.category or '未分类'
        # 只统计主产品且金额>0的
        if cat not in main_cat_names:
            continue
        amount = get_order_amount(order)
        if amount <= 0:
            continue
        
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        if cat not in category_data:
            category_data[cat] = {'amount': 0, 'count': 0}
        category_data[cat]['amount'] += amount
        category_data[cat]['count'] += 1

    return jsonify(category_data)


@bp.route('/salesman/performance/api/order_list')
@role_required('salesman')
def order_list():
    """订单列表API - 只展示金额>0的已签收订单"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month

    # 不按时间筛选，后面再判断归属月份
    orders = Order.query.filter(
        Order.salesman_id == current_user.id,
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    ).order_by(Order.sign_time.desc()).all()

    data = []
    for order in orders:
        # 判断订单归属月份（已签收订单按签收时间）
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
            if order_month == month and order_year == year:
                total_amount = get_order_amount(order)
                if total_amount > 0:  # 只展示金额>0的
                    data.append({
                        'id': order.id,
                        'customer_name': order.customer_name or '',
                        'category': order.category or '',
                        'paid_amount': order.paid_amount or '',
                        'collect_amount': order.collect_amount or 0,
                        'total_amount': total_amount,
                        'sign_time': order.sign_time.strftime('%Y-%m-%d %H:%M') if order.sign_time else '',
                        'tracking_number': order.tracking_number or ''
                    })

    return jsonify({'orders': data})


# ============ 辅助函数 ============
def calculate_performance(salesman_id, year, month):
    """计算指定月份的已签收业绩金额（只统计金额>0的）"""
    # 获取所有已签收订单，不按时间筛选
    orders = Order.query.filter(
        Order.salesman_id == salesman_id,
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    ).all()

    total = 0
    for order in orders:
        # 判断订单归属月份（已签收订单按签收时间）
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
            if order_month == month and order_year == year:
                amt = get_order_amount(order)
                if amt > 0:
                    total += amt
    return total


def count_orders(salesman_id, year, month):
    """统计指定月份的已签收订单数（只统计金额>0的）"""
    # 获取所有已签收订单，不按时间筛选
    orders = Order.query.filter(
        Order.salesman_id == salesman_id,
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    ).all()

    count = 0
    for order in orders:
        # 判断订单归属月份（已签收订单按签收时间）
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
            if order_month == month and order_year == year:
                if get_order_amount(order) > 0:
                    count += 1
    return count


def calculate_total_performance(salesman_id, year, month):
    """计算指定月份的总业绩金额（只统计主产品且金额>0）"""
    # 获取主产品类别
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}
    
    # 获取所有订单，不按时间筛选
    orders = Order.query.filter(
        Order.salesman_id == salesman_id,
        Order.status.in_(['submitted', 'shipped']),
        Order.logistics_status != '退回已签收'
    ).all()
    
    total = 0
    for order in orders:
        if (order.category or '未分类') in main_cat_names:
            # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
            order_month = None
            order_year = None
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            elif order.create_time:
                order_month = order.create_time.month
                order_year = order.create_time.year
            
            # 只统计当前查询月份的订单
            if order_month == month and order_year == year:
                amt = get_order_amount(order)
                if amt > 0:
                    total += amt
    return total


def count_total_orders(salesman_id, year, month):
    """统计指定月份的总订单数（只统计主产品且金额>0）"""
    # 获取主产品类别
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}
    
    # 获取所有订单，不按时间筛选
    orders = Order.query.filter(
        Order.salesman_id == salesman_id,
        Order.status.in_(['draft', 'submitted', 'shipped']),
        Order.logistics_status != '退回已签收'
    ).all()

    count = 0
    for order in orders:
        if (order.category or '未分类') in main_cat_names:
            # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
            order_month = None
            order_year = None
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            elif order.create_time:
                order_month = order.create_time.month
                order_year = order.create_time.year
            
            # 只统计当前查询月份的订单
            if order_month == month and order_year == year:
                if get_order_amount(order) > 0:
                    count += 1
    return count


def get_order_amount(order):
    """获取订单金额（已付定金数字 + 代收金额）"""
    paid_str = str(order.paid_amount or '')
    paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
    collect_num = float(order.collect_amount or 0)
    return paid_num + collect_num


# ============ 业绩报表导出 ============
@bp.route('/salesman/performance_report')
@role_required('salesman')
def performance_report_page():
    """业绩报表导出页面"""
    templates = PerformanceReportTemplate.query.filter_by(is_active=True).all()
    now = _now_bj()
    return render_template('performance_report.html',
                           templates=templates,
                           now_year=now.year,
                           now_month=now.month,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/salesman/performance/api/export', methods=['POST'])
@role_required('salesman')
def api_export_performance():
    """导出业绩报表 - 直接导出，使用激活的模板"""
    if request.is_json:
        data = request.get_json(silent=True) or {}
        year = data.get('year')
        month = data.get('month')
    else:
        year = request.form.get('year', type=int)
        month = request.form.get('month', type=int)

    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month

    # 自动使用激活的模板
    template = PerformanceReportTemplate.query.filter_by(is_active=True).first()
    if not template:
        return jsonify({'success': False, 'message': '没有找到激活的业绩报表模板，请联系管理员配置！'}), 400

    # 获取模板配置
    field_mapping = json.loads(template.field_mapping) if template.field_mapping else {}

    # 直接查询所有主产品
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    allowed_categories = [c.name for c in main_categories]

    # 不按时间筛选，后面再判断归属月份
    # 查询订单：当前业务员、已发货状态、已签收、有订金或代收金额
    # 时间筛选：以签收时间为准，只有已签收的订单才计入业绩
    query = Order.query.filter(
        Order.salesman_id == current_user.id,
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    )

    if allowed_categories:
        query = query.filter(Order.category.in_(allowed_categories))

    # 过滤：只有有订金或代收金额的订单才导出（金额>0），并且按归属月份筛选
    orders_to_export = []
    for order in query.all():
        # 判断订单归属月份（已签收订单按签收时间）
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
            if order_month == month and order_year == year:
                paid_num = 0
                if order.paid_amount:
                    paid_str = str(order.paid_amount)
                    match = re.match(r'[\d.]+', paid_str)
                    if match:
                        paid_num = float(match.group())
                collect_num = float(order.collect_amount or 0)
                if paid_num > 0 or collect_num > 0:
                    orders_to_export.append(order)

    orders = orders_to_export

    if not orders:
        return jsonify({'success': False, 'message': '没有找到符合条件的（金额>0）的订单！'}), 400

    # 检查模板文件是否存在
    if not template.filepath or not os.path.exists(template.filepath):
        return jsonify({'success': False, 'message': '模板文件不存在，请联系管理员重新上传模板！'}), 400

    # 使用模板文件填充数据
    import io
    from openpyxl import load_workbook
    from openpyxl.styles import numbers

    # 加载模板文件
    wb = load_workbook(template.filepath)
    ws = wb.active

    # 获取表头（第一行）
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # 找出日期列的索引
    date_col_idx = None
    for idx, h in enumerate(headers):
        if field_mapping.get(h) == 'create_time':
            date_col_idx = idx + 1  # openpyxl列索引从1开始
            break

    # 从第二行开始填充数据
    row_num = 2
    for order in orders:
        row_data = []
        for header in headers:
            order_field = field_mapping.get(header, '')
            if order_field == '__sales_amount__':
                paid_str = str(order.paid_amount or '')
                paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
                collect_num = float(order.collect_amount or 0)
                row_data.append(paid_num + collect_num)
            elif order_field == 'create_time' or order_field == 'sign_time':
                value = getattr(order, order_field, '') or ''
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M')
                row_data.append(value)
            elif order_field == 'salesman_name':
                row_data.append(current_user.name)
            elif order_field == 'group_name':
                if current_user.group:
                    row_data.append(current_user.group.name)
                else:
                    row_data.append('')
            elif order_field == 'product_info':
                row_data.append(order.product_info or order.category or '')
            elif order_field == 'total_amount':
                paid_str = str(order.paid_amount or '')
                paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
                collect_num = float(order.collect_amount or 0)
                row_data.append(paid_num + collect_num)
            elif order_field:
                value = getattr(order, order_field, '') or ''
                row_data.append(value)
            else:
                row_data.append('')
        ws.append(row_data)

        # 设置日期列格式
        if date_col_idx:
            ws.cell(row=row_num, column=date_col_idx).number_format = 'yyyy-mm-dd'
        row_num += 1

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'业绩报表_{current_user.name}_{year}年{month}月.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============ 管理员团队业绩统计 ============
@bp.route('/admin/team_performance')
@role_required('admin')
def team_performance_dashboard():
    """管理员团队业绩统计页面"""
    now = _now_bj()
    
    # 获取管理的组别列表
    if current_user.username == 'admin':
        groups = Group.query.filter_by(is_active=True).order_by(Group.level.asc(), Group.create_time.asc()).all()
    elif current_user.group_id:
        managed_group_ids = current_user.get_managed_group_ids()
        groups = Group.query.filter(Group.id.in_(managed_group_ids), Group.is_active==True).order_by(Group.level.asc(), Group.create_time.asc()).all()
    else:
        groups = []
    
    return render_template('admin_team_performance.html',
                           groups=groups,
                           now_year=now.year,
                           now_month=now.month,
                           unread_count=get_unread_count(current_user.id))


@bp.route('/admin/team_performance/api/overview')
@role_required('admin')
def team_performance_overview():
    """团队业绩概览API"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)
    show_all = request.args.get('show_all', type=int)  # 1=展开全部，0=只显示TOP10
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    # 获取管理的用户ID列表
    user_ids = get_managed_user_ids(group_id)
    
    # 计算月份范围（当月）
    start_date = datetime(year, month, 1)
    end_date = datetime(year, month, monthrange(year, month)[1], 23, 59, 59)
    
    # 计算上月范围
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1
    prev_start = datetime(prev_year, prev_month, 1)
    prev_end = datetime(prev_year, prev_month, monthrange(prev_year, prev_month)[1], 23, 59, 59)
    
    # 团队业绩统计
    # 获取主产品类别列表
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}
    
    # ========== 当月数据 ==========
    # 获取所有订单（不按时间筛选，后面再判断归属月份）
    all_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status.in_(['draft', 'submitted', 'shipped'])
    ).all()
    
    # 获取已签收订单（不按时间筛选，后面再判断归属月份）
    signed_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    ).all()
    
    # ========== 上月数据 ==========
    # 上月所有订单
    prev_all_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status.in_(['draft', 'submitted', 'shipped'])
    ).all()
    
    # 上月已签收订单
    prev_signed_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    ).all()
    
    # ========== 统计当月总业绩 ==========
    # 总业绩：当月所有订单（只统计主产品且金额>0，归属月份判断
    total_amount = 0
    total_count = 0
    for order in all_orders:
        # 只统计主产品
        if (order.category or '未分类') in main_cat_names:
            # 判断是否是退回已签收（两种情况：1. logistics_status 是退回已签收；2. logistics_warning_remark 包含退回关键词且是已签收状态）
            is_returned = False
            if order.logistics_status == '退回已签收':
                is_returned = True
            elif order.logistics_status == '已签收' and order.logistics_warning_remark:
                return_keywords = ['退回', '拒收', '退件']
                is_returned = any(keyword in order.logistics_warning_remark for keyword in return_keywords)
            
            # 排除退回已签收和拒签
            if is_returned or order.logistics_status == '拒签':
                continue
            
            # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
            order_month = None
            order_year = None
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            elif order.create_time:
                order_month = order.create_time.month
                order_year = order.create_time.year
            
            # 只统计当前查询月份的订单
            if order_month != month or order_year != year:
                continue
            
            # 金额>0的订单统计
            amt = get_order_amount(order)
            if amt > 0:
                total_amount += amt
                total_count += 1
    
    # 已签收业绩（只统计主产品且金额>0，归属月份判断
    signed_amount = 0
    signed_count = 0
    for order in signed_orders:
        if (order.category or '未分类') in main_cat_names:
            # 判断订单归属月份：以签收时间为主
            order_month = None
            order_year = None
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            
            if order_month != month or order_year != year:
                continue
            
            amt = get_order_amount(order)
            if amt > 0:
                signed_amount += amt
                signed_count += 1
    
    # ========== 统计上月总业绩 ==========
    prev_total_amount = 0
    prev_total_count = 0
    for order in prev_all_orders:
        # 只统计主产品
        if (order.category or '未分类') in main_cat_names:
            # 判断是否是退回已签收
            is_returned = False
            if order.logistics_status == '退回已签收':
                is_returned = True
            elif order.logistics_status == '已签收' and order.logistics_warning_remark:
                return_keywords = ['退回', '拒收', '退件']
                is_returned = any(keyword in order.logistics_warning_remark for keyword in return_keywords)
            
            # 排除退回已签收和拒签
            if is_returned or order.logistics_status == '拒签':
                continue
            
            # 判断订单归属月份（上月）
            order_month = None
            order_year = None
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            elif order.create_time:
                order_month = order.create_time.month
                order_year = order.create_time.year
            
            # 只统计上月的订单
            if order_month != prev_month or order_year != prev_year:
                continue
            
            amt = get_order_amount(order)
            if amt > 0:
                prev_total_amount += amt
                prev_total_count += 1
    
    prev_signed_amount = 0
    prev_signed_count = 0
    for order in prev_signed_orders:
        if (order.category or '未分类') in main_cat_names:
            # 判断归属月份
            order_month = None
            order_year = None
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            
            if order_month != prev_month or order_year != prev_year:
                continue
            
            amt = get_order_amount(order)
            if amt > 0:
                prev_signed_amount += amt
                prev_signed_count += 1

    # 按业务员统计（只统计主产品且金额>0）
    salesman_stats = {}
    
    # 第一步：先获取所有权限可见的业务员并初始化
    # 获取管理的所有用户
    managed_users = []
    if current_user.username == 'admin':
        if group_id:
            group = Group.query.get(group_id)
            if group:
                group_ids = [group.id] + group.get_all_children_ids()
                managed_users = User.query.filter(User.group_id.in_(group_ids), User.roles.contains('salesman')).all()
        else:
            managed_users = User.query.filter(User.roles.contains('salesman')).all()
    elif current_user.group_id:
        managed_ids = current_user.get_managed_group_ids()
        if group_id:
            if group_id in managed_ids:
                group = Group.query.get(group_id)
                if group:
                    group_ids = [group.id] + group.get_all_children_ids()
                    managed_users = User.query.filter(User.group_id.in_(group_ids), User.roles.contains('salesman')).all()
        else:
            managed_users = User.query.filter(User.group_id.in_(managed_ids), User.roles.contains('salesman')).all()
    
    # 初始化所有可见业务员（即使没有业绩）
    for user in managed_users:
        salesman_stats[user.id] = {
            'name': user.name,
            'group_name': user.group.name if user.group else '未分组',
            'total_amount': 0,      # 总业绩
            'total_count': 0,       # 总订单数
            'signed_amount': 0,     # 已签收业绩
            'signed_count': 0,      # 已签收订单数
            'prev_total_amount': 0, # 上月总业绩
            'prev_signed_amount': 0 # 上月已签收业绩
        }
    
    # 第二步：统计总业绩
    for order in all_orders:
        cat = order.category or '未分类'
        if cat not in main_cat_names:
            continue  # 跳过非主产品
        amt = get_order_amount(order)
        if amt <= 0:
            continue  # 跳过金额为0的
        
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        sid = order.salesman_id
        if sid in salesman_stats:  # 只统计可见用户
            salesman_stats[sid]['total_amount'] += amt
            salesman_stats[sid]['total_count'] += 1
    
    # 第三步：统计已签收业绩
    for order in signed_orders:
        sid = order.salesman_id
        if sid in salesman_stats:  # 只统计可见用户
            # 检查订单是否符合统计条件（主产品且金额>0）
            cat = order.category or '未分类'
            if cat not in main_cat_names:
                continue
            amt = get_order_amount(order)
            if amt <= 0:
                continue
            
            # 判断归属月份（已签收订单按签收时间）
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            else:
                continue  # 没有签收时间的已签收订单不统计
            
            if order_month != month or order_year != year:
                continue
            
            salesman_stats[sid]['signed_amount'] += amt
            salesman_stats[sid]['signed_count'] += 1
    
    # 第四步：统计上月总业绩
    for order in prev_all_orders:
        cat = order.category or '未分类'
        if cat not in main_cat_names:
            continue  # 跳过非主产品
        amt = get_order_amount(order)
        if amt <= 0:
            continue  # 跳过金额为0的
        
        # 判断订单归属月份（上月）：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计上月的订单
        if order_month != prev_month or order_year != prev_year:
            continue
        
        sid = order.salesman_id
        if sid in salesman_stats:  # 只统计可见用户
            salesman_stats[sid]['prev_total_amount'] += amt
    
    # 第五步：统计上月已签收业绩
    for order in prev_signed_orders:
        sid = order.salesman_id
        if sid in salesman_stats:  # 只统计可见用户
            # 检查订单是否符合统计条件（主产品且金额>0）
            cat = order.category or '未分类'
            if cat not in main_cat_names:
                continue
            amt = get_order_amount(order)
            if amt <= 0:
                continue
            
            # 判断归属月份（上月，已签收订单按签收时间）
            if order.sign_time:
                order_month = order.sign_time.month
                order_year = order.sign_time.year
            else:
                continue  # 没有签收时间的已签收订单不统计
            
            if order_month != prev_month or order_year != prev_year:
                continue
            
            salesman_stats[sid]['prev_signed_amount'] += amt
    
    # 第六步：计算每个业务员的环比
    for sid in salesman_stats:
        s = salesman_stats[sid]
        # 总业绩环比
        if s['prev_total_amount'] > 0:
            s['total_amount_rate'] = (s['total_amount'] - s['prev_total_amount']) / s['prev_total_amount'] * 100
        elif s['total_amount'] > 0:
            s['total_amount_rate'] = 100
        else:
            s['total_amount_rate'] = 0
        # 已签收环比
        if s['prev_signed_amount'] > 0:
            s['signed_amount_rate'] = (s['signed_amount'] - s['prev_signed_amount']) / s['prev_signed_amount'] * 100
        elif s['signed_amount'] > 0:
            s['signed_amount_rate'] = 100
        else:
            s['signed_amount_rate'] = 0
    
    # 排序：按总业绩金额降序
    sorted_stats = sorted(salesman_stats.values(), key=lambda x: x['total_amount'], reverse=True)
    
    # 默认只返回TOP10，展开时返回全部
    display_stats = sorted_stats if show_all else sorted_stats[:10]
    
    # 计算环比
    # 总业绩环比
    if prev_total_amount > 0:
        total_amount_rate = (total_amount - prev_total_amount) / prev_total_amount * 100
    elif total_amount > 0:
        total_amount_rate = 100
    else:
        total_amount_rate = 0
    
    # 已签收环比
    if prev_signed_amount > 0:
        signed_amount_rate = (signed_amount - prev_signed_amount) / prev_signed_amount * 100
    elif signed_amount > 0:
        signed_amount_rate = 100
    else:
        signed_amount_rate = 0
    
    return jsonify({
        'total_amount': total_amount,
        'total_count': total_count,
        'signed_amount': signed_amount,
        'signed_count': signed_count,
        'salesman_count': len(salesman_stats),
        'salesman_stats': display_stats,
        'has_more': len(sorted_stats) > 10,
        # 环比数据
        'prev_total_amount': prev_total_amount,
        'prev_total_count': prev_total_count,
        'prev_signed_amount': prev_signed_amount,
        'prev_signed_count': prev_signed_count,
        'total_amount_rate': total_amount_rate,
        'signed_amount_rate': signed_amount_rate,
        'prev_month': f'{prev_year}年{prev_month}月'
    })


@bp.route('/admin/team_performance/api/monthly_trend')
@role_required('admin')
def team_monthly_trend():
    """团队月度趋势API"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    user_ids = get_managed_user_ids(group_id)
    
    # 最近12个月
    months = []
    amounts = []
    counts = []
    
    for i in range(11, -1, -1):
        m = month - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1
        
        start = datetime(y, m, 1)
        end = datetime(y, m, monthrange(y, m)[1], 23, 59, 59)
        
        orders = Order.query.filter(
            Order.salesman_id.in_(user_ids),
            Order.status == 'shipped',
            Order.logistics_status == '已签收',
            Order.sign_time >= start,
            Order.sign_time <= end
        ).all()
        
        amount = 0
        count = 0
        for o in orders:
            o_amount = get_order_amount(o)
            if o_amount > 0:
                amount += o_amount
                count += 1
        
        months.append(f'{y}-{m:02d}')
        amounts.append(amount)
        counts.append(count)
    
    return jsonify({'months': months, 'amounts': amounts, 'counts': counts})


@bp.route('/admin/team_performance/api/category_distribution')
@role_required('admin')
def team_category_distribution():
    """团队产品类别分布API"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    user_ids = get_managed_user_ids(group_id)
    
    # 获取主产品类别
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}

    # 不按时间筛选，后面再判断归属月份
    orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status.in_(['submitted', 'shipped']),
        Order.logistics_status != '退回已签收'
    ).all()

    category_data = {}
    for order in orders:
        cat = order.category or '未分类'
        # 只统计主产品且金额>0的
        if cat not in main_cat_names:
            continue
        amount = get_order_amount(order)
        if amount <= 0:
            continue
        
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        if cat not in category_data:
            category_data[cat] = {'amount': 0, 'count': 0, 'max_amount': 0, 'amounts': []}
        category_data[cat]['amount'] += amount
        category_data[cat]['count'] += 1
        category_data[cat]['amounts'].append(amount)
        if amount > category_data[cat]['max_amount']:
            category_data[cat]['max_amount'] = amount
    
    # 计算平均金额
    for cat in category_data:
        data = category_data[cat]
        if data['count'] > 0:
            data['avg_amount'] = sum(data['amounts']) / data['count']
        else:
            data['avg_amount'] = 0
        del data['amounts']

    return jsonify(category_data)


def get_managed_user_ids(group_id=None):
    """获取当前管理员可管理的用户ID列表"""
    # 超级管理员看所有用户
    if current_user.username == 'admin':
        if group_id:
            # 如果指定了组别，只查该组及其子组的用户
            group = Group.query.get(group_id)
            if group:
                group_ids = [group.id] + group.get_all_children_ids()
                users = User.query.filter(User.group_id.in_(group_ids)).all()
                return [u.id for u in users]
        # 查所有业务员
        users = User.query.filter(User.roles.contains('salesman')).all()
        return [u.id for u in users]
    
    # 普通管理员只看管理的组别
    if current_user.group_id:
        if group_id:
            # 检查是否有权限查看该组
            managed_ids = current_user.get_managed_group_ids()
            if group_id in managed_ids:
                group = Group.query.get(group_id)
                if group:
                    group_ids = [group.id] + group.get_all_children_ids()
                    users = User.query.filter(User.group_id.in_(group_ids)).all()
                    return [u.id for u in users]
        
        # 默认查所有管理的组别
        managed_ids = current_user.get_managed_group_ids()
        users = User.query.filter(User.group_id.in_(managed_ids)).all()
        return [u.id for u in users]
    
    return []


@bp.route('/admin/team_performance/api/sign_rate_by_person')
@role_required('admin')
def sign_rate_by_person():
    """个人维度签收率API - 只统计主产品"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)

    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month

    user_ids = get_managed_user_ids(group_id)

    # 获取主产品类别列表
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}

    # 查询所有已提交和已发货订单（不按时间筛选，后面再判断归属月份）
    all_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status.in_(['submitted', 'shipped'])
    ).all()

    # 按业务员统计（只统计主产品且金额>0）
    person_stats = {}
    
    # 第一步：先获取所有权限可见的业务员并初始化
    managed_users = []
    if current_user.username == 'admin':
        if group_id:
            group = Group.query.get(group_id)
            if group:
                group_ids = [group.id] + group.get_all_children_ids()
                managed_users = User.query.filter(User.group_id.in_(group_ids), User.roles.contains('salesman')).all()
        else:
            managed_users = User.query.filter(User.roles.contains('salesman')).all()
    elif current_user.group_id:
        managed_ids = current_user.get_managed_group_ids()
        if group_id:
            if group_id in managed_ids:
                group = Group.query.get(group_id)
                if group:
                    group_ids = [group.id] + group.get_all_children_ids()
                    managed_users = User.query.filter(User.group_id.in_(group_ids), User.roles.contains('salesman')).all()
        else:
            managed_users = User.query.filter(User.group_id.in_(managed_ids), User.roles.contains('salesman')).all()
    
    # 初始化所有可见业务员（即使没有订单）
    for user in managed_users:
        person_stats[user.id] = {
            'name': user.name,
            'total': 0,
            'signed': 0
        }
    
    # 第二步：统计订单（包含退回已签收但不算签收）
    for order in all_orders:
        cat = order.category or '未分类'
        if cat not in main_cat_names:
            continue  # 跳过非主产品
        # 跳过金额为0的
        if get_order_amount(order) <= 0:
            continue
        
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        sid = order.salesman_id
        if sid in person_stats:  # 只统计可见用户
            person_stats[sid]['total'] += 1
            # 只统计"已签收"为签收成功，退回已签收不算签收
            if order.logistics_status == '已签收':
                person_stats[sid]['signed'] += 1

    # 计算签收率并排序
    result = []
    for sid, stat in person_stats.items():
        rate = (stat['signed'] / stat['total'] * 100) if stat['total'] > 0 else 0
        result.append({
            'name': stat['name'],
            'total': stat['total'],
            'signed': stat['signed'],
            'unsigned': stat['total'] - stat['signed'],
            'rate': round(rate, 1)
        })

    result.sort(key=lambda x: x['rate'], reverse=True)
    return jsonify(result)


@bp.route('/admin/team_performance/api/sign_rate_by_category')
@role_required('admin')
def sign_rate_by_category():
    """产品维度签收率API - 只展示主产品"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)

    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month

    user_ids = get_managed_user_ids(group_id)

    # 获取主产品类别列表
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}

    # 查询所有已提交和已发货订单（不按时间筛选，后面再判断归属月份）
    all_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status.in_(['submitted', 'shipped'])
    ).all()

    # 按产品类别统计（只统计主产品且金额>0，包含退回已签收但不算签收）
    cat_stats = {}
    for order in all_orders:
        cat = order.category or '未分类'
        if cat not in main_cat_names:
            continue  # 跳过非主产品
        # 跳过金额为0的
        if get_order_amount(order) <= 0:
            continue
        
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        if cat not in cat_stats:
            cat_stats[cat] = {'total': 0, 'signed': 0}
        cat_stats[cat]['total'] += 1
        # 只统计"已签收"为签收成功，退回已签收不算签收
        if order.logistics_status == '已签收':
            cat_stats[cat]['signed'] += 1

    result = []
    for cat, stat in cat_stats.items():
        rate = (stat['signed'] / stat['total'] * 100) if stat['total'] > 0 else 0
        result.append({
            'category': cat,
            'total': stat['total'],
            'signed': stat['signed'],
            'unsigned': stat['total'] - stat['signed'],
            'rate': round(rate, 1)
        })

    result.sort(key=lambda x: x['rate'], reverse=True)
    return jsonify(result)


@bp.route('/admin/team_performance/api/salesman_detail')
@role_required('admin')
def salesman_order_detail():
    """获取业务员订单详情API"""
    salesman_name = request.args.get('salesman_name', '')
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)

    if not salesman_name or not year or not month:
        return jsonify({'error': '参数不全'}), 400

    # 查找业务员
    salesman = User.query.filter_by(name=salesman_name).first()
    if not salesman:
        return jsonify({'error': '业务员不存在'}), 404

    # 检查权限
    managed_user_ids = get_managed_user_ids(group_id)
    if salesman.id not in managed_user_ids:
        return jsonify({'error': '无权限查看该业务员数据'}), 403

    # 获取主产品类别列表
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}

    # 查询该业务员的所有订单（不按时间筛选，后面再判断归属月份）
    orders = Order.query.filter(
        Order.salesman_id == salesman.id,
        Order.status.in_(['draft', 'submitted', 'shipped']),
        Order.logistics_status != '退回已签收'
    ).order_by(Order.create_time.desc()).all()

    order_list = []
    for order in orders:
        cat = order.category or '未分类'
        if cat not in main_cat_names:
            continue  # 跳过非主产品
        amt = get_order_amount(order)
        if amt <= 0:
            continue  # 跳过金额为0的
        
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        # 合并状态：和发货单列表保持一致的逻辑
        if order.status == 'shipped':
            # 已发货订单，优先显示物流状态
            if order.logistics_status == '已签收':
                combined_status = '已签收'
            elif order.logistics_status == '退回已签收':
                combined_status = '退回已签收'
            elif order.logistics_status == '拒签':
                combined_status = '拒签'
            elif order.logistics_status == '派送中':
                combined_status = '派送中'
            elif order.logistics_status == '待派送':
                combined_status = '待派送'
            elif order.logistics_status == '运送中':
                combined_status = '运送中'
            elif order.logistics_status == '已揽收':
                combined_status = '已揽收'
            elif order.logistics_status == '待取件':
                combined_status = '待取件'
            else:
                combined_status = '已发货'
        else:
            # 未发货订单，直接显示订单状态
            combined_status = get_order_status_text(order.status)
        
        order_list.append({
            'id': order.id,
            'customer_name': order.customer_name or '未知客户',
            'category': cat,
            'paid_amount': order.paid_amount or '',
            'collect_amount': order.collect_amount or 0,
            'total_amount': amt,
            'status': combined_status,
            'logistics_status': order.logistics_status or '',
            'tracking_number': order.tracking_number or '',
            'create_time': order.create_time.strftime('%Y-%m-%d %H:%M') if order.create_time else '',
            'sign_time': order.sign_time.strftime('%Y-%m-%d %H:%M') if order.sign_time else ''
        })

    return jsonify({
        'salesman_name': salesman_name,
        'orders': order_list
    })


def get_order_status_text(status):
    """获取订单状态文本 - 和发货单列表保持一致"""
    status_map = {
        'draft': '草稿',
        'submitted': '待发货',
        'shipped': '已发货'
    }
    return status_map.get(status, status)


@bp.route('/admin/team_performance/api/export')
@role_required('admin')
def export_team_performance():
    """导出当月团队业绩 - 使用业绩报表模板"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)

    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month

    # 获取第一个激活的模板
    template = PerformanceReportTemplate.query.filter_by(is_active=True).first()
    if not template:
        return jsonify({'success': False, 'message': '没有找到业绩报表模板，请先配置模板！'}), 400

    # 检查模板文件是否存在
    if not template.filepath or not os.path.exists(template.filepath):
        return jsonify({'success': False, 'message': '模板文件不存在，请联系管理员重新上传模板！'}), 400

    # 获取模板配置
    field_mapping = json.loads(template.field_mapping) if template.field_mapping else {}

    # 直接查询所有主产品
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    allowed_categories = [c.name for c in main_categories]

    # 获取管理的用户ID列表
    user_ids = get_managed_user_ids(group_id)
    if not user_ids:
        return jsonify({'success': False, 'message': '没有找到可导出的业务员！'}), 400

    # 查询所有已发货、已签收的订单（不按时间筛选，后面再判断归属月份）
    query = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    )

    if allowed_categories:
        query = query.filter(Order.category.in_(allowed_categories))

    # 过滤：只有有订金或代收金额的订单才导出（金额>0），并且按归属月份筛选
    orders_to_export = []
    for order in query.all():
        # 判断订单归属月份（已签收订单按签收时间）
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
            if order_month == month and order_year == year:
                paid_num = 0
                if order.paid_amount:
                    paid_str = str(order.paid_amount)
                    match = re.match(r'[\d.]+', paid_str)
                    if match:
                        paid_num = float(match.group())
                collect_num = float(order.collect_amount or 0)
                # 只有金额>0才导出
                if paid_num > 0 or collect_num > 0:
                    orders_to_export.append(order)

    # 按组别、业务员名字排序
    # 先获取所有相关用户信息
    user_cache = {}
    for order in orders_to_export:
        if order.salesman_id not in user_cache:
            user_cache[order.salesman_id] = db.session.get(User, order.salesman_id)
    
    # 排序：先按组别，再按业务员姓名
    def get_sort_key(order):
        user = user_cache.get(order.salesman_id)
        group_name = user.group.name if (user and user.group) else ''
        user_name = user.name if user else ''
        return (group_name, user_name)
    
    orders = sorted(orders_to_export, key=get_sort_key)

    if not orders:
        return jsonify({'success': False, 'message': '没有找到符合条件的订单！'}), 400

    # 使用模板文件填充数据
    import io
    from openpyxl import load_workbook
    from openpyxl.styles import numbers

    # 加载模板文件
    wb = load_workbook(template.filepath)
    ws = wb.active

    # 获取表头（第一行）
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # 找出日期列的索引
    date_col_idx = None
    for idx, h in enumerate(headers):
        if field_mapping.get(h) == 'create_time':
            date_col_idx = idx + 1
            break

    # 从第二行开始填充数据
    row_num = 2
    for order in orders:
        row_data = []
        salesman = user_cache.get(order.salesman_id)
        for header in headers:
            order_field = field_mapping.get(header, '')
            if order_field == '__sales_amount__':
                paid_str = str(order.paid_amount or '')
                paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
                collect_num = float(order.collect_amount or 0)
                row_data.append(paid_num + collect_num)
            elif order_field == 'create_time' or order_field == 'sign_time':
                value = getattr(order, order_field, '') or ''
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M')
                row_data.append(value)
            elif order_field == 'salesman_name':
                row_data.append(salesman.name if salesman else '')
            elif order_field == 'group_name':
                if salesman and salesman.group:
                    row_data.append(salesman.group.name)
                else:
                    row_data.append('')
            elif order_field == 'product_info':
                row_data.append(order.product_info or order.category or '')
            elif order_field == 'total_amount':
                paid_str = str(order.paid_amount or '')
                paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
                collect_num = float(order.collect_amount or 0)
                row_data.append(paid_num + collect_num)
            elif order_field:
                value = getattr(order, order_field, '') or ''
                row_data.append(value)
            else:
                row_data.append('')
        ws.append(row_data)

        if date_col_idx:
            ws.cell(row=row_num, column=date_col_idx).number_format = 'yyyy-mm-dd'
        row_num += 1

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'团队业绩报表_{year}年{month}月.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ==================== 产品维度 API ====================
@bp.route('/admin/team_performance/api/product_overview')
@role_required('admin')
def product_overview():
    """产品维度概览API - 简化版，确保不会卡死"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group_id = request.args.get('group_id', type=int)
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    # 计算月份范围
    start_date = datetime(year, month, 1)
    end_date = datetime(year, month, monthrange(year, month)[1], 23, 59, 59)
    
    # 获取管理的用户ID列表
    user_ids = get_managed_user_ids(group_id)
    if not user_ids:
        return jsonify({
            'total_amount': 0,
            'total_count': 0,
            'signed_amount': 0,
            'signed_count': 0,
            'products': []
        })
    
    # 获取主产品类别列表
    main_categories = Category.query.filter_by(is_main_product=True, is_active=True).all()
    main_cat_names = {c.name for c in main_categories}
    
    # 获取所有订单（不按时间筛选，后面再判断归属月份）
    all_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status.in_(['draft', 'submitted', 'shipped'])
    ).all()
    
    # 获取已签收订单（不按时间筛选，后面再判断归属月份）
    signed_orders = Order.query.filter(
        Order.salesman_id.in_(user_ids),
        Order.status == 'shipped',
        Order.logistics_status == '已签收'
    ).all()
    
    # 统计产品数据
    product_stats = {}
    total_amount = 0
    total_count = 0
    signed_amount = 0
    signed_count = 0
    male_count = 0
    female_count = 0
    unknown_gender_count = 0
    
    # 统计所有订单
    for order in all_orders:
        category = order.category or '未知产品'
        # 只统计主产品
        if category not in main_cat_names:
            continue
        # 计算订单金额
        paid_str = str(order.paid_amount or '')
        paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
        collect_num = float(order.collect_amount or 0)
        amount = paid_num + collect_num
        
        if amount <= 0:
            continue  # 金额为0的订单不统计
        
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        if category not in product_stats:
            product_stats[category] = {
                'category': category,
                'total_amount': 0,
                'total_count': 0,
                'sign_rate_denominator': 0,  # 签收率分母：包含退回已签收
                'valid_amount': 0,
                'signed_amount': 0,
                'signed_count': 0,
                'male_count': 0,
                'female_count': 0,
                'unknown_gender_count': 0
            }
        
        # 判断是否是退回已签收（两种情况：1. logistics_status 是退回已签收；2. logistics_warning_remark 包含退回关键词且是已签收状态）
        is_returned = False
        if order.logistics_status == '退回已签收':
            is_returned = True
        elif order.logistics_status == '已签收' and order.logistics_warning_remark:
            return_keywords = ['退回', '拒收', '退件']
            is_returned = any(keyword in order.logistics_warning_remark for keyword in return_keywords)
        
        # 排除退回已签收和拒签的订单（跟个人维度统一）
        if is_returned or order.logistics_status == '拒签':
            continue
        
        # 签收率分母、总数统计（排除退回已签收）
        product_stats[category]['sign_rate_denominator'] += 1
        product_stats[category]['total_amount'] += amount
        product_stats[category]['total_count'] += 1
        product_stats[category]['valid_amount'] += amount
        
        # 总体统计
        total_amount += amount
        total_count += 1
        
        # 统计性别
        gender = order.gender or ''
        if gender == '男':
            product_stats[category]['male_count'] += 1
            male_count += 1
        elif gender == '女':
            product_stats[category]['female_count'] += 1
            female_count += 1
        else:
            product_stats[category]['unknown_gender_count'] += 1
            unknown_gender_count += 1
    
    # 统计已签收订单
    for order in signed_orders:
        category = order.category or '未知产品'
        # 只统计主产品
        if category not in main_cat_names:
            continue
        # 计算订单金额
        paid_str = str(order.paid_amount or '')
        paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
        collect_num = float(order.collect_amount or 0)
        amount = paid_num + collect_num
        
        if amount <= 0:
            continue
        
        # 判断订单归属月份（已签收订单按签收时间）
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        else:
            continue  # 没有签收时间的已签收订单不统计
        
        # 只统计当前查询月份的订单
        if order_month != month or order_year != year:
            continue
        
        # 判断是否是退回已签收（两种情况：1. logistics_status 是退回已签收；2. logistics_warning_remark 包含退回关键词）
        is_returned = False
        if order.logistics_status == '退回已签收':
            is_returned = True
        elif order.logistics_warning_remark:
            return_keywords = ['退回', '拒收', '退件']
            is_returned = any(keyword in order.logistics_warning_remark for keyword in return_keywords)
        
        if not is_returned and category in product_stats:
            product_stats[category]['signed_amount'] += amount
            product_stats[category]['signed_count'] += 1
            
            # 总体统计
            signed_amount += amount
            signed_count += 1
    
    # 转换为列表并按总金额排序
    products_list = list(product_stats.values())
    products_list.sort(key=lambda x: x['total_amount'], reverse=True)
    
    return jsonify({
        'total_amount': total_amount,
        'total_count': total_count,
        'signed_amount': signed_amount,
        'signed_count': signed_count,
        'products': products_list
    })


@bp.route('/admin/team_performance/api/product_orders')
@role_required('admin')
def get_product_orders():
    """获取指定产品的订单列表"""
    category = request.args.get('category', '')
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    group = request.args.get('group', '')
    
    if not year or not month:
        now = _now_bj()
        year, month = now.year, now.month
    
    # 不按时间筛选，后面再判断归属月份
    # 构建查询
    query = Order.query
    
    # 类别筛选
    if category:
        query = query.filter(Order.category == category)
    
    # 组别筛选
    if group:
        query = query.filter(Order.group_name == group)
    
    # 获取订单
    all_orders = query.order_by(Order.create_time.desc()).limit(1000).all()
    
    # 筛选归属月份正确的订单
    orders = []
    for order in all_orders:
        # 判断订单归属月份：以签收时间为主，没有签收时间时用创建时间
        order_month = None
        order_year = None
        if order.sign_time:
            order_month = order.sign_time.month
            order_year = order.sign_time.year
        elif order.create_time:
            order_month = order.create_time.month
            order_year = order.create_time.year
        
        # 只统计当前查询月份的订单
        if order_month == month and order_year == year:
            orders.append(order)
    
    # 转换为字典列表
    orders_data = []
    for order in orders:
        # 计算总金额（已付+代收）
        try:
            paid_str = str(order.paid_amount or '')
            paid_num = float(re.match(r'[\d.]+', paid_str).group()) if re.match(r'[\d.]+', paid_str) else 0
        except Exception:
            paid_num = 0
        
        collect_num = float(order.collect_amount or 0)
        total_amount = paid_num + collect_num
        
        # 只展示金额>0的订单
        if total_amount <= 0:
            continue
            
        try:
            create_time_str = order.create_time.isoformat() if order.create_time else None
        except Exception:
            create_time_str = None
            
        try:
            signed_time_str = order.sign_time.strftime('%Y-%m-%d') if order.sign_time else None
        except Exception:
            signed_time_str = None
            
        orders_data.append({
            'id': order.id,
            'group_name': order.group_name,
            'salesman_name': order.salesman_name,
            'customer_name': order.customer_name,
            'phone': order.phone,
            'tracking_number': order.tracking_number,
            'amount': total_amount,
            'product_info': order.product_info,
            'status': order.status,
            'logistics_status': order.logistics_status,
            'create_time': create_time_str,
            'signed_time': signed_time_str
        })
    
    return jsonify({
        'success': True,
        'orders': orders_data,
        'total': len(orders_data)
    })
