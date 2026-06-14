import hashlib
import base64
import json
import os
import time
import uuid
import zipfile
import tempfile
import requests
from datetime import datetime
from urllib.parse import quote

from config import (
    SF_APP_ID, SF_APP_SECRET, SF_API_URL, SF_SERVICE_CODE,
    DINGTALK_CORP_ID, DINGTALK_APP_KEY, DINGTALK_APP_SECRET, DINGTALK_AGENT_ID,
    DINGTALK_CHAT_ID, DINGTALK_UNION_ID, DINGTALK_SPACE_ID, DINGTALK_OPERATOR_ID, DINGTALK_WORKSPACE_ID
)
from models import db, Order, OrderReminder, _now_bj

# ============ 物流信息缓存 ============
CACHE_DIR = os.path.join(os.path.dirname(__file__), 'cache')
LOGISTICS_CACHE_DIR = os.path.join(CACHE_DIR, 'logistics')  # 已签收，按月存储
TEMP_CACHE_DIR = os.path.join(CACHE_DIR, 'temp')  # 未签收，临时缓存

def _ensure_cache_dirs():
    """确保缓存目录存在"""
    os.makedirs(LOGISTICS_CACHE_DIR, exist_ok=True)
    os.makedirs(TEMP_CACHE_DIR, exist_ok=True)

def _get_logistics_cache_path(tracking_number, is_signed=False, sign_time=None):
    """获取物流缓存文件路径"""
    _ensure_cache_dirs()
    if is_signed and sign_time:
        # 已签收：按月存储 cache/logistics/{year}/{month}/{tracking_number}.json
        year_dir = os.path.join(LOGISTICS_CACHE_DIR, str(sign_time.year))
        month_dir = os.path.join(year_dir, str(sign_time.month))
        os.makedirs(month_dir, exist_ok=True)
        return os.path.join(month_dir, f"{tracking_number}.json")
    else:
        # 未签收：临时缓存 cache/temp/{tracking_number}.json
        return os.path.join(TEMP_CACHE_DIR, f"{tracking_number}.json")

def save_logistics_cache(tracking_number, routes, is_signed=False, sign_time=None):
    """保存物流信息缓存"""
    cache_path = _get_logistics_cache_path(tracking_number, is_signed, sign_time)
    cache_data = {
        'tracking_number': tracking_number,
        'routes': routes,
        'is_signed': is_signed,
        'cache_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2)
    print(f"[CACHE] 保存物流缓存: {cache_path}")
    print(f"[CACHE] 保存内容: 单号={tracking_number}, is_signed={is_signed}, routes数量={len(routes)}")

def load_logistics_cache(tracking_number, is_signed=False, sign_time=None, create_time=None):
    """加载物流信息缓存
    
    Returns:
        dict with 'routes' and 'cache_time', or None if not found/expired
    """
    # 尝试常规路径
    cache_path = _get_logistics_cache_path(tracking_number, is_signed, sign_time)
    print(f"[CACHE] 尝试读取缓存: 单号={tracking_number}, is_signed={is_signed}, 路径={cache_path}")
    
    if os.path.exists(cache_path):
        return _try_load_cache(cache_path, is_signed)
    
    # 如果是已签收但没找到，而且 sign_time 是 None
    if is_signed and not sign_time:
        print(f"[CACHE] 已签收但 sign_time 为 None，尝试搜索缓存")
        # 搜索订单创建时间的前 2 个月、创建月、下一个月
        search_months = []
        if create_time:
            # 前 2 个月
            for i in [2, 1, 0, -1]:  # i=2: 前2个月，i=1: 前1个月，i=0: 当前月，i=-1: 下一个月
                target_month = create_time.month - i
                target_year = create_time.year
                while target_month < 1:
                    target_month += 12
                    target_year -= 1
                while target_month > 12:
                    target_month -= 12
                    target_year += 1
                search_months.append((target_year, target_month))
            print(f"[CACHE] 基于订单创建时间搜索: {search_months}")
        
        if search_months:
            import glob
            for year, month in search_months:
                search_pattern = os.path.join(LOGISTICS_CACHE_DIR, str(year), str(month), f"{tracking_number}.json")
                print(f"[CACHE] 搜索模式: {search_pattern}")
                matching_files = glob.glob(search_pattern)
                if matching_files:
                    cache_path = matching_files[0]
                    print(f"[CACHE] 找到缓存文件: {cache_path}")
                    return _try_load_cache(cache_path, is_signed)
        else:
            # 没有创建时间，搜索所有月份
            import glob
            search_pattern = os.path.join(LOGISTICS_CACHE_DIR, '*', '*', f"{tracking_number}.json")
            print(f"[CACHE] 无创建时间，搜索所有: {search_pattern}")
            matching_files = glob.glob(search_pattern)
            if matching_files:
                # 取最近修改的那个
                matching_files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
                cache_path = matching_files[0]
                print(f"[CACHE] 找到缓存文件: {cache_path}")
                return _try_load_cache(cache_path, is_signed)
    
    print(f"[CACHE] 缓存文件不存在: {cache_path}")
    return None

def _try_load_cache(cache_path, is_signed):
    """尝试加载缓存文件"""
    try:
        with open(cache_path, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
        
        print(f"[CACHE] 读取到缓存: routes数量={len(cache_data.get('routes', []))}")
        
        # 未签收的缓存检查是否过期（2小时）
        # 已签收和退回已签收都是最终状态，永久保留，不删除
        if not is_signed:
            cache_time = datetime.strptime(cache_data['cache_time'], '%Y-%m-%d %H:%M:%S')
            elapsed = (datetime.now() - cache_time).total_seconds()
            if elapsed > 2 * 3600:  # 2小时
                print(f"[CACHE] 缓存已过期: {cache_path}, elapsed={elapsed/3600:.1f}h")
                os.remove(cache_path)  # 删除过期缓存
                return None
        
        print(f"[CACHE] 命中缓存: {cache_path}")
        return cache_data
    except Exception as e:
        print(f"[CACHE] 读取缓存失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_logistics_with_cache(order, force_refresh=False):
    """获取物流信息
    
    逻辑：所有状态都缓存，非最终状态有效期2小时，已签收和退回已签收按月缓存（永久保留）
    强制刷新时直接调用API
    
    Args:
        order: 订单对象
        force_refresh: 是否强制刷新
    
    Returns:
        dict with 'routes', 'from_cache', 'is_signed'
    """
    print(f"[LOGISTICS_CACHE] get_logistics_with_cache 被调用: 单号={order.tracking_number}, force_refresh={force_refresh}")
    print(f"[LOGISTICS_CACHE] 当前状态: logistics_status={order.logistics_status}, sign_time={order.sign_time}, create_time={order.create_time}")
    
    if not order.tracking_number or order.express_type != '顺丰':
        print(f"[LOGISTICS_CACHE] 不是顺丰或无单号，返回空")
        return {'routes': [], 'from_cache': False, 'is_signed': False}
    
    tracking_number = order.tracking_number
    
    # 已签收和退回已签收都是最终状态，永久保留缓存
    final_statuses = ['已签收', '退回已签收']
    is_final_status = order.logistics_status in final_statuses
    sign_time = order.sign_time
    
    print(f"[LOGISTICS_CACHE] 最终状态判断: is_final_status={is_final_status}")
    
    # 非强制刷新时，先尝试读缓存
    if not force_refresh:
        print(f"[LOGISTICS_CACHE] 尝试读取缓存...")
        cache_data = load_logistics_cache(tracking_number, is_final_status, sign_time, order.create_time)
        if cache_data:
            print(f"[LOGISTICS_CACHE] 缓存命中，使用缓存数据")
            _update_order_status_from_routes(order, cache_data['routes'])
            return {
                'routes': cache_data['routes'],
                'from_cache': True,
                'is_signed': is_final_status
            }
        else:
            print(f"[LOGISTICS_CACHE] 缓存未命中，将调用API")
    
    # 调用顺丰API查询（强制刷新或缓存不存在/过期）
    phone_last4 = order.phone[-4:] if order.phone and len(order.phone) >= 4 else ''
    print(f"[LOGISTICS_CACHE] 调用顺丰API: phone_last4={phone_last4}")
    routes = get_sf_routes(tracking_number, phone_last4)
    print(f"[LOGISTICS_CACHE] API返回routes数量: {len(routes)}")
    
    # 更新订单状态
    _update_order_status_from_routes(order, routes)
    
    # 保存缓存（所有状态都保存，包括空结果）
    # 如果更新后变成最终状态，按最终状态的方式保存（永久保留）
    new_is_final = order.logistics_status in final_statuses
    print(f"[LOGISTICS_CACHE] 更新后的新状态: logistics_status={order.logistics_status}, new_is_final={new_is_final}")
    save_logistics_cache(tracking_number, routes, new_is_final, order.sign_time if new_is_final else None)
    
    return {
        'routes': routes,
        'from_cache': False,
        'is_signed': new_is_final
    }


def _update_order_status_from_routes(order, routes):
    """根据路由信息更新订单物流状态
    Returns:
        dict: {'success': bool, 'status': str or None}
    """
    if not routes:
        return {'success': False, 'status': None}

    # 取最后一条路由节点（最新的状态）
    latest = routes[-1] if routes else {}
    
    # 方式1（最高优先级）：优先使用顺丰官方返回的状态名称（最准确）
    # secondaryStatusName 和 firstStatusName 都是顺丰官方返回的中文状态
    new_status = latest.get('secondaryStatusName', '') or latest.get('firstStatusName', '')
    
    if not new_status:
        # 方式2（次优先级）：根据 secondaryStatusCode / firstStatusCode 状态码映射
        status_code = latest.get('secondaryStatusCode', '') or latest.get('firstStatusCode', '')
        if status_code:
            status_map = {
                '1': '已揽收',
                '101': '已揽收',
                '2': '运送中',
                '201': '运送中',
                '3': '派送中',
                '301': '派送中',
                '4': '已签收',
                '401': '已签收',
                '5': '拒签',
                '501': '拒签',
                '11': '待取件',
                '1101': '待取件',
            }
            new_status = status_map.get(status_code)
    
    if not new_status:
        # 方式3（最低优先级）：根据 remark 关键词匹配
        opcode = latest.get('opCode', '') or latest.get('opcode', '') or ''
        remark = latest.get('remark', '') or ''
        if '已签收' in remark:
            new_status = '已签收'
        elif '拒签' in remark:
            new_status = '拒签'
        elif '派送' in remark:
            new_status = '派送中'
        elif '运送' in remark or '发往' in remark or '离开' in remark or '分拣' in remark or '到达' in remark:
            new_status = '运送中'
        elif '收取' in remark:
            new_status = '已揽收'
    
    # 更新订单状态
    remark = latest.get('remark', '') or ''
    return_keywords = ['退回', '拒收', '无人签收', '退回寄件人', '返程']
    is_return = any(keyword in remark for keyword in return_keywords)
    
    # 检查物流异常关键词
    warning_keywords = ['派送不成功', '送不成功', '送不出去', '无法送达', '无法派送', '拒收', '退回', '异常', '滞留']
    is_warning = any(keyword in remark for keyword in warning_keywords)
    
    # 更新异常标识（无论状态是否变化，都要更新异常标识）
    from models import db
    if is_warning:
        # 检查是否之前已经有异常标记（避免重复发送通知）
        was_warning = order.logistics_warning
        order.logistics_warning = True
        order.logistics_warning_remark = remark
        print(f"[物流异常] 订单 {order.id}: 检测到异常 - {remark}")
        
        # 只有当异常标记从 False 变成 True 时，才发送通知（每个单据只发送一次）
        if not was_warning and order.salesman_id:
            try:
                from models import User, BroadcastNotification, NotificationReceipt
                
                salesman = db.session.get(User, order.salesman_id)
                if salesman:
                    # 查询一个存在的用户作为发送者
                    sender_id = order.salesman_id
                    if not db.session.get(User, sender_id):
                        first_user = db.session.query(User).first()
                        if first_user:
                            sender_id = first_user.id
                        else:
                            sender_id = None
                    
                    notification = BroadcastNotification(
                        title='📦 物流异常提醒',
                        content=f"[物流异常] 客户 {order.customer_name or '未知'} 的订单出现物流异常：\n\n快递单号：{order.tracking_number or '未填写'}\n收货地址：{order.address or '未知'}\n异常信息：{remark}\n\n请及时处理！",
                        priority='important',
                        target_type='user',
                        target_ids=str(order.salesman_id),
                        sender_id=sender_id,
                        status='sent',
                        sent_time=_now_bj()
                    )
                    db.session.add(notification)
                    db.session.flush()

                    receipt = NotificationReceipt(
                        notification_id=notification.id,
                        user_id=order.salesman_id
                    )
                    db.session.add(receipt)

                    print(f"[物流异常通知] 已保存通知给业务员 {salesman.name} (ID: {order.salesman_id})")
            except Exception as notify_err:
                print(f"[物流异常通知] 发送通知失败: {notify_err}")
                import traceback
                traceback.print_exc()
    else:
        order.logistics_warning = False
        order.logistics_warning_remark = None
    
    if new_status:
        # 如果remark包含退回关键词，并且是已签收相关状态，强制设置为退回已签收
        if is_return and new_status in ['已签收', '退回已签收']:
            target_status = '退回已签收'
            if order.logistics_status != target_status:
                old_status = order.logistics_status
                order.logistics_status = target_status
                from models import db
                db.session.commit()
                print(f"[物流状态更新] 订单 {order.id}: {old_status} -> 退回已签收 (remark: {remark})")
                # 记录签收时间
                accept_time = latest.get('acceptTime', '') or latest.get('accepttime', '') or ''
                if accept_time:
                    try:
                        order.sign_time = datetime.strptime(accept_time, '%Y-%m-%d %H:%M:%S')
                        db.session.commit()
                    except ValueError:
                        pass
            else:
                # 状态没变，但异常标识可能变了，需要commit
                db.session.commit()
            return {'success': True, 'status': target_status}  # 已更新为退回已签收，无需继续
        
        # 正常状态更新
        if new_status != order.logistics_status:
            old_status = order.logistics_status
            order.logistics_status = new_status
            
            if new_status in ['已签收', '退回已签收']:
                accept_time = latest.get('acceptTime', '') or latest.get('accepttime', '') or ''
                if accept_time:
                    from models import db
                    try:
                        order.sign_time = datetime.strptime(accept_time, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        pass
            
            from models import db
            db.session.commit()
            print(f"[物流状态更新] 订单 {order.id}: {old_status} -> {order.logistics_status} (remark: {remark})")
        else:
            # 状态没变，但异常标识可能变了，需要commit
            db.session.commit()

    return {'success': True, 'status': order.logistics_status}

# ============ uapis.cn 物流查询 ============
UAPIS_API_KEY = "uapi--5ndmrhw66hOT4V0WfAF96QPEZh_ruE2BtgHPLlF"
UAPIS_API_URL = "https://uapis.cn/api/v1/misc/tracking/query"

def query_logistics_uapis(tracking_number, phone_last4=None):
    """通过 uapis.cn 查询物流信息 (无缓存，直接查询)"""
    try:
        params = {
            'tracking_number': tracking_number
        }
        if phone_last4:
            params['phone'] = phone_last4
        
        headers = {
            'Authorization': f'Bearer {UAPIS_API_KEY}',
            'Accept': 'application/json'
        }
        
        print(f"uapis.cn 查询: tracking_number={tracking_number}, phone_last4={phone_last4}")
        print(f"uapis.cn 请求URL: {UAPIS_API_URL}")
        response = requests.get(UAPIS_API_URL, params=params, headers=headers, timeout=10)
        response.raise_for_status()
        result = response.json()
        print(f"uapis.cn 响应: {result}")
        
        # 解析结果格式，转换为与顺丰API兼容的格式
        # 支持多种响应结构
        data = result
        
        # 如果有 data 字段，先取 data
        if 'data' in data:
            data = data['data']
        
        # 调试输出
        print(f"[DEBUG] 完整响应: {result}")
        print(f"[DEBUG] data 字段: {data}")
        
        # 检查是否是成功响应
        # 检查返回码是否为成功
        is_success_code = (result.get('code') == 200 or result.get('success') is True)
        # 或者直接检查是否有 tracks 或 tracking_number
        has_data = ('tracking_number' in data or 'tracks' in data or 'status' in data)
        success = is_success_code or has_data
        
        print(f"[DEBUG] 是否成功响应: {success}")
        
        if success:
            # 获取轨迹列表，支持多种字段名
            track_list = []
            
            # 优先从 data 里找
            for key in ['tracks', 'list', 'traces', 'data']:
                if key in data and isinstance(data[key], list):
                    track_list = data[key]
                    print(f"[DEBUG] 从 data.{key} 找到轨迹列表")
                    break
            
            # 如果外层没找到，在原始 result 里找
            if not track_list:
                for key in ['tracks', 'list', 'traces', 'data']:
                    if key in result and isinstance(result[key], list):
                        track_list = result[key]
                        print(f"[DEBUG] 从 result.{key} 找到轨迹列表")
                        break
            
            print(f"[DEBUG] 最终轨迹列表长度: {len(track_list)}")
            if track_list:
                print(f"[DEBUG] 第一条轨迹: {track_list[0]}")
            
            # 获取物流状态
            status_text = data.get('status', result.get('status', ''))
            status_code = data.get('status_code', result.get('status_code', ''))
            
            # 转换为我们系统使用的状态
            status_map = {
                'pending': '待揽收',
                'picked_up': '已揽收',
                'in_transit': '运送中',
                'out_for_delivery': '派送中',
                'delivered': '已签收',
                'exception': '异常',
                'unknown': '未知'
            }
            
            # 使用 status_code 映射，或者直接使用 status_text
            system_status = status_map.get(status_code, status_text[:20] if status_text else '')
            
            # 如果没有状态，尝试从最后一条轨迹推断
            if not system_status and track_list:
                last_track = track_list[-1]
                context = last_track.get('context', '') or last_track.get('remark', '')
                if '签收' in context:
                    system_status = '已签收'
                elif '派送' in context:
                    system_status = '派送中'
                elif '揽收' in context or '取件' in context:
                    system_status = '已揽收'
                elif '运输' in context or '发往' in context or '离开' in context:
                    system_status = '运送中'
            
            routes = []
            if track_list:
                for track in track_list:
                    # 获取时间，支持多种字段名
                    time_str = ''
                    for key in ['time', 'acceptTime', 'datetime', 'date', 'timestamp']:
                        if key in track:
                            time_str = track[key]
                            break
                    
                    # 获取描述，支持多种字段名
                    context = ''
                    for key in ['context', 'remark', 'desc', 'status']:
                        if key in track:
                            context = track[key]
                            break
                    
                    if context:
                        routes.append({
                            'acceptTime': time_str,
                            'remark': context,
                            'secondaryStatusName': system_status  # 使用简短的状态
                        })
            
            print(f"解析结果: {len(routes)} 条轨迹, 状态: {system_status}")
            
            return {
                'success': True, 
                'routes': routes,
                'status': system_status
            }
        else:
            print(f"uapis.cn 查询返回失败: {result}")
            return {'success': False, 'error': result.get('message', '查询失败')}
            
    except Exception as e:
        print(f"uapis.cn 查询失败: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


def get_logistics_uapis_with_cache(order, force_refresh=False, update_db=True):
    """通过 uapis.cn 获取物流信息 (带缓存)
    
    逻辑：所有状态都缓存，非最终状态有效期2小时，已签收按月缓存（永久保留）
    强制刷新时直接调用API
    
    Args:
        order: 订单对象
        force_refresh: 是否强制刷新
        update_db: 是否更新数据库（默认是，普通查看详情时设为False）
    
    Returns:
        dict with 'routes', 'from_cache', 'status'
    """
    print(f"[uapis 缓存] 调用 get_logistics_uapis_with_cache: tracking={order.tracking_number}, force_refresh={force_refresh}, update_db={update_db}")
    if not order.tracking_number:
        return {'routes': [], 'from_cache': False, 'status': ''}
    
    tracking_number = order.tracking_number
    
    # 已签收是最终状态，永久保留缓存
    final_statuses = ['已签收']
    is_final_status = order.logistics_status in final_statuses
    sign_time = order.sign_time
    print(f"[uapis 缓存] 当前状态: logistics_status={order.logistics_status}, is_final_status={is_final_status}, sign_time={sign_time}, create_time={order.create_time}")
    
    # 非强制刷新时，先尝试读缓存
    if not force_refresh:
        cache_data = load_logistics_cache(tracking_number, is_final_status, sign_time, order.create_time)
        if cache_data:
            print(f"[uapis 缓存] 命中缓存: {tracking_number}, routes数量={len(cache_data['routes'])}")
            # 从缓存中更新状态
            if cache_data['routes'] and update_db:
                _update_order_status_from_routes(order, cache_data['routes'])
            return {
                'routes': cache_data['routes'],
                'from_cache': True,
                'status': order.logistics_status or ''
            }
        else:
            print(f"[uapis 缓存] 未命中缓存或缓存已过期: {tracking_number}")
    
    # 调用 uapis.cn API查询（强制刷新或缓存不存在/过期）
    print(f"[uapis 缓存] 调用 uapis.cn API: {tracking_number}")
    phone_last4 = order.phone[-4:] if order.phone and len(order.phone) >= 4 else None
    result = query_logistics_uapis(tracking_number, phone_last4)
    
    if not result['success']:
        print(f"[uapis 缓存] uapis.cn API调用失败: {result.get('error')}")
        return {
            'routes': [],
            'from_cache': False,
            'status': order.logistics_status or '',
            'error': result.get('error')
        }
    
    # 使用 uapis.cn 返回的状态更新
    routes = result['routes']
    status = result.get('status', '')
    print(f"[uapis 缓存] uapis.cn API返回: status={status}, routes数量={len(routes)}")
    
    # 只有 update_db 为 True 时才更新数据库
    if update_db:
        print(f"[uapis 缓存] 更新数据库")
        if status:
            order.logistics_status = status
            # 如果是已签收，记录签收时间
            if status == '已签收' and routes:
                last_track = routes[0]
                time_str = last_track.get('acceptTime', '')
                if time_str:
                    try:
                        from datetime import datetime
                        order.sign_time = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
                    except:
                        pass
        elif routes:
            # 从轨迹推断状态
            _update_order_status_from_routes(order, routes)
    
    # 保存缓存（所有状态都保存，包括空结果）
    # 如果有状态，按状态判断是否为最终状态
    new_is_final = status in final_statuses if status else (order.logistics_status in final_statuses)
    # 如果不更新数据库，使用旧的 sign_time
    cache_sign_time = None
    if new_is_final:
        if status == '已签收' and routes:
            # 尝试从轨迹获取时间
            last_track = routes[0]
            time_str = last_track.get('acceptTime', '')
            if time_str:
                try:
                    from datetime import datetime
                    cache_sign_time = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
                except:
                    pass
        # 如果没找到，使用订单已有的 sign_time
        if not cache_sign_time:
            cache_sign_time = order.sign_time
    
    print(f"[uapis 缓存] 保存缓存: tracking={tracking_number}, new_is_final={new_is_final}, cache_sign_time={cache_sign_time}")
    save_logistics_cache(tracking_number, routes, new_is_final, cache_sign_time)
    
    return {
        'routes': routes,
        'from_cache': False,
        'status': status or order.logistics_status or ''
    }

# ============ 下载令牌管理 ============
_download_tokens = {}


def create_download_token(file_path, file_name, expire_seconds=3600):
    """创建临时下载令牌"""
    token = str(uuid.uuid4())
    _download_tokens[token] = {
        'file_path': file_path,
        'file_name': file_name,
        'expire': time.time() + expire_seconds,
    }
    return token


def get_download_info(token):
    """根据令牌获取下载信息"""
    info = _download_tokens.get(token)
    if not info:
        return None
    if time.time() > info['expire']:
        del _download_tokens[token]
        return None
    return info


# ============ 顺丰物流 API ============

def _generate_sf_msg_digest(msg_data, timestamp, check_word):
    """
    生成顺丰API msgDigest签名
    按照顺丰文档：
    1. 拼接字符串：msgData + timestamp + checkWord
    2. UTF-8编码
    3. MD5加密（得到字节数组）
    4. Base64编码（对MD5字节数组编码）
    """
    # 拼接字符串
    str_to_sign = f"{msg_data}{timestamp}{check_word}"
    # MD5加密（得到字节数组）
    md5_bytes = hashlib.md5(str_to_sign.encode('utf-8')).digest()
    # Base64编码（对字节数组编码）
    return base64.b64encode(md5_bytes).decode('utf-8')


def call_sf_api(service_code, msg_data):
    """调用顺丰API（使用partnerID/requestID/msgDigest方式）"""
    # 时间戳精确到毫秒（14位）
    timestamp = str(int(time.time() * 1000))
    request_id = str(uuid.uuid4())
    msg_data_json = json.dumps(msg_data, ensure_ascii=False)

    headers = {
        'Content-Type': 'application/x-www-form-urlencoded;charset=UTF-8'
    }

    # 使用顺丰标准API格式
    data = {
        'partnerID': SF_APP_ID,  # 顾客编码
        'requestID': request_id,  # 请求唯一号
        'serviceCode': service_code,
        'timestamp': timestamp,
        'msgDigest': _generate_sf_msg_digest(msg_data_json, timestamp, SF_APP_SECRET),
        'msgData': msg_data_json
    }

    try:
        resp = requests.post(SF_API_URL, data=data, headers=headers, timeout=30)
        result = resp.json()

        if result.get('apiResultCode') == 'A1000':
            return {'success': True, 'data': result.get('apiResultData', {})}
        else:
            return {'success': False, 'error': result.get('apiErrorMsg', '未知错误'), 'raw': result}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def get_sf_routes(tracking_number, phone_last4=''):
    """
    查询顺丰物流路由信息

    Args:
        tracking_number: 顺丰运单号
        phone_last4: 收件人手机号后4位（可选，用于验证）

    Returns:
        list: 路由节点列表
    """
    # 构建请求参数
    msg_data = {
        'language': '0',
        'trackingType': '1',
        'trackingNumber': [tracking_number],
        'methodType': '1'
    }

    # 如果有手机号后4位，添加到请求中
    if phone_last4:
        msg_data['checkPhoneNo'] = phone_last4

    print(f"[SF_API] 查询路由: tracking_number={tracking_number}, phone_last4={phone_last4}")
    result = call_sf_api(SF_SERVICE_CODE, msg_data)

    if result['success']:
        # API返回的data可能是字符串，需要解析
        data = result['data']
        if isinstance(data, str):
            data = json.loads(data)
        if data.get('success'):
            # 解析路由响应
            route_resps = data.get('msgData', {}).get('routeResps', [])
            print(f"[SF_API] routeResps数量: {len(route_resps)}")
            if route_resps and len(route_resps) > 0:
                routes = route_resps[0].get('routes', [])
                print(f"[SF_API] 原始路由数量: {len(routes)}")
                if routes:
                    print(f"[SF_API] 第一条时间: {routes[0].get('acceptTime', 'N/A')}")
                    print(f"[SF_API] 最后一条时间: {routes[-1].get('acceptTime', 'N/A')}")
                    print(f"[SF_API] 第一条remark: {routes[0].get('remark', '')[:50]}")
                    print(f"[SF_API] 最后一条remark: {routes[-1].get('remark', '')[:50]}")
                return routes

    print(f"[SF_API] 未获取到路由信息")
    return []


def get_sf_routes_batch(tracking_numbers, phone_last4_list=None):
    """
    批量查询顺丰物流路由信息（最多10条）

    Args:
        tracking_numbers: 顺丰运单号列表（最多10个）
        phone_last4_list: 收件人手机号后4位列表（可选，需与tracking_numbers顺序对应）

    Returns:
        dict: {tracking_number: routes_list} 路由字典
    """
    if not tracking_numbers:
        return {}

    # 限制最多10条
    tracking_numbers = tracking_numbers[:10]
    if phone_last4_list:
        phone_last4_list = phone_last4_list[:10]

    # 构建请求参数
    msg_data = {
        'language': '0',
        'trackingType': '1',
        'trackingNumber': tracking_numbers,
        'methodType': '1'
    }

    # 如果有手机号后4位，添加到请求中（逗号分隔的字符串，与单号顺序对应）
    if phone_last4_list:
        msg_data['checkPhoneNo'] = ','.join(phone_last4_list)

    print(f"[SF_API] 批量查询路由: tracking_numbers={tracking_numbers}, phone_last4_list={phone_last4_list}")
    result = call_sf_api(SF_SERVICE_CODE, msg_data)
    print(f"[SF_API] 批量查询原始完整返回: {result}")

    # 返回字典：{单号: 路由列表}
    routes_dict = {}

    if result['success']:
        data = result['data']
        print(f"[SF_API] 批量查询data字段: {data}")
        if isinstance(data, str):
            print(f"[SF_API] data是字符串，尝试解析")
            data = json.loads(data)
            print(f"[SF_API] 解析后data: {data}")
        if data.get('success'):
            route_resps = data.get('msgData', {}).get('routeResps', [])
            print(f"[SF_API] 批量查询返回routeResps数量: {len(route_resps)}")
            print(f"[SF_API] routeResps内容: {route_resps}")
            for route_resp in route_resps:
                # 每个route_resp包含mailNo字段（顺丰API返回的是mailNo，不是trackingNumber）
                tn = route_resp.get('mailNo', '') or route_resp.get('trackingNumber', '')
                print(f"[SF_API] route_resp中的单号: {tn}")
                if tn and tn in tracking_numbers:
                    routes = route_resp.get('routes', [])
                    routes_dict[tn] = routes
                    print(f"[SF_API] 单号{tn}路由数量: {len(routes)}, routes内容: {routes}")
            return routes_dict

    print(f"[SF_API] 批量查询未获取到路由信息")
    return {tn: [] for tn in tracking_numbers}


def update_single_order_logistics(order):
    """更新单个订单的物流状态
    
    策略：
    - 优先读缓存，有效期2小时（非已签收）或按月（已签收）
    - 缓存不存在/过期，调用API获取最新状态
    - 空结果也保存缓存
    """
    try:
        print(f"[UPDATE_LOGISTICS] 开始更新订单 {order.id}, tracking={order.tracking_number}, express={order.express_type}")
        if not order.tracking_number or order.express_type != '顺丰':
            print(f"[UPDATE_LOGISTICS] 跳过: 非顺丰或无单号")
            return {'success': False, 'error': '非顺丰订单或无运单号'}

        # 先尝试读缓存
        is_signed = order.logistics_status == '已签收'
        sign_time = order.sign_time
        cache_data = load_logistics_cache(order.tracking_number, is_signed, sign_time)
        if cache_data:
            print(f"[UPDATE_LOGISTICS] 使用缓存数据")
            _update_order_status_from_routes(order, cache_data['routes'])
            return {'success': True, 'status': order.logistics_status}
        
        # 缓存不存在/过期，调用API
        phone_last4 = order.phone[-4:] if order.phone and len(order.phone) >= 4 else ''
        print(f"[UPDATE_LOGISTICS] 调用顺丰API查询, phone_last4={phone_last4}")
        routes = get_sf_routes(order.tracking_number, phone_last4)

        # 取最后一条路由节点（最新的状态）
        latest = routes[-1] if routes else {}
        print(f"[UPDATE_LOGISTICS] 完整路由节点: {latest}")

        # 方式1（最高优先级）：优先使用顺丰官方返回的状态名称（最准确）
        # secondaryStatusName 和 firstStatusName 都是顺丰官方返回的中文状态
        status_name = latest.get('secondaryStatusName', '') or latest.get('firstStatusName', '')
        if status_name:
            new_status = status_name
            print(f"[UPDATE_LOGISTICS] 方式1（顺丰官方状态名称）匹配成功: {new_status}")
        else:
            # 方式2（次优先级）：根据 secondaryStatusCode / firstStatusCode 状态码映射
            status_code = latest.get('secondaryStatusCode', '') or latest.get('firstStatusCode', '')
            if status_code:
                status_map = {
                    '1': '已揽收',
                    '101': '已揽收',
                    '2': '运送中',
                    '201': '运送中',
                    '3': '派送中',
                    '301': '派送中',
                    '4': '已签收',
                    '401': '已签收',
                    '5': '拒签',
                    '501': '拒签',
                    '11': '待取件',
                    '1101': '待取件',
                }
                new_status = status_map.get(status_code)
                if new_status:
                    print(f"[UPDATE_LOGISTICS] 方式2（状态码映射）匹配成功: {status_code} -> {new_status}")
                else:
                    print(f"[UPDATE_LOGISTICS] 方式2未匹配到状态码: {status_code}")
            else:
                new_status = None
            
            # 方式3（最低优先级）：如果以上都不行，根据 remark 关键词匹配
            if not new_status:
                opcode = latest.get('opCode', '') or latest.get('opcode', '') or ''
                remark = latest.get('remark', '') or ''
                print(f"[UPDATE_LOGISTICS] 方式3尝试关键词匹配: opCode={opcode!r}, remark={remark[:50]!r}")
                if '已签收' in remark:
                    new_status = '已签收'
                elif '拒签' in remark:
                    new_status = '拒签'
                elif '派送' in remark:
                    new_status = '派送中'
                elif '运送' in remark or '发往' in remark or '离开' in remark or '分拣' in remark or '到达' in remark:
                    new_status = '运送中'
                elif '收取' in remark:
                    new_status = '已揽收'
                print(f"[UPDATE_LOGISTICS] 方式3匹配结果: new_status={new_status}")

        old_status = order.logistics_status
        print(f"[UPDATE_LOGISTICS] 订单 {order.id}: old_status={old_status}, new_status={new_status}")
        
        # 检查退回关键词
        remark = latest.get('remark', '') or ''
        return_keywords = ['退回', '拒收', '无人签收', '退回寄件人', '返程']
        is_return = any(keyword in remark for keyword in return_keywords)
        
        # 检查物流异常关键词
        warning_keywords = ['派送不成功', '送不成功', '送不出去', '无法送达', '无法派送', '拒收', '退回', '异常', '滞留']
        is_warning = any(keyword in remark for keyword in warning_keywords)
        
        # 更新异常标识
        if is_warning:
            order.logistics_warning = True
            order.logistics_warning_remark = remark
            print(f"[物流异常] 订单 {order.id}: 检测到异常 - {remark}")
        else:
            order.logistics_warning = False
            order.logistics_warning_remark = None
        
        if new_status:
            # 如果当前状态是已签收，但remark包含退回关键词，需要更新为退回已签收
            if old_status == '已签收' and is_return and new_status == '已签收':
                print(f"[UPDATE_LOGISTICS] 订单 {order.id}: 已签收 -> 退回已签收 (remark: {remark})")
                order.logistics_status = '退回已签收'
                accept_time = latest.get('acceptTime', '') or latest.get('accepttime', '') or ''
                if accept_time:
                    try:
                        order.sign_time = datetime.strptime(accept_time, '%Y-%m-%d %H:%M:%S')
                        print(f"[UPDATE_LOGISTICS] 订单 {order.id} 签收时间: {order.sign_time}")
                    except ValueError:
                        print(f"[UPDATE_LOGISTICS] 签收时间解析失败: {accept_time}")
                db.session.commit()
                print(f"[UPDATE_LOGISTICS] 订单 {order.id}: 提交数据库 commit")
                # 保存缓存
                save_logistics_cache(order.tracking_number, routes, False, None)
                return {'success': True, 'status': order.logistics_status}
            
            # 正常状态更新
            if new_status != old_status:
                print(f"[UPDATE_LOGISTICS] 订单 {order.id}: 需要更新状态")
                order.logistics_status = new_status
                # 已签收时记录签收时间（取最新路由节点的acceptTime）
                if new_status == '已签收':
                    accept_time = latest.get('acceptTime', '') or latest.get('accepttime', '') or ''
                    if accept_time:
                        try:
                            order.sign_time = datetime.strptime(accept_time, '%Y-%m-%d %H:%M:%S')
                            print(f"[UPDATE_LOGISTICS] 订单 {order.id} 签收时间: {order.sign_time}")
                        except ValueError:
                            print(f"[UPDATE_LOGISTICS] 签收时间解析失败: {accept_time}")
                print(f"[UPDATE_LOGISTICS] 订单 {order.id}: 提交数据库 commit")
                db.session.commit()
                print(f"[UPDATE_LOGISTICS] 订单 {order.id} 状态更新: {old_status} -> {new_status}")
            else:
                print(f"[UPDATE_LOGISTICS] 订单 {order.id}: 状态相同，无需更新")
        else:
            print(f"[UPDATE_LOGISTICS] 订单 {order.id}: new_status 为 None，无法更新")
        
        # 保存缓存（所有状态都保存，包括空结果）
        new_is_signed = order.logistics_status == '已签收'
        save_logistics_cache(order.tracking_number, routes, new_is_signed, order.sign_time if new_is_signed else None)
        
        return {'success': True, 'status': order.logistics_status}
    except Exception as e:
        print(f"[UPDATE_LOGISTICS] 订单 {order.id} 更新失败: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


def update_sf_logistics(order_ids=None, app=None):
    """批量更新顺丰物流信息（支持定时任务调用）

    Args:
        order_ids: 要更新的订单ID列表，None表示更新所有顺丰订单
        app: Flask应用实例（用于app_context）
    """
    # 兼容早期把 Flask app 当第一个参数传入的调用方式
    try:
        from flask import Flask
        if isinstance(order_ids, Flask):
            app = order_ids
            order_ids = None
    except Exception:
        pass

    def _do_update():
        query = Order.query.filter(
            Order.express_type == '顺丰',
            Order.tracking_number.isnot(None)
        )
        if order_ids:
            query = query.filter(Order.id.in_(order_ids))

        orders = query.all()
        updated = 0
        failed = 0

        for order in orders:
            result = update_single_order_logistics(order)
            if result['success']:
                updated += 1
            else:
                failed += 1

        return {'updated': updated, 'failed': failed, 'total': len(orders)}

    # 尝试自动获取Flask app（支持两种调用方式）
    if app is None:
        try:
            from app import app as _flask_app
            app = _flask_app
        except Exception:
            pass

    if app:
        with app.app_context():
            return _do_update()
    else:
        # 最后的兜底：尝试从当前模块的db获取session
        from flask import current_app
        with current_app.app_context():
            return _do_update()


# ============ 钉钉API Token缓存 ============
_dingtalk_token_cache = {'token': None, 'expire_time': 0}
_dingtalk_api_token_cache = {'token': None, 'expire_time': 0}

# 钉钉在线表格配置（从config导入）
DINGTALK_WORKSPACE_ID = None  # 将在函数中从config导入
DINGTALK_OPERATOR_ID = None  # 将在函数中从config导入

def _get_workspace_config():
    """获取工作空间配置"""
    global DINGTALK_WORKSPACE_ID, DINGTALK_OPERATOR_ID
    if DINGTALK_WORKSPACE_ID is None:
        from config import DINGTALK_WORKSPACE_ID as WS, DINGTALK_OPERATOR_ID as OP
        DINGTALK_WORKSPACE_ID = WS
        DINGTALK_OPERATOR_ID = OP
    return DINGTALK_WORKSPACE_ID, DINGTALK_OPERATOR_ID


def _get_dingtalk_api_token():
    """获取钉钉v1.0 API访问令牌"""
    now = time.time()
    if _dingtalk_api_token_cache['token'] and now < _dingtalk_api_token_cache['expire_time']:
        return _dingtalk_api_token_cache['token']

    url = 'https://api.dingtalk.com/v1.0/oauth2/accessToken'
    resp = requests.post(url, json={
        'appKey': DINGTALK_APP_KEY,
        'appSecret': DINGTALK_APP_SECRET
    }, timeout=10)
    result = resp.json()

    if 'accessToken' in result:
        _dingtalk_api_token_cache['token'] = result['accessToken']
        _dingtalk_api_token_cache['expire_time'] = now + 5000
        return result['accessToken']
    raise Exception(f"获取钉钉API Token失败: {result}")


def _get_dingtalk_token():
    """获取钉钉旧版API访问令牌"""
    now = time.time()
    if _dingtalk_token_cache['token'] and now < _dingtalk_token_cache['expire_time']:
        return _dingtalk_token_cache['token']

    url = 'https://oapi.dingtalk.com/gettoken'
    params = {
        'appkey': DINGTALK_APP_KEY,
        'appsecret': DINGTALK_APP_SECRET
    }
    resp = requests.get(url, params=params, timeout=10)
    result = resp.json()

    if result.get('errcode') == 0:
        _dingtalk_token_cache['token'] = result['access_token']
        _dingtalk_token_cache['expire_time'] = now + result.get('expires_in', 7200) - 300
        return result['access_token']
    raise Exception(f"获取钉钉Token失败: {result}")



# ============ 钉钉存储空间功能（支持在线预览编辑） ============
_dingtalk_space_cache = {'space_id': None}


def _get_or_create_storage_space():
    """获取或创建钉钉存储空间"""
    if _dingtalk_space_cache['space_id']:
        return _dingtalk_space_cache['space_id']
    
    from config import DINGTALK_OPERATOR_ID
    token = _get_dingtalk_api_token()
    
    # 先尝试查询已有空间
    url = f'https://api.dingtalk.com/v1.0/storage/spaces?unionId={DINGTALK_OPERATOR_ID}'
    headers = {'x-acs-dingtalk-access-token': token}
    resp = requests.get(url, headers=headers, timeout=10)
    result = resp.json()
    
    if 'spaces' in result and len(result['spaces']) > 0:
        space_id = result['spaces'][0]['id']
        _dingtalk_space_cache['space_id'] = space_id
        print(f"[DEBUG] 使用已有存储空间: {space_id}")
        return space_id
    
    # 创建新空间
    create_url = f'https://api.dingtalk.com/v1.0/storage/spaces?unionId={DINGTALK_OPERATOR_ID}'
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    payload = {
        'option': {
            'name': '发货单存储空间',
            'scene': 'deliveryorder',
            'sceneId': '001',
            'ownerType': 'APP'
        }
    }
    
    resp = requests.post(create_url, headers=headers, json=payload, timeout=10)
    result = resp.json()
    
    if 'space' in result:
        space_id = result['space']['id']
        _dingtalk_space_cache['space_id'] = space_id
        print(f"[DEBUG] 创建存储空间成功: {space_id}")
        return space_id
    else:
        raise Exception(f"创建存储空间失败: {result}")


def _upload_file_to_storage(file_path, file_name):
    """
    上传文件到钉钉存储空间（使用 v2.0 API）
    流程: 1.获取上传信息 -> 2.上传到OSS -> 3.提交文件
    """
    from config import DINGTALK_OPERATOR_ID
    import os
    import hashlib
    import uuid
    
    space_id = _get_or_create_storage_space()
    token = _get_dingtalk_api_token()
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    
    file_size = os.path.getsize(file_path)
    print(f"[DEBUG] 文件大小: {file_size} bytes")
    
    # 读取文件内容
    with open(file_path, 'rb') as f:
        file_content = f.read()
    file_md5 = hashlib.md5(file_content).hexdigest()
    print(f"[DEBUG] 文件MD5: {file_md5}")
    
    # 空间根目录的 uuid（需要先获取或使用固定值）
    # 钉钉存储空间的根目录 uuid 就是 space_id 本身
    parent_dentry_uuid = space_id
    print(f"[DEBUG] 使用 parentDentryUuid: {parent_dentry_uuid}")
    
    # 步骤1: 获取上传信息
    upload_url = f'https://api.dingtalk.com/v2.0/storage/spaces/files/{parent_dentry_uuid}/uploadInfos/query?unionId={DINGTALK_OPERATOR_ID}'
    payload = {
        'protocol': 'HEADER_SIGNATURE',
        'option': {
            'storageDriver': 'DINGTALK',
            'preCheckParam': {
                'size': file_size,
                'name': file_name
            }
        }
    }
    
    print(f"[DEBUG] 获取上传信息...")
    resp = requests.post(upload_url, headers=headers, json=payload, timeout=10)
    result = resp.json()
    
    print(f"[DEBUG] 上传信息响应: {result}")
    
    if 'uploadKey' not in result:
        raise Exception(f"获取上传信息失败: {result}")
    
    upload_key = result['uploadKey']
    resource_url = result['headerSignatureInfo']['resourceUrls'][0]
    oss_headers = result['headerSignatureInfo'].get('headers', {})
    print(f"[DEBUG] uploadKey: {upload_key}")
    print(f"[DEBUG] resourceUrl: {resource_url}")
    
    # 步骤2: 上传文件到OSS
    print(f"[DEBUG] 上传文件到OSS...")
    oss_headers['Content-Type'] = 'application/octet-stream'
    oss_resp = requests.put(resource_url, headers=oss_headers, data=file_content, timeout=60)
    
    print(f"[DEBUG] OSS响应状态: {oss_resp.status_code}")
    
    if oss_resp.status_code not in [200, 201, 204]:
        raise Exception(f"上传到OSS失败: {oss_resp.status_code} - {oss_resp.text}")
    
    print(f"[DEBUG] OSS上传成功")
    
    # 步骤3: 提交文件
    commit_url = f'https://api.dingtalk.com/v2.0/storage/spaces/files/{parent_dentry_uuid}/commit?unionId={DINGTALK_OPERATOR_ID}'
    commit_payload = {
        'uploadKey': upload_key,
        'name': file_name,
        'option': {
            'size': file_size,
            'conflictStrategy': 'AUTO_RENAME',
            'convertToOnlineDoc': True  # 转换为在线文档，支持在线编辑
        }
    }
    
    print(f"[DEBUG] 提交文件...")
    commit_resp = requests.post(commit_url, headers=headers, json=commit_payload, timeout=10)
    commit_result = commit_resp.json()
    
    print(f"[DEBUG] 提交响应: {commit_result}")
    
    if 'dentry' not in commit_result:
        raise Exception(f"提交文件失败: {commit_result}")
    
    dentry = commit_result['dentry']
    dentry_id = dentry.get('id')
    print(f"[DEBUG] 文件提交成功, dentryId: {dentry_id}")
    
    return space_id, dentry_id


def _get_file_preview_url(space_id, dentry_id, open_type='EDIT'):
    """获取文件预览/编辑链接"""
    from config import DINGTALK_OPERATOR_ID
    token = _get_dingtalk_api_token()
    
    url = f'https://api.dingtalk.com/v1.0/storage/spaces/{space_id}/dentries/{dentry_id}/openInfos/query?unionId={DINGTALK_OPERATOR_ID}'
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    
    payload = {
        'option': {
            'type': open_type,  # PREVIEW 或 EDIT
            'waterMark': False,
            'checkLogin': True
        }
    }
    
    resp = requests.post(url, headers=headers, json=payload, timeout=10)
    result = resp.json()
    
    if 'url' in result:
        return result['url']
    else:
        raise Exception(f"获取预览链接失败: {result}")


def send_dingtalk_file_preview(file_path, title, order_count=0):
    """
    上传文件到钉钉存储空间并发送在线预览/编辑链接到群
    实现图二那种可在线预览编辑的效果
    """
    print(f"[DEBUG] send_dingtalk_file_preview 被调用: file_path={file_path}, title={title}")
    
    try:
        import os
        import json
        from config import DINGTALK_OPEN_CONVERSATION_ID, DINGTALK_APP_KEY
        
        # 检查文件
        if not os.path.exists(file_path):
            return {'success': False, 'error': f'文件不存在: {file_path}'}
        
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            return {'success': False, 'error': '文件大小为0'}
        
        file_name = os.path.basename(file_path)
        
        # 1. 上传文件到存储空间
        print(f"[DEBUG] 上传文件到存储空间...")
        space_id, dentry_id = _upload_file_to_storage(file_path, file_name)
        print(f"[DEBUG] 文件上传成功: space_id={space_id}, dentry_id={dentry_id}")
        
        # 2. 获取文件预览/编辑链接
        print(f"[DEBUG] 获取文件预览链接...")
        preview_url = _get_file_preview_url(space_id, dentry_id, 'EDIT')
        print(f"[DEBUG] 预览链接: {preview_url}")
        
        # 3. 发送链接消息到群（使用 v2 API）
        token = _get_dingtalk_api_token()
        send_url = 'https://api.dingtalk.com/v1.0/robot/groupMessages/send'
        headers = {
            'x-acs-dingtalk-access-token': token,
            'Content-Type': 'application/json'
        }
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # 使用链接消息，钉钉会自动解析为文档卡片
        payload = {
            'robotCode': DINGTALK_APP_KEY,
            'openConversationId': DINGTALK_OPEN_CONVERSATION_ID,
            'msgKey': 'sampleLink',
            'msgParam': json.dumps({
                'title': f'发货通知 - {title}',
                'text': f'时间: {current_time} | 订单数: {order_count}\n点击卡片在线预览/编辑Excel',
                'picUrl': '',
                'messageUrl': preview_url
            })
        }
        
        print(f"[DEBUG] 发送链接消息...")
        resp = requests.post(send_url, headers=headers, json=payload, timeout=10)
        result = resp.json()
        
        print(f"[DEBUG] 发送结果: {result}")
        
        if 'processQueryKey' in result:
            return {
                'success': True,
                'message': '文件已上传并发送在线预览链接',
                'preview_url': preview_url,
                'space_id': space_id,
                'dentry_id': dentry_id
            }
        else:
            return {'success': False, 'error': f'发送消息失败: {result}'}
            
    except Exception as e:
        import traceback
        print(f"[ERROR] 发送文件预览失败: {e}")
        print(f"[ERROR] {traceback.format_exc()}")
        return {'success': False, 'error': str(e)}


# ============ 钉钉在线表格功能 ============
def create_dingtalk_sheet(title):
    """在钉钉知识库中创建在线表格
    
    返回: (workbook_id, sheet_url)
    注意: workbook_id 使用 dentryUuid（不是 nodeId）
    """
    workspace_id, operator_id = _get_workspace_config()
    token = _get_dingtalk_api_token()
    url = f'https://api.dingtalk.com/v1.0/doc/workspaces/{workspace_id}/docs'
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    payload = {
        'name': title,
        'docType': 'WORKBOOK',
        'operatorId': operator_id
    }

    resp = requests.post(url, json=payload, headers=headers, timeout=10)
    result = resp.json()

    if 'dentryUuid' in result:
        # dentryUuid 才是正确的 workbookId
        workbook_id = result['dentryUuid']
        node_id = result.get('nodeId')  # 用于权限设置
        sheet_url = result.get('url', '')
        print(f"[DEBUG] 创建在线表格成功: workbook_id={workbook_id}, node_id={node_id}, url={sheet_url}")
        # 如果 URL 为空，尝试构造
        if not sheet_url and node_id:
            sheet_url = f"https://docs.dingtalk.com/document/spreadsheet/spreadsheet-detail?dentryUuid={workbook_id}&docId={node_id}"
            print(f"[DEBUG] 构造表格URL: {sheet_url}")
        return workbook_id, sheet_url, node_id
    raise Exception(f"创建在线表格失败: {result}")


def set_document_permissions(node_id, member_id, member_type='DEPT', role_type='EDITOR'):
    """
    设置知识库文档权限（旧API，仅支持部门和用户）
    
    API: PUT /v1.0/doc/workspaces/{workspaceId}/docs/{nodeId}/members
    
    参数:
        node_id: 文档ID
        member_id: 部门ID或用户unionId
        member_type: 'DEPT' 部门, 'USER' 用户
        role_type: 'ONLY_VIEWER' 只读, 'VIEWER' 可查看下载, 'EDITOR' 可编辑
    
    注意: 钉钉API不支持群组(GROUP)，只能设置部门或用户权限
    """
    workspace_id, operator_id = _get_workspace_config()
    token = _get_dingtalk_api_token()
    url = f'https://api.dingtalk.com/v1.0/doc/workspaces/{workspace_id}/docs/{node_id}/members'
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    payload = {
        'operatorId': operator_id,
        'members': [{
            'memberId': member_id,
            'memberType': member_type,
            'roleType': role_type
        }]
    }
    
    print(f"[DEBUG] 设置文档权限: url={url}, payload={payload}")
    resp = requests.put(url, json=payload, headers=headers, timeout=10)
    result = resp.json()
    print(f"[DEBUG] 设置权限响应: status={resp.status_code}, result={result}")
    
    if resp.status_code == 200 or result.get('success'):
        print(f"[DEBUG] 文档权限设置成功: node_id={node_id}, member_id={member_id}, role={role_type}")
        return True
    else:
        print(f"[WARN] 文档权限设置失败: {result}")
        return False


def set_document_public_permission(dentry_uuid, corp_id, role_type='EDITOR'):
    """
    设置文档全员可见（使用v2.0 API）
    
    API: PUT /v2.0/storage/spaces/dentries/{dentryUuid}/permissions
    
    参数:
        dentry_uuid: 文件uuid（workbook_id）
        corp_id: 企业ID
        role_type: 'READER' 仅查看, 'DOWNLOADER' 查看下载, 'EDITOR' 可编辑, 'MANAGER' 管理者
    """
    from config import DINGTALK_UNION_ID
    token = _get_dingtalk_api_token()
    url = f'https://api.dingtalk.com/v2.0/storage/spaces/dentries/{dentry_uuid}/permissions'
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    payload = {
        'roleId': role_type,
        'members': [{
            'type': 'ORG',  # ORG表示企业全员
            'id': corp_id,
            'corpId': corp_id
        }]
    }
    
    print(f"[DEBUG] 设置全员权限: url={url}, payload={payload}")
    resp = requests.put(url, json=payload, headers=headers, params={'unionId': DINGTALK_UNION_ID}, timeout=10)
    result = resp.json()
    print(f"[DEBUG] 全员权限响应: status={resp.status_code}, result={result}")
    
    if resp.status_code == 200 or result.get('success'):
        print(f"[DEBUG] 全员权限设置成功: dentry_uuid={dentry_uuid}")
        return True
    else:
        print(f"[WARN] 全员权限设置失败: {result}")
        return False


def extract_template_style(template_path):
    """从Excel模板中提取标题行样式
    
    返回: {'bg_color': '#RRGGBB', 'font_color': '#RRGGBB', 'font_bold': bool}
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 获取第一个单元格（A1）的样式作为标题行样式
        cell = ws['A1']
        
        # 提取背景色
        bg_color = '#D9D9D9'  # 默认浅灰色
        if cell.fill and cell.fill.fgColor:
            color = cell.fill.fgColor
            if color.type == 'rgb' and color.rgb:
                # 转换为 #RRGGBB 格式
                bg_color = f'#{color.rgb[-6:]}'
            elif color.type == 'theme':
                # 主题色，使用默认
                bg_color = '#D9D9D9'
        
        # 提取字体颜色
        font_color = '#000000'  # 默认黑色
        if cell.font and cell.font.color:
            color = cell.font.color
            if color.type == 'rgb' and color.rgb:
                font_color = f'#{color.rgb[-6:]}'
        
        # 提取字体加粗
        font_bold = cell.font.bold if cell.font else False
        
        wb.close()
        
        print(f"[DEBUG] 从模板提取样式: bg={bg_color}, font={font_color}, bold={font_bold}")
        return {
            'bg_color': bg_color,
            'font_color': font_color,
            'font_bold': font_bold
        }
    except Exception as e:
        print(f"[WARN] 提取模板样式失败，使用默认样式: {e}")
        return None


def write_sheet_data(workbook_id, headers, rows, template_style=None):
    """向钉钉在线表格写入数据（使用正确的 ranges API）
    
    API路径: PUT /v1.0/doc/workbooks/{workbookId}/sheets/{sheetId}/ranges/{rangeAddress}
    
    参数:
        workbook_id: 表格ID
        headers: 列标题列表
        rows: 数据行列表
        template_style: 模板样式（从Excel模板提取）
    """
    _, operator_id = _get_workspace_config()
    token = _get_dingtalk_api_token()

    # 1. 获取工作表列表
    meta_url = f'https://api.dingtalk.com/v1.0/doc/workbooks/{workbook_id}/sheets'
    meta_headers = {'x-acs-dingtalk-access-token': token}
    meta_resp = requests.get(meta_url, headers=meta_headers, params={'operatorId': operator_id}, timeout=10)
    meta_result = meta_resp.json()

    sheets = meta_result.get('value', meta_result.get('sheets', []))
    if not sheets:
        raise Exception(f"获取表格工作表失败: {meta_result}")

    sheet_id = sheets[0]['id']
    col_count = len(headers)
    last_col = chr(64 + col_count) if col_count <= 26 else chr(64 + col_count // 26) + chr(65 + col_count % 26 - 1)
    
    values_headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }

    # 从模板样式中获取颜色，默认使用浅灰色
    if template_style:
        bg_color = template_style.get('bg_color', '#D9D9D9')  # 默认浅灰色
        font_color = template_style.get('font_color', '#000000')  # 默认黑色
        font_bold = template_style.get('font_bold', False)
    else:
        bg_color = '#D9D9D9'  # 浅灰色背景
        font_color = '#000000'  # 黑色字体
        font_bold = False

    print(f"[DEBUG] 使用样式: bg={bg_color}, font={font_color}, bold={font_bold}")

    # 2. 写入标题行（使用模板样式）
    header_range = f"A1:{last_col}1"
    header_url = f'https://api.dingtalk.com/v1.0/doc/workbooks/{workbook_id}/sheets/{sheet_id}/ranges/{header_range}'
    header_payload = {
        'values': [headers],
        'backgroundColors': [[bg_color] * col_count],
        'fontColors': [[font_color] * col_count],
        'fontWeights': [['bold' if font_bold else 'normal'] * col_count],
        'fontSizes': [[11] * col_count],
        'horizontalAlignments': [['center'] * col_count],
        'verticalAlignments': [['middle'] * col_count],
        'rowHeights': [25],
    }
    print(f"[DEBUG] 写入标题行: url={header_url}")
    resp = requests.put(header_url, json=header_payload, headers=values_headers, 
                       params={'operatorId': operator_id}, timeout=10)
    print(f"[DEBUG] 标题行响应: status={resp.status_code}")
    if resp.status_code != 200:
        print(f"[WARN] 写入标题行失败: {resp.text}")

    # 3. 写入数据行（大批次写入，减少API调用次数）
    if rows:
        # 使用更大的批次大小，减少网络请求次数
        batch_size = 500
        total_rows = len(rows)
        print(f"[DEBUG] 开始写入 {total_rows} 行数据，批次大小: {batch_size}")
        
        for i in range(0, total_rows, batch_size):
            batch = rows[i:i + batch_size]
            start_row = i + 2
            end_row = start_row + len(batch) - 1
            data_range = f"A{start_row}:{last_col}{end_row}"
            data_url = f'https://api.dingtalk.com/v1.0/doc/workbooks/{workbook_id}/sheets/{sheet_id}/ranges/{data_range}'
            
            data_payload = {
                'values': batch,
                'fontSizes': [[10] * col_count] * len(batch),
                'horizontalAlignments': [['center'] + ['left'] * (col_count - 1)] * len(batch),
                'verticalAlignments': [['middle'] * col_count] * len(batch),
                'rowHeights': [20] * len(batch),
            }
            resp = requests.put(data_url, json=data_payload, headers=values_headers, 
                               params={'operatorId': operator_id}, timeout=60)
            if resp.status_code != 200:
                print(f"[WARN] 写入数据行 {start_row}-{end_row} 失败: {resp.text[:200]}")
        
        print(f"[DEBUG] 数据写入完成: {total_rows} 行数据已写入表格")


def send_dingtalk_sheet(title, headers, rows, order_count=0, workbook_id=None, sheet_url=None, template_style=None):
    """
    创建钉钉在线表格并发送到群聊
    
    参数:
        title: 表格标题
        headers: 列标题列表
        rows: 数据行列表
        order_count: 订单数量（用于消息显示）
        workbook_id: 已有的表格ID（可选，传入则跳过创建步骤）
        sheet_url: 已有的表格URL（可选，传入则跳过创建步骤）
        template_style: 模板样式（从Excel提取）
    """
    try:
        from config import DINGTALK_OPEN_CONVERSATION_ID, DINGTALK_APP_KEY
        
        # 如果没有传入已有的表格，则创建新的（带重试）
        node_id = None
        if not workbook_id or not sheet_url:
            import time
            for attempt in range(3):
                try:
                    workbook_id, sheet_url, node_id = create_dingtalk_sheet(title)
                    break
                except Exception as e:
                    if attempt < 2:
                        print(f"[WARN] 创建表格失败，重试 ({attempt+1}/3): {e}")
                        time.sleep(1)
                    else:
                        raise
        
        write_sheet_data(workbook_id, headers, rows, template_style)
        
        # 权限设置改为后台异步执行，不阻塞消息发送
        def set_permission_async():
            try:
                from config import DINGTALK_CORP_ID
                if DINGTALK_CORP_ID and workbook_id:
                    result = set_document_public_permission(workbook_id, DINGTALK_CORP_ID, 'EDITOR')
                    print(f"[DEBUG] [ASYNC] 设置全员权限结果: {result}")
            except Exception as e:
                print(f"[ERROR] [ASYNC] 设置文档权限失败: {e}")
        
        import threading
        threading.Thread(target=set_permission_async, daemon=True).start()
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
        total_rows = len(rows)
        
        print(f"[DEBUG] 在线表格创建成功: {workbook_id}")
        print(f"[DEBUG] 表格URL: {sheet_url}")
        
        # 使用 v2 API 发送链接消息
        token = _get_dingtalk_token()
        send_url = 'https://api.dingtalk.com/v1.0/robot/groupMessages/send'
        headers = {
            'x-acs-dingtalk-access-token': token,
            'Content-Type': 'application/json'
        }
        
        # 消息标题和内容使用文件名
        import json
        payload = {
            'robotCode': DINGTALK_APP_KEY,
            'openConversationId': DINGTALK_OPEN_CONVERSATION_ID,
            'msgKey': 'sampleLink',
            'msgParam': json.dumps({
                'title': title,  # 文件名作为标题
                'text': f'文件名: {title}\n时间: {current_time} | 订单数: {order_count}\n点击查看在线表格，支持实时编辑',
                'picUrl': '',
                'messageUrl': sheet_url
            })
        }
        
        print(f"[DEBUG] 发送链接消息: {payload}")
        
        resp = requests.post(send_url, headers=headers, json=payload, timeout=10)
        result = resp.json()
        
        print(f"[DEBUG] 发送结果: {result}")
        
        if 'processQueryKey' in result:
            return {
                'success': True, 
                'message': '在线表格创建并发送成功', 
                'workbook_id': workbook_id, 
                'url': sheet_url
            }
        else:
            # 如果链接消息失败，尝试发送文本消息
            print(f"[WARN] 链接消息发送失败，尝试发送文本消息: {result}")
            return _send_sheet_as_text(title, sheet_url, order_count, total_rows)
            
    except Exception as e:
        import traceback
        print(f"[ERROR] 发送在线表格失败: {e}")
        print(f"[ERROR] {traceback.format_exc()}")
        return {'success': False, 'error': str(e)}




def send_excel_as_online_sheet(file_path, file_name, order_count=0, workbook_id=None, sheet_url=None):
    """
    读取 Excel 文件并发送到钉钉在线表格
    
    参数:
        file_path: Excel 文件路径
        file_name: 显示的文件名
        order_count: 订单数量
        workbook_id: 已有的表格ID（可选）
        sheet_url: 已有的表格URL（可选）
    
    返回:
        {'success': True/False, 'error': ..., 'url': ...}
    """
    try:
        import openpyxl
        
        # 读取 Excel 文件
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 获取标题行
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                break
        
        if not headers:
            headers = ['序号', '客户名称', '运单号', '发货时间', '状态']  # 默认标题
        
        # 提取标题行样式：如果没有提取到有效颜色，使用浅灰色
        template_style = None
        try:
            cell = ws['A1']
            bg_color = '#E5E5E5'  # 浅灰色默认
            font_color = '#000000'  # 黑色字体
            font_bold = False
            
            if cell.fill and cell.fill.fgColor:
                color = cell.fill.fgColor
                if color.type == 'rgb' and color.rgb:
                    rgb = color.rgb[-6:]
                    # 过滤掉透明色和黑色，使用浅灰色默认值
                    if rgb.upper() not in ('000000', 'FFFFFF', 'FFFFFF') and rgb != '000000':
                        bg_color = f'#{rgb}'
            
            if cell.font and cell.font.color:
                color = cell.font.color
                if color.type == 'rgb' and color.rgb:
                    font_color = f'#{color.rgb[-6:]}'
            
            if cell.font:
                font_bold = cell.font.bold or False
            
            template_style = {
                'bg_color': bg_color,
                'font_color': font_color,
                'font_bold': font_bold
            }
            print(f"[DEBUG] 从Excel提取样式: bg={bg_color}, font={font_color}, bold={font_bold}")
        except Exception as e:
            print(f"[WARN] 提取样式失败: {e}")
        
        # 获取数据行
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(str(cell.value))
                else:
                    row_data.append('')
            if any(row_data):  # 跳过空行
                rows.append(row_data)
        
        wb.close()
        
        print(f"[DEBUG] 读取 Excel: {len(headers)} 列, {len(rows)} 行数据")
        
        # 使用原始文件名作为标题
        title = file_name.replace('.xlsx', '').replace('.xls', '')
        
        # 调用在线表格发送
        return send_dingtalk_sheet(title, headers, rows, order_count, workbook_id, sheet_url, template_style)
        
    except Exception as e:
        import traceback
        print(f"[ERROR] 读取 Excel 发送失败: {e}")
        print(f"[ERROR] {traceback.format_exc()}")
        return {'success': False, 'error': str(e)}

def _send_sheet_as_text(title, sheet_url, order_count, total_rows):
    """备用：以文本方式发送在线表格链接"""
    try:
        from config import DINGTALK_OPEN_CONVERSATION_ID, DINGTALK_APP_KEY
        import json
        
        token = _get_dingtalk_token()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        send_url = 'https://api.dingtalk.com/v1.0/robot/groupMessages/send'
        headers = {
            'x-acs-dingtalk-access-token': token,
            'Content-Type': 'application/json'
        }
        
        payload = {
            'robotCode': DINGTALK_APP_KEY,
            'openConversationId': DINGTALK_OPEN_CONVERSATION_ID,
            'msgKey': 'sampleMarkdown',
            'msgParam': json.dumps({
                'title': f'发货通知 - {title}',
                'text': f'**发货通知**\n\n- 时间: {current_time}\n- 订单数: {order_count}\n- 数据行数: {total_rows}\n\n[点击打开在线表格]({sheet_url})\n\n> 数据已同步到钉钉在线表格，支持实时编辑'
            })
        }
        
        resp = requests.post(send_url, headers=headers, json=payload, timeout=10)
        result = resp.json()
        
        if 'processQueryKey' in result:
            return {'success': True, 'message': '在线表格链接已发送（Markdown格式）'}
        else:
            return {'success': False, 'error': f'发送失败: {result}'}
            
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============ WPS模板填充功能 ============
def fill_wps_template(template_path, mapping, data_rows, output_path):
    """
    填充WPS Excel模板（支持批量数据行）
    支持 .xlsx, .xls 格式，以及无扩展名的文件（自动检测）
    
    Args:
        template_path: 模板文件路径
        mapping: 字段映射（Excel列名 → 订单字段）
        data_rows: 数据行列表，每行是一个字典
        output_path: 输出文件路径
    """
    import tempfile
    import shutil
    
    temp_copy = None
    
    try:
        import openpyxl
        from openpyxl.styles import Alignment
        import os
        
        print(f"[DEBUG] fill_wps_template 开始执行")
        print(f"[DEBUG] 模板路径: {template_path}")
        print(f"[DEBUG] 输出路径: {output_path}")
        print(f"[DEBUG] 数据行数: {len(data_rows)}")
        print(f"[DEBUG] 字段映射: {mapping}")
        
        # 检查模板文件是否存在
        if not os.path.exists(template_path):
            error_msg = f"模板文件不存在: {template_path}"
            print(f"[ERROR] {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # 检查模板文件大小
        template_size = os.path.getsize(template_path)
        print(f"[DEBUG] 模板文件大小: {template_size} bytes")
        if template_size == 0:
            error_msg = "模板文件大小为0"
            print(f"[ERROR] {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # 检查文件扩展名
        file_ext = os.path.splitext(template_path)[1].lower()
        print(f"[DEBUG] 模板文件扩展名: '{file_ext}'")
        
        # 如果没有扩展名或是 .xlsx/.xlsm/.xltx/.xltm 之外的格式，尝试自动检测
        actual_path = template_path
        
        if file_ext not in ('.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'):
            print(f"[DEBUG] 无扩展名或不支持的扩展名，尝试检测文件格式")
            
            # 检查是否是 ZIP 格式（.xlsx 是 ZIP）
            import zipfile
            if zipfile.is_zipfile(template_path):
                print(f"[DEBUG] 检测到 ZIP 格式，可能是 .xlsx 文件")
                # 复制到临时文件并添加 .xlsx 扩展名
                temp_copy = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                temp_copy.close()
                shutil.copy(template_path, temp_copy.name)
                actual_path = temp_copy.name
                print(f"[DEBUG] 已复制到临时文件: {actual_path}")
            else:
                print(f"[DEBUG] 不是 ZIP 格式，可能是 .xls 文件")
                # 可能是 .xls 格式
                file_ext = '.xls'
        
        # 根据文件类型选择读取方式
        wb = None
        ws = None
        header_map = {}
        
        if file_ext == '.xls':
            # 旧版 .xls 格式，使用 xlrd 读取
            print(f"[DEBUG] 检测到 .xls 格式，使用 xlrd 读取")
            try:
                import xlrd
                rb = xlrd.open_workbook(actual_path, formatting_info=True)
                rs = rb.sheet_by_index(0)
                
                # 创建新的 openpyxl 工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # 复制表头（第一行）
                for col_idx in range(rs.ncols):
                    cell_value = rs.cell_value(0, col_idx)
                    ws.cell(row=1, column=col_idx + 1, value=cell_value)
                    if cell_value:
                        header_map[cell_value] = col_idx + 1
                
                print(f"[DEBUG] .xls 表头映射: {header_map}")
                
            except ImportError:
                print(f"[ERROR] 读取 .xls 文件需要 xlrd 库，请安装: pip install xlrd")
                return {'success': False, 'error': '读取 .xls 文件需要 xlrd 库，请安装: pip install xlrd'}
        else:
            # .xlsx 格式
            # 使用 read_only=True 读取模板（避免 openpyxl 的绘图/图片解析 bug）
            # 然后创建新的工作簿写入数据
            print(f"[DEBUG] 使用 openpyxl 读取模板: {actual_path}")
            try:
                # 第一步：用 read_only=True 读取表头
                wb_read = openpyxl.load_workbook(actual_path, read_only=True, data_only=True)
                ws_read = wb_read.active
                
                print(f"[DEBUG] 模板工作表行数: {ws_read.max_row}, 列数: {ws_read.max_column}")
                
                # 读取表头
                for col in range(1, ws_read.max_column + 1):
                    cell_value = ws_read.cell(row=1, column=col).value
                    if cell_value:
                        header_map[cell_value] = col
                
                print(f"[DEBUG] 表头映射: {header_map}")
                
                # 关闭只读工作簿
                wb_read.close()
                
                # 第二步：创建新的工作簿用于写入
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # 复制表头到新的工作簿
                for col_name, col_idx in header_map.items():
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                print(f"[DEBUG] 已创建新工作簿并复制表头")
                
            except Exception as e:
                print(f"[ERROR] openpyxl 读取失败: {e}")
                return {'success': False, 'error': f'无法读取模板文件: {e}'}
        
        # 检查是否有数据需要填充
        if not data_rows:
            print(f"[WARN] 没有数据行需要填充")
        
        # 从第二行开始填充数据
        start_row = 2
        filled_count = 0
        for row_idx, row_data in enumerate(data_rows):
            current_row = start_row + row_idx
            print(f"[DEBUG] 填充第 {row_idx + 1} 行数据: {row_data}")
            for excel_col_name, order_field in mapping.items():
                if excel_col_name in header_map:
                    col_idx = header_map[excel_col_name]
                    cell = ws.cell(row=current_row, column=col_idx)
                    # 获取字段值
                    if order_field == '__seq__':
                        cell.value = row_data.get(order_field, row_idx + 1)
                    else:
                        cell.value = row_data.get(order_field, '')
                    # 设置居中对齐
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    filled_count += 1
            print(f"[DEBUG] 第 {row_idx + 1} 行填充完成")
        
        print(f"[DEBUG] 总共填充了 {filled_count} 个单元格")
        
        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"[DEBUG] 创建输出目录: {output_dir}")
        
        wb.save(output_path)
        
        # 验证输出文件
        if os.path.exists(output_path):
            output_size = os.path.getsize(output_path)
            print(f"[DEBUG] 输出文件保存成功，大小: {output_size} bytes")
            if output_size == 0:
                return {'success': False, 'error': '输出文件大小为0'}
        else:
            return {'success': False, 'error': '输出文件未生成'}
        
        return {'success': True, 'path': output_path}
        
    except ImportError as e:
        print(f"[WARN] openpyxl 导入失败: {e}，使用XML方式")
        return _fill_wps_template_xml_batch(template_path, mapping, data_rows, output_path)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[ERROR] fill_wps_template 异常: {e}")
        print(f"[ERROR] 详细错误: {error_detail}")
        return {'success': False, 'error': str(e), 'detail': error_detail}
    finally:
        # 清理临时文件
        if temp_copy and os.path.exists(temp_copy.name):
            try:
                os.remove(temp_copy.name)
                print(f"[DEBUG] 已清理临时文件: {temp_copy.name}")
            except Exception as e:
                print(f"[WARN] 清理临时文件失败: {e}")


def _fill_wps_template_xml_batch(template_path, mapping, data_rows, output_path):
    """使用XML方式批量填充WPS模板（简化版）"""
    try:
        import zipfile
        with zipfile.ZipFile(template_path, 'r') as zip_in:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for item in zip_in.namelist():
                    zip_out.writestr(item, zip_in.read(item))
        return {'success': True, 'path': output_path, 'warning': 'XML批量填充未完全实现，建议安装openpyxl'}
    except Exception as e:
        return {'success': False, 'error': f"XML批量填充失败: {str(e)}"}


# 保留旧版函数用于兼容
def _fill_wps_template_single(template_path, output_path, data):
    """填充WPS Excel模板（单条数据，旧版兼容）"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value
                    for key, val in data.items():
                        placeholder = f'{{{{{key}}}}}'
                        if placeholder in value:
                            value = value.replace(placeholder, str(val))
                    cell.value = value
        wb.save(output_path)
        return {'success': True, 'path': output_path}
    except ImportError:
        return _fill_wps_template_xml(template_path, output_path, data)
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _fill_wps_template_xml(template_path, output_path, data):
    """使用XML方式填充WPS模板"""
    try:
        with zipfile.ZipFile(template_path, 'r') as zip_in:
            if 'xl/sharedStrings.xml' in zip_in.namelist():
                shared_strings_bytes = zip_in.read('xl/sharedStrings.xml')
                if shared_strings_bytes is not None:
                    shared_strings = shared_strings_bytes.decode('utf-8')
                else:
                    shared_strings = ''
                for key, val in data.items():
                    placeholder = f'{{{{{key}}}}}'
                    shared_strings = shared_strings.replace(placeholder, str(val))
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                    for item in zip_in.namelist():
                        if item == 'xl/sharedStrings.xml':
                            zip_out.writestr(item, shared_strings.encode('utf-8'))
                        else:
                            content = zip_in.read(item)
                            if content is not None:
                                zip_out.writestr(item, content)
                return {'success': True, 'path': output_path}
            else:
                return {'success': False, 'error': '模板格式不支持'}
    except Exception as e:
        return {'success': False, 'error': f"XML填充失败: {str(e)}"}


def send_dingtalk_preview(file_path, title, order_count=0):
    """
    发送文件到钉钉群（支持 v2 API 和旧版 API）
    
    优先使用 v2 API（通过 openConversationId + 机器人发送）
    如果未配置 openConversationId，回退到旧版 API（通过 chatId）
    
    Args:
        file_path: 文件路径
        title: 文件标题
        order_count: 订单数量（未使用，保持兼容）
    """
    print(f"[DEBUG] send_dingtalk_preview 被调用: file_path={file_path}, title={title}")
    
    try:
        import os
        import requests
        from config import DINGTALK_CHAT_ID, DINGTALK_OPEN_CONVERSATION_ID, DINGTALK_APP_KEY
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            error_msg = f"文件不存在: {file_path}"
            print(f"[ERROR] {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # 获取文件大小
        file_size = os.path.getsize(file_path)
        print(f"[DEBUG] 文件大小: {file_size} bytes")
        
        if file_size == 0:
            error_msg = "文件大小为0，无法上传"
            print(f"[ERROR] {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # 获取钉钉token
        print(f"[DEBUG] 正在获取钉钉token...")
        token = _get_dingtalk_token()
        print(f"[DEBUG] 获取token成功")
        
        # 1. 上传文件到钉钉获取media_id
        print(f"[DEBUG] 上传文件到钉钉...")
        upload_url = 'https://oapi.dingtalk.com/media/upload'
        
        filename = os.path.basename(file_path)
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        print(f"[DEBUG] 读取文件内容: {len(file_content)} bytes")
        
        files = {
            'media': (filename, file_content, 'application/octet-stream')
        }
        data = {
            'access_token': token,
            'type': 'file'
        }
        
        upload_resp = requests.post(upload_url, params=data, files=files, timeout=30)
        upload_result = upload_resp.json()
        
        print(f"[DEBUG] 上传结果: {upload_result}")
        
        if upload_result.get('errcode') != 0:
            error_msg = f"文件上传失败: {upload_result}"
            print(f"[ERROR] {error_msg}")
            return {'success': False, 'error': error_msg}
        
        media_id = upload_result.get('media_id')
        print(f"[DEBUG] 获取media_id成功: {media_id}")
        
        # 2. 发送文件消息到群
        # 优先使用 v2 API（openConversationId + 机器人）
        if DINGTALK_OPEN_CONVERSATION_ID:
            print(f"[DEBUG] 使用 v2 API 发送文件 (openConversationId={DINGTALK_OPEN_CONVERSATION_ID})")
            return _send_file_v2_api(token, media_id, filename)
        else:
            print(f"[DEBUG] 使用旧版 API 发送文件 (chatId={DINGTALK_CHAT_ID})")
            return _send_file_legacy_api(token, media_id)
            
    except Exception as e:
        import traceback
        error_detail = f"{str(e)}\n{traceback.format_exc()}"
        print(f"[ERROR] 钉钉发送异常: {error_detail}")
        return {'success': False, 'error': str(e), 'traceback': traceback.format_exc()}


def _send_file_v2_api(token, media_id, filename):
    """使用 v2 机器人 API 发送文件到群"""
    import requests
    import json
    from config import DINGTALK_OPEN_CONVERSATION_ID, DINGTALK_APP_KEY
    
    send_url = 'https://api.dingtalk.com/v1.0/robot/groupMessages/send'
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    
    # 获取文件扩展名作为 fileType
    file_ext = filename.rsplit('.', 1)[-1] if '.' in filename else 'xlsx'
    
    payload = {
        'robotCode': DINGTALK_APP_KEY,
        'openConversationId': DINGTALK_OPEN_CONVERSATION_ID,
        'msgKey': 'sampleFile',
        'msgParam': json.dumps({
            'mediaId': media_id,
            'fileName': filename,
            'fileType': file_ext
        })
    }
    
    print(f"[DEBUG] v2 API 请求: robotCode={DINGTALK_APP_KEY}, openConversationId={DINGTALK_OPEN_CONVERSATION_ID}")
    
    resp = requests.post(send_url, headers=headers, json=payload, timeout=10)
    result = resp.json()
    
    print(f"[DEBUG] v2 API 发送结果: {result}")
    
    if 'processQueryKey' in result:
        print(f"[DEBUG] 钉钉文件发送成功 (v2 API)")
        return {'success': True, 'message': '文件发送成功'}
    else:
        error_msg = f"v2 API发送失败: {result}"
        print(f"[ERROR] {error_msg}")
        # 回退到旧版 API
        print(f"[WARN] v2 API 失败，回退到旧版 API")
        from config import DINGTALK_CHAT_ID
        return _send_file_legacy_api(token, media_id)


def _send_file_legacy_api(token, media_id):
    """使用旧版 chat/send API 发送文件到群"""
    import requests
    from config import DINGTALK_CHAT_ID
    
    send_url = 'https://oapi.dingtalk.com/chat/send'
    params = {'access_token': token}
    
    message = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    
    payload = {
        'chatid': DINGTALK_CHAT_ID,
        'msg': message
    }
    
    resp = requests.post(send_url, params=params, json=payload, timeout=10)
    result = resp.json()
    
    print(f"[DEBUG] 旧版API发送结果: {result}")
    
    if result.get('errcode') == 0:
        print(f"[DEBUG] 钉钉文件发送成功 (旧版API)")
        return {'success': True, 'message': '文件发送成功'}
    else:
        error_msg = f"发送文件失败: {result}"
        print(f"[ERROR] {error_msg}")
        return {'success': False, 'error': error_msg}


# ============ 订单发货提醒功能 ============
def check_order_reminders(app=None):
    """
    检查并发送订单发货提醒
    
    定时任务调用此函数，检查所有未发送的提醒，
    如果预计发货时间已到或即将到达，发送桌面通知给业务员
    
    Args:
        app: Flask应用实例（用于创建应用上下文）
    
    Returns:
        dict with 'sent_count' and 'total_count'
    """
    def _do_check():
        from datetime import datetime
        
        now = datetime.now()
        print(f"[REMINDER] 检查订单发货提醒 - {now.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # 查询所有未发送的提醒，且预计时间已到或在5分钟内
        reminders = OrderReminder.query.filter(
            OrderReminder.is_sent == False,
            OrderReminder.expected_shipping_time <= now
        ).all()
        
        print(f"[REMINDER] 找到 {len(reminders)} 条待发送提醒")
        
        sent_count = 0
        for reminder in reminders:
            try:
                # 获取订单和用户信息
                order = reminder.order
                user = reminder.user
                
                if not order or not user:
                    print(f"[REMINDER] 订单或用户不存在，跳过提醒 ID:{reminder.id}")
                    continue
                
                # 构建提醒消息
                customer_info = f"{order.customer_name or '未知客户'}-{order.phone}"
                message = f"【发货提醒】订单 #{order.id} ({customer_info}) 已到预计发货时间，请及时处理！"
                
                print(f"[REMINDER] 发送提醒给 {user.name} (ID:{user.id}): {message}")
                
                # 通过Socket.IO发送通知（需要应用上下文）
                try:
                    from flask_socketio import emit
                    emit('notification', {
                        'type': 'shipping_reminder',
                        'title': '发货提醒',
                        'content': message,
                        'order_id': order.id,
                        'user_id': user.id,
                        'time': now.strftime('%Y-%m-%d %H:%M:%S')
                    }, room=f'user_{user.id}')
                    print(f"[REMINDER] Socket.IO 发送成功")
                except Exception as e:
                    print(f"[REMINDER] Socket.IO 发送失败: {e}")
                    # 降级处理：记录到日志，后续可通过其他方式通知
                
                # 更新提醒状态
                reminder.is_sent = True
                reminder.sent_time = now
                db.session.commit()
                sent_count += 1
                
                print(f"[REMINDER] 提醒 ID:{reminder.id} 已标记为已发送")
                
            except Exception as e:
                print(f"[REMINDER] 处理提醒 ID:{reminder.id} 失败: {e}")
        
        print(f"[REMINDER] 检查完成，成功发送 {sent_count} 条提醒")
        return {'sent_count': sent_count, 'total_count': len(reminders)}
    
    if app:
        with app.app_context():
            return _do_check()
    else:
        return _do_check()


# ============ 定时任务统一执行入口 ============

_SCHEDULED_TASK_FUNCS = {}


def _register_scheduled_task_func(key, func):
    """注册任务执行函数"""
    _SCHEDULED_TASK_FUNCS[key] = func


# 项目内置任务注册
_register_scheduled_task_func('update_sf_logistics', update_sf_logistics)
_register_scheduled_task_func('check_order_reminders', check_order_reminders)


def run_scheduled_task_by_key(task_key, app=None):
    """按 task_key 执行一个定时任务（供调度器 / 手动执行调用）

    Returns:
        dict: {'status': 'success'/'failed', 'message': str, 'duration': int}
    """
    import time
    start_time = time.time()

    func = _SCHEDULED_TASK_FUNCS.get(task_key)
    if func is None:
        return {'status': 'failed', 'message': f'未知任务: {task_key}', 'duration': 0}

    try:
        result = func() if task_key == 'check_order_reminders' else func(order_ids=None)

        if isinstance(result, dict):
            if 'updated' in result:
                msg = f'共 {result.get("total", 0)} 条，成功 {result.get("updated", 0)} 条，失败 {result.get("failed", 0)} 条'
            elif 'sent_count' in result:
                msg = f'共检查 {result.get("total_count", 0)} 条，发送 {result.get("sent_count", 0)} 条提醒'
            else:
                msg = str(result)
        else:
            msg = str(result)

        duration = int(time.time() - start_time)
        print(f'[定时任务] {task_key} 执行完成 - {msg} (耗时 {duration}s)')
        return {'status': 'success', 'message': msg, 'duration': duration}
    except Exception as e:
        import traceback
        traceback.print_exc()
        duration = int(time.time() - start_time)
        return {'status': 'failed', 'message': f'执行异常: {e}', 'duration': duration}


